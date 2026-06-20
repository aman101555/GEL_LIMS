"""
payment_advice.py
FastAPI router for the Walk-In "Generate Payment Advice" feature.

Mirrors the patterns already used in invoices.py / walkin.py:
- Numbering style copied from generate_delivery_note_number() / generate_invoice_no()
- Excel-fill style copied from generate_excel_invoice()
- Template download copied from download_template_from_supabase()

DB additions required (see migration.sql):
    ALTER TABLE walk_in_items ADD COLUMN IF NOT EXISTS pa_generated BOOLEAN DEFAULT FALSE;
    CREATE TABLE payment_advices ( ... )
    CREATE TABLE payment_advice_items ( ... )
"""

from fastapi import APIRouter, HTTPException
from fastapi.responses import FileResponse
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, date
import os
import re
import tempfile
import traceback

import requests
import openpyxl

from db import get_connection

# Reuse the amount-in-words helper that already exists for invoices.
from invoices import number_to_words

router = APIRouter(prefix="/payment-advice", tags=["Payment Advice"])

VAT_RATE = 0.05

TEMPLATE_URL = (
    "https://hqwgkmbjmcxpxbwccclo.supabase.co/storage/v1/object/public/"
    "templates/invoices/paymentadvice.xlsx"
)

# Template cell map (per spec: D18:D32 / I18:I32 / J18:J32 for items,
# K35 = SUM(K18:K34) for the subtotal — the sum range intentionally
# extends two rows past the last data row to match the template exactly).
FIRST_ITEM_ROW = 18
LAST_TEMPLATE_ITEM_ROW = 32      # D18:D32 / I18:I32 / J18:J32 -> 15 data rows
SUM_RANGE_END_ROW = 34           # K35 sums K18:K34 (per template, 2 rows wider than data)
SUBTOTAL_ROW = 35                # K35 = SUM(K18:K34)
VAT_ROW = 36                     # K36 = K35 * 5%
GRAND_TOTAL_ROW = 37             # K37 = K35 + K36
WORDS_ROW = 38                   # C38 = amount in words
TEMPLATE_ITEM_ROWS = LAST_TEMPLATE_ITEM_ROW - FIRST_ITEM_ROW + 1  # 15


# ─── Pydantic models ──────────────────────────────────────────────────────────

class GeneratePaymentAdviceRequest(BaseModel):
    project_id: int
    selected_item_ids: Optional[List[int]] = None
    include_all_items: bool = True


# ─── Helper: download template from Supabase (same pattern as invoices.py) ───

def _download_template() -> str:
    try:
        response = requests.get(TEMPLATE_URL, timeout=30)
        response.raise_for_status()
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            tmp.write(response.content)
            return tmp.name
    except requests.exceptions.RequestException as e:
        raise HTTPException(status_code=500, detail=f"Failed to download Payment Advice template: {e}")


# ─── Helper: generate PA number (0001-26, 0002-26, ... resets each year) ─────

def _generate_pa_number(cur) -> str:
    year_short = str(datetime.now().year)[-2:]

    cur.execute("""
        SELECT pa_no FROM payment_advices
        WHERE pa_no LIKE %s
        ORDER BY pa_id DESC
        LIMIT 1
    """, (f"%-{year_short}",))
    last = cur.fetchone()

    if last and last[0]:
        try:
            last_number = int(last[0].split('-')[0])
            next_number = last_number + 1
        except (ValueError, IndexError):
            next_number = 1
    else:
        next_number = 1

    return f"{next_number:04d}-{year_short}"


def _clean_filename(text: str) -> str:
    if not text:
        return ""
    text = re.sub(r'[\\/*?:"<>|]', '-', text)
    text = re.sub(r'\s+', '-', text)
    return text.strip('- ')


def _safe_set(ws, coord: str, value):
    """
    Write a value to a cell even if that cell is part of a merged range.

    openpyxl raises 'MergedCell object attribute value is read-only' if you
    write directly to anything other than the top-left anchor cell of a
    merge. This helper finds the merge range covering `coord` (if any),
    unmerges it, writes the value to the original top-left anchor cell,
    then re-merges the same range so formatting/layout is unaffected.
    """
    cell = ws[coord]
    # Plain, unmerged cell — just write directly.
    if cell.__class__.__name__ != "MergedCell":
        cell.value = value
        return

    for merged_range in list(ws.merged_cells.ranges):
        if coord in merged_range:
            anchor = merged_range.coord.split(":")[0]
            ws.unmerge_cells(str(merged_range))
            ws[anchor] = value
            ws.merge_cells(str(merged_range))
            return

    # Shouldn't happen (a MergedCell with no covering range), but don't
    # silently swallow it — surface a clear error instead of a cryptic one.
    raise RuntimeError(f"Cell {coord} is a MergedCell but no merge range covers it")


# ─── 1. List tests under an LP, available for Payment Advice ─────────────────

@router.get("/{project_id}/items", summary="Get walk-in tests available for Payment Advice")
def get_items_for_payment_advice(project_id: int):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT project_id, project_no, walk_in_client, project_name,
                   consultant, plot_no, walk_in_phone, client_name
            FROM projects
            WHERE project_id = %s AND is_walk_in = TRUE
        """, (project_id,))
        proj = cur.fetchone()
        if not proj:
            raise HTTPException(404, "Walk-in not found")

        if not proj[1] or proj[1] == "PENDING":
            raise HTTPException(400, "LP number has not been generated for this walk-in yet")

        cur.execute("""
            SELECT item_id, description, test_standard, unit_rate,
                   quantity, amount, net_unit, item_code, pa_generated
            FROM walk_in_items
            WHERE project_id = %s
            ORDER BY item_id
        """, (project_id,))
        rows = cur.fetchall()

        items = [
            {
                "item_id": r[0],
                "description": r[1],
                "test_standard": r[2],
                "unit_rate": float(r[3]) if r[3] is not None else 0.0,
                "quantity": r[4],
                "amount": float(r[5]) if r[5] is not None else 0.0,
                "net_unit": r[6],
                "item_code": r[7],
                "already_advised": bool(r[8]),
            }
            for r in rows
        ]

        return {
            "project_id": proj[0],
            "lp_number": proj[1],
            "client_name": proj[2] or proj[3],   # Contractor (kept under this key — see GeneratePaymentAdvice.jsx)
            "project_name": proj[3],
            "consultant": proj[4],
            "plot_no": proj[5],
            "phone": proj[6],
            "client": proj[7],                    # NEW: the actual Client/Owner name (distinct from Contractor)
            "items": items,
            "advisable_count": len([i for i in items if not i["already_advised"]]),
            "advised_count": len([i for i in items if i["already_advised"]]),
        }
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        cur.close()
        conn.close()


# ─── 2. Generate the Payment Advice Excel file ────────────────────────────────

@router.post("/generate", summary="Generate Payment Advice Excel for selected tests")
def generate_payment_advice(payload: GeneratePaymentAdviceRequest):
    conn = get_connection()
    cur = conn.cursor()

    try:
        # ----------------------------------------------------
        # 1. Validate walk-in / LP
        # ----------------------------------------------------
        cur.execute("""
            SELECT project_id, project_no, walk_in_client, project_name,
                   consultant, plot_no, walk_in_phone, client_name
            FROM projects
            WHERE project_id = %s AND is_walk_in = TRUE
        """, (payload.project_id,))
        proj = cur.fetchone()
        if not proj:
            raise HTTPException(404, "Walk-in not found")

        lp_number = proj[1]
        contractor = proj[2] or proj[3] or " - "   # walk_in_client — labelled "Contractor" in the UI
        project_name = proj[3] or contractor
        consultant = proj[4] or ""
        plot_no = proj[5] or ""
        phone = proj[6] or ""
        client_field = proj[7] or ""                 # NEW: the actual Client/Owner name → goes to C12

        if not lp_number or lp_number == "PENDING":
            raise HTTPException(400, "LP number has not been generated for this walk-in yet")

        # ----------------------------------------------------
        # 2. Resolve which items go on this Payment Advice
        # ----------------------------------------------------
        if payload.include_all_items:
            cur.execute("""
                SELECT item_id, description, test_standard, unit_rate, quantity, amount
                FROM walk_in_items
                WHERE project_id = %s AND pa_generated = FALSE
                ORDER BY item_id
            """, (payload.project_id,))
        else:
            if not payload.selected_item_ids:
                raise HTTPException(400, "No tests selected")
            cur.execute("""
                SELECT item_id, description, test_standard, unit_rate, quantity, amount
                FROM walk_in_items
                WHERE project_id = %s AND item_id = ANY(%s) AND pa_generated = FALSE
                ORDER BY item_id
            """, (payload.project_id, payload.selected_item_ids))

        rows = cur.fetchall()
        if not rows:
            raise HTTPException(400, "No advisable tests found (they may already be on a Payment Advice)")

        items = [
            {
                "item_id": r[0],
                "description": r[1] or " - ",
                "test_standard": r[2] or "",
                "unit_rate": float(r[3]) if r[3] is not None else 0.0,
                "quantity": r[4] or 0,
                "amount": float(r[5]) if r[5] is not None else 0.0,
            }
            for r in rows
        ]

        if len(items) > TEMPLATE_ITEM_ROWS:
            raise HTTPException(
                400,
                f"Template supports a maximum of {TEMPLATE_ITEM_ROWS} test rows (D18:D32). "
                f"You selected {len(items)}. Please generate multiple Payment Advices instead."
            )

        subtotal = sum(i["amount"] for i in items)
        vat = round(subtotal * VAT_RATE, 2)
        grand_total = round(subtotal + vat, 2)

        # ----------------------------------------------------
        # 3. Generate the PA number
        # ----------------------------------------------------
        pa_no = _generate_pa_number(cur)
        pa_date = date.today()

        # ----------------------------------------------------
        # 4. Fill the Excel template
        # ----------------------------------------------------
        template_path = _download_template()
        if not os.path.exists(template_path):
            raise HTTPException(404, "Payment Advice template not found in Supabase storage")

        wb = openpyxl.load_workbook(template_path, data_only=False)
        ws = wb.active

        # Header fields
        _safe_set(ws, "A5", contractor)
        _safe_set(ws, "I4", pa_no)
        _safe_set(ws, "I5", pa_date.strftime("%d-%b-%Y"))
        _safe_set(ws, "I7", lp_number)

        # Additional header fields (per template v2 spec)
        _safe_set(ws, "B7", phone)          # Tel
        _safe_set(ws, "C11", consultant)    # Consultant
        _safe_set(ws, "C12", client_field)  # Cl Name — the dedicated Client/Owner field, NOT the Contractor
        _safe_set(ws, "C13", plot_no)       # Plot No
        _safe_set(ws, "C15", project_name)  # Project Name

        # Clear existing item rows first (defensive, in case template has stray data)
        for row in range(FIRST_ITEM_ROW, LAST_TEMPLATE_ITEM_ROW + 1):
            for col in ['D', 'I', 'J', 'K']:
                _safe_set(ws, f"{col}{row}", None)

        # Fill item rows + per-row formula in K
        for index, item in enumerate(items):
            row = FIRST_ITEM_ROW + index
            _safe_set(ws, f"D{row}", item["description"])
            _safe_set(ws, f"I{row}", item["quantity"])
            _safe_set(ws, f"J{row}", item["unit_rate"])
            _safe_set(ws, f"K{row}", f'=IF(I{row}="","",(I{row}*J{row}))')

        # Fill the IF formula in every row covered by the K35 SUM range
        # (K18:K34) so unused rows correctly evaluate to "" rather than 0.
        for row in range(FIRST_ITEM_ROW, SUM_RANGE_END_ROW + 1):
            if row > FIRST_ITEM_ROW + len(items) - 1:
                _safe_set(ws, f"K{row}", f'=IF(I{row}="","",(I{row}*J{row}))')

        # Totals + amount in words
        _safe_set(ws, f"K{SUBTOTAL_ROW}", f"=SUM(K{FIRST_ITEM_ROW}:K{SUM_RANGE_END_ROW})")
        _safe_set(ws, f"K{VAT_ROW}", f"=K{SUBTOTAL_ROW}*5%")
        _safe_set(ws, f"K{GRAND_TOTAL_ROW}", f"=K{SUBTOTAL_ROW}+K{VAT_ROW}")
        _safe_set(ws, f"C{WORDS_ROW}", number_to_words(grand_total))

        # ----------------------------------------------------
        # 5. Save to server
        # ----------------------------------------------------
        output_dir = "generated_payment_advices"
        os.makedirs(output_dir, exist_ok=True)

        pa_no_hyphen = pa_no.replace('/', '-')
        clean_client = _clean_filename(contractor)
        download_filename = f"PA-{pa_no_hyphen}-{clean_client}.xlsx"
        output_path = os.path.join(output_dir, f"{pa_no_hyphen}.xlsx")
        wb.save(output_path)

        # ----------------------------------------------------
        # 6. Record the Payment Advice + mark items as advised
        # ----------------------------------------------------
        cur.execute("""
            INSERT INTO payment_advices
                (pa_no, project_id, lp_number, client_name, subtotal, vat, grand_total, pa_date)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING pa_id
        """, (pa_no, payload.project_id, lp_number, contractor, subtotal, vat, grand_total, pa_date))
        pa_id = cur.fetchone()[0]

        for item in items:
            cur.execute("""
                INSERT INTO payment_advice_items
                    (pa_id, item_id, description, test_standard, unit_rate, quantity, amount)
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (pa_id, item["item_id"], item["description"], item["test_standard"],
                  item["unit_rate"], item["quantity"], item["amount"]))

        item_ids = [i["item_id"] for i in items]
        cur.execute("""
            UPDATE walk_in_items
            SET pa_generated = TRUE
            WHERE item_id = ANY(%s)
        """, (item_ids,))

        conn.commit()

        import urllib.parse
        encoded_filename = urllib.parse.quote(download_filename)

        return FileResponse(
            output_path,
            filename=download_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}; filename=\"{download_filename}\""
            }
        )

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        print(f"ERROR in generate_payment_advice: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating Payment Advice: {str(e)}")
    finally:
        cur.close()
        conn.close()


# ─── 3. List Payment Advice history for a project ─────────────────────────────

@router.get("/{project_id}/history", summary="List Payment Advices already generated for this LP")
def get_payment_advice_history(project_id: int):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("""
            SELECT pa_id, pa_no, subtotal, vat, grand_total, pa_date, created_at
            FROM payment_advices
            WHERE project_id = %s
            ORDER BY pa_id DESC
        """, (project_id,))
        rows = cur.fetchall()
        return [
            {
                "pa_id": r[0],
                "pa_no": r[1],
                "subtotal": float(r[2]),
                "vat": float(r[3]),
                "grand_total": float(r[4]),
                "pa_date": str(r[5]) if r[5] else None,
                "created_at": str(r[6]) if r[6] else None,
            }
            for r in rows
        ]
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        cur.close()
        conn.close()