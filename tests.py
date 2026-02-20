from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from datetime import datetime
from typing import Optional, List
from db import get_connection
from psycopg2.extras import DictCursor
from openpyxl import load_workbook
from utils import resource_path

import tempfile
import os
from fastapi.responses import FileResponse
import requests  # For Supabase downloads
from io import BytesIO  # For handling template bytes

router = APIRouter(prefix="/test-requests", tags=["5. Test Requests"])


def download_test_request_template_from_supabase(template_url: str = None):
    """
    Download test request template from Supabase storage.
    
    Args:
        template_url: Optional specific URL. If not provided, uses default.
    
    Returns:
        BytesIO object containing the template
    """
    try:
        # Default template URL
        default_url = "https://hqwgkmbjmcxpxbwccclo.supabase.co/storage/v1/object/public/templates/test-requests/ST_Test_Request.xlsx"
        
        url = template_url or default_url
        print(f"DEBUG: Downloading test request template from Supabase: {url}")
        
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        
        # Return the template as BytesIO
        template_bytesio = BytesIO(response.content)
        print(f"DEBUG: Successfully downloaded test request template ({len(response.content)} bytes)")
        
        return template_bytesio
        
    except requests.exceptions.RequestException as e:
        print(f"ERROR: Failed to download test request template from {url}: {e}")
        raise ValueError(f"Failed to download test request template from Supabase: {e}")
    except Exception as e:
        print(f"ERROR in download_test_request_template_from_supabase: {e}")
        raise


class TestRequestExcelGenerator:
    def __init__(self, template_source=None, template_url=None):
        """
        Initialize test request Excel generator.
        
        template_source can be:
        1. A string (local file path)
        2. A BytesIO object (file downloaded from Supabase)
        3. None if template_url is provided
        
        template_url: Supabase URL to download template from
        """
        self.template_source = None
        
        # If template_url is provided, download from Supabase
        if template_url:
            print(f"DEBUG: Downloading test request template from URL: {template_url}")
            self.template_source = download_test_request_template_from_supabase(template_url)
        elif template_source is None:
            # Use default Supabase URL
            print(f"DEBUG: Using default test request template from Supabase")
            self.template_source = download_test_request_template_from_supabase()
        elif isinstance(template_source, str):
            # Local file path
            self.template_source = resource_path(template_source)
            if not os.path.exists(self.template_source):
                raise FileNotFoundError(f"Template not found: {self.template_source}")
            print(f"DEBUG: Using local test request template: {self.template_source}")
        else:
            # Already a BytesIO object or similar
            self.template_source = template_source
            print(f"DEBUG: Using provided test request template source (BytesIO)")

    def generate_excel(self, test_request_data, project_data, client_data, items):
        """
        Fill the Excel template with test request data
        """
        try:
            # Load the template from BytesIO or file path
            if isinstance(self.template_source, BytesIO):
                # Reset BytesIO position if needed
                self.template_source.seek(0)
                wb = load_workbook(self.template_source)
            else:
                # It's a file path
                wb = load_workbook(self.template_source)
            
            ws = wb.active
            
            # Fill in the header information
            # Based on your specifications:
            ws['C5'] = test_request_data.get('request_no', '')  # Request No
            ws['C6'] = project_data.get('lpo_no', '')  # LPO No
            ws['C7'] = test_request_data.get('project_name', '')  # Project Name
            ws['C9'] = client_data.get('address', '') if client_data else ''  # Address
            ws['C10'] = client_data.get('name', '') if client_data else ''  # Client Name
            ws['C14'] = project_data.get('location', '')  # Location
            
            ws['F5'] = test_request_data.get('created_at', '')  # Date
            ws['F6'] = datetime.now().strftime("%H:%M")  # Current Time
            ws['F7'] = test_request_data.get('project_no', '')  # Project No
            ws['F8'] = test_request_data.get('created_at', '')  # Date (again if needed)
            ws['F9'] = datetime.now().strftime("%H:%M")  # Time (again if needed)
            
            # Fill in test items starting from row 24
            start_row = 24
            current_row = start_row
            
            for index, item in enumerate(items, 1):
                # Only fill up to row 43 (20 rows max)
                if current_row > 43:
                    break
                    
                # Fill data in the specified columns
                ws[f'A{current_row}'] = index  # SI No
                ws[f'C{current_row}'] = item.get('description', '')  # Test Description
                ws[f'D{current_row}'] = item.get('test_standard', '')  # Standard
                ws[f'G{current_row}'] = item.get('quantity', 1)  # No. of Tests
                
                current_row += 1
            
            # Create a temporary file
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
                temp_file_path = tmp.name
                wb.save(temp_file_path)
            
            return temp_file_path
            
        except Exception as e:
            raise Exception(f"Error generating Excel: {str(e)}")


# ---------------------------
# Pydantic Models
# ---------------------------
class TestRequestCreate(BaseModel):
    project_id: int
    requested_by: Optional[str] = None

class TestRequestItemAdd(BaseModel):
    quotation_item_id: int   # Index (1,2,3)
    quantity: Optional[int] = None  # Make optional

class TestRequestMultiItem(BaseModel):
    items: List[TestRequestItemAdd]

class TestRequestStatusUpdate(BaseModel):
    status: str

class ItemQuantityUpdate(BaseModel):
    quantity: int


# ---------------------------
# Helper - Generate Test Request No (NEW FORMAT)
# --------------------------- 
def generate_request_no(cur):
    # Get current date components
    now = datetime.utcnow()
    date_str = now.strftime("%d%m%y")  # DDMMYY format
    
    # Count existing requests for today
    cur.execute("""
        SELECT COUNT(*)
        FROM test_requests
        WHERE DATE(created_at) = CURRENT_DATE
    """)
    count = cur.fetchone()[0] + 1
    
    # Format the sequential number
    # 01-09, then 010, 011, 012, etc.
    if count < 10:
        seq_no = f"0{count}"
    else:
        seq_no = f"0{count}"  # This gives 010, 011, etc.
    
    return f"GQ-{date_str}-{seq_no}"


# ---------------------------
# Get All Test Requests
# ---------------------------
@router.get("/")
def get_all_test_requests():
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT 
                tr.test_request_id,
                tr.request_no,
                tr.status,
                tr.requested_by,
                tr.created_at,
                p.project_id,
                p.project_no,
                p.project_name
            FROM test_requests tr
            JOIN projects p ON tr.project_id = p.project_id
            ORDER BY tr.created_at DESC
        """)

        test_requests = [
            {
                "test_request_id": row[0],
                "request_no": row[1],
                "status": row[2],
                "requested_by": row[3],
                "created_at": row[4],
                "project_id": row[5],
                "project_no": row[6],
                "project_name": row[7]
            }
            for row in cur.fetchall()
        ]

        return test_requests

    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        cur.close()
        conn.close()


# ---------------------------
# Create Test Request - UPDATED WITH DEFAULT STATUS
# ---------------------------
@router.post("/")
def create_test_request(payload: TestRequestCreate):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("SELECT project_id FROM projects WHERE project_id = %s", (payload.project_id,))
        if cur.fetchone() is None:
            raise HTTPException(404, "Project not found")

        request_no = generate_request_no(cur)

        cur.execute("""
            INSERT INTO test_requests (project_id, request_no, requested_by, status)
            VALUES (%s, %s, %s, 'PENDING_SAMPLES')
            RETURNING test_request_id
        """, (payload.project_id, request_no, payload.requested_by))

        req_id = cur.fetchone()[0]
        conn.commit()

        return {
            "message": "Test request created",
            "test_request_id": req_id,
            "request_no": request_no
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, str(e))
    finally:
        cur.close()
        conn.close()


# ---------------------------
# Update Test Request Status
# ---------------------------
@router.post("/{test_request_id}/status")
def update_test_request_status(test_request_id: int, payload: TestRequestStatusUpdate):
    conn = get_connection()
    cur = conn.cursor()

    try:
        # UPDATE THIS LINE: Add all statuses from frontend
        valid_statuses = [
            'PENDING_SAMPLES', 
            'SAMPLES_RECEIVED', 
            'TESTING_IN_PROGRESS',  
            'COMPLETED',           
            'CANCELLED', 
            'APPROVED'
        ]
        if payload.status not in valid_statuses:
            raise HTTPException(400, f"Invalid status. Must be one of: {valid_statuses}")

        cur.execute("SELECT test_request_id FROM test_requests WHERE test_request_id = %s", (test_request_id,))
        if cur.fetchone() is None:
            raise HTTPException(404, "Test request not found")

        cur.execute("""
            UPDATE test_requests 
            SET status = %s, updated_at = NOW()
            WHERE test_request_id = %s
        """, (payload.status, test_request_id))

        conn.commit()

        return {
            "message": f"Test request status updated to {payload.status}",
            "test_request_id": test_request_id,
            "new_status": payload.status
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, str(e))
    finally:
        cur.close()
        conn.close()


# ================================================================
# ADD ITEM (WITH INDEX → ACTUAL ITEM ID MAPPING) - FIXED VERSION
# ================================================================
@router.post("/{test_request_id}/items")
def add_test_item(test_request_id: int, payload: TestRequestItemAdd):
    conn = get_connection()
    cur = conn.cursor()

    try:
        # 1️⃣ Get project for this test request
        cur.execute("SELECT project_id FROM test_requests WHERE test_request_id = %s", (test_request_id,))
        row = cur.fetchone()

        if row is None:
            raise HTTPException(404, "Test request not found")

        project_id = row[0]

        # 2️⃣ Get quotation for that project
        cur.execute("SELECT quotation_id FROM projects WHERE project_id = %s", (project_id,))
        row = cur.fetchone()

        if row is None:
            raise HTTPException(404, "Project not found")

        quotation_id = row[0]

        # 3️⃣ Fetch all items for this quotation WITH THEIR QUANTITIES
        cur.execute("""
            SELECT item_id, quantity
            FROM quotation_items
            WHERE quotation_id = %s
            ORDER BY item_id
        """, (quotation_id,))

        items = cur.fetchall()  # Now each item is (item_id, quantity)

        if not items:
            raise HTTPException(404, "No quotation items found!")

        # 4️⃣ User enters INDEX — convert to REAL item_id and get quotation quantity
        user_index = payload.quotation_item_id

        if user_index < 1 or user_index > len(items):
            raise HTTPException(404, f"Invalid test index: {user_index}. Must be between 1 and {len(items)}")

        actual_item_id = items[user_index - 1][0]
        quotation_quantity = items[user_index - 1][1] or 1  # Default to 1 if null

        # 5️⃣ Use user-provided quantity OR quotation quantity
        quantity_to_use = payload.quantity if payload.quantity is not None else quotation_quantity

        # 6️⃣ Insert with the proper quantity
        cur.execute("""
            INSERT INTO test_request_items (test_request_id, quotation_item_id, quantity)
            VALUES (%s, %s, %s)
            RETURNING tri_id
        """, (test_request_id, actual_item_id, quantity_to_use))

        tri_id = cur.fetchone()[0]
        conn.commit()

        return {
            "message": "Item added",
            "tri_id": tri_id,
            "actual_item_id": actual_item_id,
            "quantity_used": quantity_to_use,
            "source": "user_input" if payload.quantity is not None else "quotation"
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"Error adding item: {str(e)}")
    finally:
        cur.close()
        conn.close()


# ================================================================
# ADD MULTIPLE ITEMS (ALSO INDEX-BASED) - FIXED VERSION
# ================================================================
@router.post("/{test_request_id}/items/bulk")
def add_multiple_items(test_request_id: int, payload: TestRequestMultiItem):
    conn = get_connection()
    cur = conn.cursor()

    try:
        # 1️⃣ Get project
        cur.execute("SELECT project_id FROM test_requests WHERE test_request_id = %s", (test_request_id,))
        row = cur.fetchone()

        if row is None:
            raise HTTPException(404, "Test request not found")

        project_id = row[0]

        # 2️⃣ Get quotation
        cur.execute("SELECT quotation_id FROM projects WHERE project_id = %s", (project_id,))
        row = cur.fetchone()

        if row is None:
            raise HTTPException(404, "Project not found")

        quotation_id = row[0]

        # 3️⃣ Get list of available item_ids WITH THEIR QUANTITIES
        cur.execute("""
            SELECT item_id, quantity
            FROM quotation_items
            WHERE quotation_id = %s
            ORDER BY item_id
        """, (quotation_id,))

        items = cur.fetchall()  # Now contains (item_id, quantity)

        if not items:
            raise HTTPException(404, "No quotation items found")

        added_items = []
        quantity_details = []

        # 4️⃣ Process each item in bulk input
        for entry in payload.items:
            if entry.quotation_item_id < 1 or entry.quotation_item_id > len(items):
                raise HTTPException(404, f"Invalid test index: {entry.quotation_item_id}")

            actual_item_id = items[entry.quotation_item_id - 1][0]
            quotation_quantity = items[entry.quotation_item_id - 1][1] or 1
            
            # Use user quantity or quotation quantity
            quantity_to_use = entry.quantity if entry.quantity is not None else quotation_quantity

            cur.execute("""
                INSERT INTO test_request_items (test_request_id, quotation_item_id, quantity)
                VALUES (%s, %s, %s)
                RETURNING tri_id
            """, (test_request_id, actual_item_id, quantity_to_use))

            tri_id = cur.fetchone()[0]
            added_items.append(tri_id)
            
            quantity_details.append({
                "tri_id": tri_id,
                "item_id": actual_item_id,
                "quantity_used": quantity_to_use,
                "source": "user_input" if entry.quantity is not None else "quotation"
            })

        conn.commit()

        return {
            "message": "Items added",
            "added_item_ids": added_items,
            "quantity_details": quantity_details
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"Error adding items: {str(e)}")
    finally:
        cur.close()
        conn.close()


# ================================================================
# COPY ALL ITEMS FROM QUOTATION WITH QUANTITIES
# ================================================================
@router.post("/{test_request_id}/items/copy-all")
def copy_all_items_from_quotation(test_request_id: int):
    """Copy all items from the project's quotation with their original quantities"""
    conn = get_connection()
    cur = conn.cursor()

    try:
        # 1️⃣ Get project
        cur.execute("SELECT project_id FROM test_requests WHERE test_request_id = %s", (test_request_id,))
        row = cur.fetchone()

        if row is None:
            raise HTTPException(404, "Test request not found")

        project_id = row[0]

        # 2️⃣ Get quotation
        cur.execute("SELECT quotation_id FROM projects WHERE project_id = %s", (project_id,))
        row = cur.fetchone()

        if row is None:
            raise HTTPException(404, "Project not found")

        quotation_id = row[0]

        # 3️⃣ Get all items from quotation with quantities
        cur.execute("""
            SELECT item_id, description, quantity, unit_rate
            FROM quotation_items
            WHERE quotation_id = %s
            ORDER BY item_id
        """, (quotation_id,))

        quotation_items = cur.fetchall()

        if not quotation_items:
            raise HTTPException(404, "No items in quotation to copy")

        # 4️⃣ Check if test request already has items
        cur.execute("SELECT COUNT(*) FROM test_request_items WHERE test_request_id = %s", (test_request_id,))
        existing_count = cur.fetchone()[0]
        
        if existing_count > 0:
            raise HTTPException(400, "Test request already has items. Please use add items individually.")

        # 5️⃣ Insert all items with their original quantities
        added_items = []
        for item in quotation_items:
            item_id, description, quantity, unit_rate = item
            
            cur.execute("""
                INSERT INTO test_request_items (test_request_id, quotation_item_id, quantity)
                VALUES (%s, %s, %s)
                RETURNING tri_id
            """, (test_request_id, item_id, quantity or 1))

            tri_id = cur.fetchone()[0]
            added_items.append({
                "tri_id": tri_id,
                "item_id": item_id,
                "description": description,
                "quantity": quantity or 1,
                "unit_rate": float(unit_rate) if unit_rate else 0.0
            })

        conn.commit()

        return {
            "message": f"Copied {len(added_items)} items from quotation with original quantities",
            "test_request_id": test_request_id,
            "items_copied": added_items,
            "total_items": len(added_items)
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"Error copying items: {str(e)}")
    finally:
        cur.close()
        conn.close()


# ---------------------------
# Get Test Request Details
# ---------------------------
@router.get("/{test_request_id}")
def get_test_request(test_request_id: int):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT tr.test_request_id, tr.request_no, tr.status, tr.project_id,
                   tr.requested_by, tr.created_at,
                   p.project_no, p.project_name
            FROM test_requests tr
            JOIN projects p ON tr.project_id = p.project_id
            WHERE tr.test_request_id = %s
        """, (test_request_id,))

        header = cur.fetchone()
        if not header:
            raise HTTPException(404, "Test request not found")

        # Items
        cur.execute("""
            SELECT tri.tri_id, tri.quantity,
                   qi.description, qi.test_standard, qi.unit_rate, qi.item_id
            FROM test_request_items tri
            JOIN quotation_items qi ON tri.quotation_item_id = qi.item_id
            WHERE tri.test_request_id = %s
        """, (test_request_id,))

        items = []
        for r in cur.fetchall():
            unit_rate = float(r[4]) if r[4] is not None else 0.0
            quantity = r[1] if r[1] is not None else 0
            
            items.append({
                "tri_id": r[0],
                "quantity": quantity,
                "description": r[2],
                "test_standard": r[3],
                "unit_rate": unit_rate,
                "amount": unit_rate * quantity,
                "item_id": r[5]  # Include actual item_id for reference
            })

        return {
            "test_request_id": header[0],
            "request_no": header[1],
            "status": header[2],
            "project_id": header[3],
            "requested_by": header[4],
            "created_at": header[5],
            "project_no": header[6],
            "project_name": header[7],
            "items": items
        }

    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        cur.close()
        conn.close()


@router.patch("/{test_request_id}")
def update_test_request(test_request_id: int, updated_data: dict):
    conn = get_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute("SELECT test_request_id FROM test_requests WHERE test_request_id = %s", (test_request_id,))
        if not cursor.fetchone():
            raise HTTPException(404, "Test Request not found")
        
        # ADD THIS: Handle status updates
        if "status" in updated_data:
            # Validate status
            valid_statuses = [
                'PENDING_SAMPLES', 
                'SAMPLES_RECEIVED', 
                'TESTING_IN_PROGRESS',  
                'COMPLETED',           
                'CANCELLED', 
                'APPROVED'
            ]
            if updated_data["status"] not in valid_statuses:
                raise HTTPException(400, f"Invalid status. Must be one of: {valid_statuses}")
            
            cursor.execute(
                "UPDATE test_requests SET status = %s WHERE test_request_id = %s",
                (updated_data["status"], test_request_id)
            )
        
        # Keep existing code for requested_by
        if "requested_by" in updated_data:
            cursor.execute(
                "UPDATE test_requests SET requested_by = %s WHERE test_request_id = %s",
                (updated_data["requested_by"], test_request_id)
            )
        
        if "items" in updated_data:
            for item in updated_data["items"]:
                cursor.execute(
                    "UPDATE test_request_items SET quantity = %s WHERE tri_id = %s AND test_request_id = %s",
                    (item["quantity"], item["tri_id"], test_request_id)
                )
        
        conn.commit()
        return {"status": "success", "message": "Test Request updated"}
        
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, str(e))
    finally:
        cursor.close()
        conn.close()


# ---------------------------
# Get Allowed Items (Index-Based) - ENHANCED VERSION
# ---------------------------
@router.get("/{test_request_id}/items/available")
def get_available_items(test_request_id: int):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("SELECT project_id FROM test_requests WHERE test_request_id = %s", (test_request_id,))
        row = cur.fetchone()

        if row is None:
            raise HTTPException(404, "Test request not found")

        project_id = row[0]

        cur.execute("SELECT quotation_id FROM projects WHERE project_id = %s", (project_id,))
        qrow = cur.fetchone()

        if qrow is None:
            raise HTTPException(404, "Project not found")

        quotation_id = qrow[0]

        # Get ALL quotation items
        cur.execute("""
            SELECT qi.item_id, qi.description, qi.test_standard, qi.quantity, qi.unit_rate, qi.amount
            FROM quotation_items qi
            WHERE qi.quotation_id = %s
            ORDER BY qi.item_id
        """, (quotation_id,))

        all_items = cur.fetchall()
        
        # Get already used item_ids from test_request_items for this project
        cur.execute("""
            SELECT DISTINCT tri.quotation_item_id
            FROM test_request_items tri
            JOIN test_requests tr ON tri.test_request_id = tr.test_request_id
            WHERE tr.project_id = %s
        """, (project_id,))
        
        used_item_ids = [row[0] for row in cur.fetchall()]
        
        # Filter out already used items
        available_items = []
        index = 1
        for item in all_items:
            item_id = item[0]
            if item_id not in used_item_ids:
                available_items.append({
                    "index": index,
                    "item_id": item_id,
                    "description": item[1],
                    "test_standard": item[2],
                    "quotation_quantity": item[3] or 1,
                    "unit_rate": float(item[4]) if item[4] else 0.0,
                    "amount": float(item[5]) if item[5] else 0.0,
                    "is_used": False
                })
                index += 1
            else:
                # You could also include used items but mark them as used
                available_items.append({
                    "index": index,
                    "item_id": item_id,
                    "description": item[1],
                    "test_standard": item[2],
                    "quotation_quantity": item[3] or 1,
                    "unit_rate": float(item[4]) if item[4] else 0.0,
                    "amount": float(item[5]) if item[5] else 0.0,
                    "is_used": True  # Mark as already used
                })
                index += 1

        return available_items

    except Exception as e:
        raise HTTPException(500, f"Error loading available items: {str(e)}")
    finally:
        cur.close()
        conn.close()


# ================================================================
# GET TEST REPORTS FOR PROJECT (For Invoice Generation)
# ================================================================
@router.get("/project/{project_id}")
def get_test_reports_for_project(project_id: int):
    """Get all test reports for a specific project - used for invoice generation"""
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("SELECT project_id FROM projects WHERE project_id = %s", (project_id,))
        if not cur.fetchone():
            raise HTTPException(404, "Project not found")
        
        cur.execute("""
            SELECT 
                tr.test_request_id,
                tr.request_no,
                tr.status,
                tr.created_at,
                tr.requested_by,
                COUNT(DISTINCT s.sample_id) as sample_count,
                COUNT(DISTINCT tri.tri_id) as test_count,
                COALESCE(SUM(qi.unit_rate * tri.quantity), 0) as total_amount
            FROM test_requests tr
            LEFT JOIN test_request_items tri ON tr.test_request_id = tri.test_request_id
            LEFT JOIN quotation_items qi ON tri.quotation_item_id = qi.item_id
            LEFT JOIN samples s ON tr.test_request_id = s.request_id
            WHERE tr.project_id = %s
            GROUP BY tr.test_request_id, tr.request_no, tr.status, tr.created_at, tr.requested_by
            ORDER BY tr.created_at DESC
        """, (project_id,))
        
        test_requests = []
        for row in cur.fetchall():
            test_requests.append({
                "test_request_id": row[0],
                "request_no": row[1],
                "status": row[2],
                "created_at": str(row[3]),
                "requested_by": row[4],
                "sample_count": row[5],
                "test_count": row[6],
                "total_amount": float(row[7]) if row[7] else 0.0
            })
        
        invoice_items = []
        if test_requests:
            cur.execute("""
                SELECT 
                    tr.request_no,
                    qi.description,
                    qi.test_standard,
                    qi.unit_rate,
                    tri.quantity,
                    (qi.unit_rate * tri.quantity) as amount,
                    tr.created_at
                FROM test_request_items tri
                JOIN quotation_items qi ON tri.quotation_item_id = qi.item_id
                JOIN test_requests tr ON tri.test_request_id = tr.test_request_id
                WHERE tr.project_id = %s
                ORDER BY tr.request_no, qi.description
            """, (project_id,))
            
            for idx, row in enumerate(cur.fetchall(), 1):
                invoice_items.append({
                    "index": idx,
                    "report_no": row[0],
                    "description": row[1],
                    "test_standard": row[2],
                    "unit_rate": float(row[3]) if row[3] else 0.0,
                    "quantity": row[4],
                    "amount": float(row[5]) if row[5] else 0.0,
                    "report_date": str(row[6])
                })
        
        return {
            "project_id": project_id,
            "test_requests": test_requests,
            "invoice_items": invoice_items,
            "total_test_requests": len(test_requests),
            "total_tests": sum(tr["test_count"] for tr in test_requests),
            "total_amount": sum(tr["total_amount"] for tr in test_requests)
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to fetch test reports: {str(e)}")
    finally:
        cur.close()
        conn.close()


@router.get("/{test_request_id}/download-doc")
def download_test_request_doc(test_request_id: int):
    """Download test request as Excel document (replaces Word)"""
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        # Fetch test request details
        cur.execute("""
            SELECT tr.test_request_id, tr.request_no, tr.status, tr.project_id,
                   tr.requested_by, tr.created_at,
                   p.project_no, p.project_name, p.location, p.lpo_no,
                   p.client_id, p.lpo_date
            FROM test_requests tr
            JOIN projects p ON tr.project_id = p.project_id
            WHERE tr.test_request_id = %s
        """, (test_request_id,))
        
        header = cur.fetchone()
        if not header:
            raise HTTPException(404, "Test request not found")
        
        # Fetch client details
        client_id = header[10]  # client_id from projects
        client_data = {}
        if client_id:
            cur.execute("""
                SELECT name, contact_person, email, phone, address
                FROM clients
                WHERE client_id = %s
            """, (client_id,))
            client_row = cur.fetchone()
            if client_row:
                client_data = {
                    "name": client_row[0],
                    "contact_person": client_row[1],
                    "email": client_row[2],
                    "phone": client_row[3],
                    "address": client_row[4]
                }
        
        # Fetch items
        cur.execute("""
            SELECT tri.tri_id, tri.quantity,
                   qi.description, qi.test_standard, qi.unit_rate, qi.item_id
            FROM test_request_items tri
            JOIN quotation_items qi ON tri.quotation_item_id = qi.item_id
            WHERE tri.test_request_id = %s
            ORDER BY tri.tri_id
        """, (test_request_id,))
        
        items = []
        for r in cur.fetchall():
            items.append({
                "tri_id": r[0],
                "quantity": r[1] if r[1] is not None else 1,
                "description": r[2],
                "test_standard": r[3],
                "unit_rate": float(r[4]) if r[4] else 0.0,
                "item_id": r[5]
            })
        
        # Prepare data for Excel
        test_request_data = {
            "test_request_id": header[0],
            "request_no": header[1],
            "status": header[2],
            "project_id": header[3],
            "requested_by": header[4],
            "created_at": header[5].strftime("%d-%m-%Y") if header[5] else datetime.now().strftime("%d-%m-%Y"),
            "project_no": header[6],
            "project_name": header[7]
        }
        
        project_data = {
            "project_no": header[6],
            "project_name": header[7],
            "location": header[8],
            "lpo_no": header[9],
            "client_id": header[10],
            "lpo_date": header[11].strftime("%d-%m-%Y") if header[11] else ""
        }
        
        # Generate Excel file using Supabase template
        # The TestRequestExcelGenerator will download the template from Supabase by default
        excel_generator = TestRequestExcelGenerator()
        temp_file = excel_generator.generate_excel(test_request_data, project_data, client_data, items)
        
        # Create filename with request number
        filename = f"Test_Request_{header[1]}.xlsx"
        
        # Return file response
        return FileResponse(
            path=temp_file,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error generating Excel document: {str(e)}")
    finally:
        cur.close()
        conn.close()


@router.get("/projects/remaining-tests")
def get_projects_with_remaining_tests():
    """Get all projects with count of remaining unused tests"""
    conn = get_connection()
    cur = conn.cursor()

    try:
        # Get all active projects
        cur.execute("""
            SELECT p.project_id, p.project_no, p.project_name, p.status,
                   q.quotation_id, COUNT(qi.item_id) as total_tests
            FROM projects p
            JOIN quotations q ON p.quotation_id = q.quotation_id
            JOIN quotation_items qi ON q.quotation_id = qi.quotation_id
            WHERE p.status = 'ACTIVE'
            GROUP BY p.project_id, p.project_no, p.project_name, p.status, q.quotation_id
            ORDER BY p.project_id DESC
        """)
        
        projects = cur.fetchall()
        
        result = []
        for project in projects:
            project_id = project[0]
            total_tests = project[5]
            
            # Count used tests for this project
            cur.execute("""
                SELECT COUNT(DISTINCT tri.quotation_item_id)
                FROM test_request_items tri
                JOIN test_requests tr ON tri.test_request_id = tr.test_request_id
                WHERE tr.project_id = %s
            """, (project_id,))
            
            used_tests = cur.fetchone()[0]
            remaining_tests = total_tests - used_tests
            
            result.append({
                "project_id": project_id,
                "project_no": project[1],
                "project_name": project[2],
                "status": project[3],
                "total_tests": total_tests,
                "used_tests": used_tests,
                "remaining_tests": remaining_tests,
                "has_remaining_tests": remaining_tests > 0
            })
        
        return result

    except Exception as e:
        raise HTTPException(500, f"Error fetching projects with remaining tests: {str(e)}")
    finally:
        cur.close()
        conn.close()


# Add to tests.py after the existing endpoints
@router.put("/{test_request_id}/items/{item_index}/quantity")
def update_test_item_quantity(test_request_id: int, item_index: int, payload: ItemQuantityUpdate):
    """Update quantity of a test request item AND sync to quotation"""
    conn = get_connection()
    cur = conn.cursor()

    try:
        # 1. Get test request item details including quotation_item_id
        cur.execute("""
            SELECT tri.tri_id, tri.quotation_item_id, tri.quantity
            FROM test_request_items tri
            WHERE tri.test_request_id = %s
            ORDER BY tri.tri_id
            OFFSET %s LIMIT 1
        """, (test_request_id, item_index))
        
        item = cur.fetchone()
        
        if not item:
            raise HTTPException(404, "Test request item not found")
        
        tri_id, quotation_item_id, old_quantity = item
        
        # 2. Validate test request status allows changes
        cur.execute("""
            SELECT status FROM test_requests 
            WHERE test_request_id = %s
        """, (test_request_id,))
        
        test_request = cur.fetchone()
        if not test_request:
            raise HTTPException(404, "Test request not found")
        
        # Only allow updates if status is PENDING_SAMPLES or SAMPLES_RECEIVED
        if test_request[0] not in ['PENDING_SAMPLES', 'SAMPLES_RECEIVED']:
            raise HTTPException(400, f"Cannot modify items in {test_request[0]} status")
        
        if payload.quantity <= 0:
            raise HTTPException(400, "Quantity must be greater than zero")
        
        # 3. Update test request item quantity
        cur.execute("""
            UPDATE test_request_items
            SET quantity = %s
            WHERE tri_id = %s
            RETURNING tri_id, quantity
        """, (payload.quantity, tri_id))
        
        updated = cur.fetchone()
        
        # 4. Sync to quotation item (if needed)
        # Get the quotation_id from the project
        cur.execute("""
            SELECT q.quotation_id
            FROM test_requests tr
            JOIN projects p ON tr.project_id = p.project_id
            JOIN quotations q ON p.quotation_id = q.quotation_id
            WHERE tr.test_request_id = %s
        """, (test_request_id,))
        
        quotation_result = cur.fetchone()
        if quotation_result:
            quotation_id = quotation_result[0]
            
            # Update quotation item quantity
            cur.execute("""
                UPDATE quotation_items
                SET quantity = %s
                WHERE item_id = %s AND quotation_id = %s
                RETURNING amount
            """, (payload.quantity, quotation_item_id, quotation_id))
            
            # Recalculate quotation totals
            cur.execute("""
                UPDATE quotations
                SET total_amount = sub.total,
                    vat = sub.total * 0.05,
                    grand_total = sub.total * 1.05
                FROM (
                    SELECT SUM(unit_rate * quantity) as total
                    FROM quotation_items
                    WHERE quotation_id = %s
                ) sub
                WHERE quotation_id = %s
                RETURNING total_amount, vat, grand_total
            """, (quotation_id, quotation_id))
            
            totals = cur.fetchone()
        
        conn.commit()
        
        return {
            "message": "Quantity updated and synced to quotation",
            "tri_id": updated[0],
            "new_quantity": updated[1],
            "quotation_updated": bool(quotation_result),
            "totals": {
                "total_amount": float(totals[0]) if quotation_result else 0,
                "vat": float(totals[1]) if quotation_result else 0,
                "grand_total": float(totals[2]) if quotation_result else 0,
            } if quotation_result else None
        }
        
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"Error updating quantity: {str(e)}")
    finally:
        cur.close()
        conn.close()