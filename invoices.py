from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Optional, List, Literal
from datetime import date, datetime
from db import get_connection
from decimal import Decimal
from fastapi.responses import HTMLResponse
import traceback
from utils import resource_path  # ADD THIS LINE
import requests
import tempfile


import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from fastapi.responses import FileResponse
import openpyxl
from datetime import datetime
import os

router = APIRouter(prefix="/invoices", tags=["6. Invoices"])

# ----------------------------
# Pydantic Models
# ----------------------------
class InvoiceItemCreate(BaseModel):
    quotation_item_id: int
    quantity: int = 1
    sample_id: Optional[int] = None

class InvoiceCreate(BaseModel):
    project_id: int
    invoice_type: Literal['CASH', 'CREDIT', 'PROFORMA', 'TAX']
    payment_method: Optional[Literal['CASH', 'CREDIT']] = 'CASH'  # NEW
    invoice_date: Optional[date] = None
    client_reference: Optional[str] = None
    lpo_reference: Optional[str] = None
    lpo_date: Optional[date] = None
    payment_terms: Optional[str] = "30 days"
    services_description: Optional[str] = None
    remarks: Optional[str] = None
    payment_status: Optional[str] = "UNPAID"


class InvoiceOut(BaseModel):
    invoice_id: int
    invoice_no: str
    project_id: int
    invoice_type: str
    payment_method: str  # NEW
    invoice_date: date
    client_reference: Optional[str]
    lpo_reference: Optional[str]
    lpo_date: Optional[date]
    payment_terms: Optional[str]
    subtotal: float
    vat: float
    total: float
    amount_in_words: str
    services_description: Optional[str]
    remarks: Optional[str]
    payment_status: str
    paid_date: Optional[date]
    items: List[dict]
    project_details: dict


# =====================================================
# NEW: Models for Invoice Reports
# =====================================================
class InvoiceReportSelection(BaseModel):
    project_id: int
    selected_report_ids: Optional[List[int]] = None
    include_all_reports: bool = True
    invoice_type: Literal['PROFORMA', 'TAX']  # New field

# ----------------------------
# Utility Functions - FIXED
# ----------------------------
# ----------------------------
# Utility Functions - FIXED
# ----------------------------
def number_to_words(num: float) -> str:
    """Convert number to words for amount in words field"""
    # First, separate whole dirhams and fils
    whole_part = int(num)
    decimal_part = round((num - whole_part) * 100)
    
    def convert_less_than_thousand(n):
        ones = ["", "One", "Two", "Three", "Four", "Five", "Six", "Seven", "Eight", "Nine"]
        teens = ["Ten", "Eleven", "Twelve", "Thirteen", "Fourteen", "Fifteen", "Sixteen", 
                 "Seventeen", "Eighteen", "Nineteen"]
        tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]
        
        if n == 0:
            return ""
        elif n < 10:
            return ones[n]
        elif n < 20:
            return teens[n - 10]
        elif n < 100:
            return tens[n // 10] + (" " + ones[n % 10] if n % 10 != 0 else "")
        else:
            return ones[n // 100] + " Hundred" + (" " + convert_less_than_thousand(n % 100) if n % 100 != 0 else "")
    
    if whole_part == 0:
        words = "Zero"
    else:
        # Handle millions
        millions = whole_part // 1000000
        remainder = whole_part % 1000000
        
        # Handle thousands
        thousands = remainder // 1000
        remainder = remainder % 1000
        
        words_parts = []
        
        if millions > 0:
            words_parts.append(convert_less_than_thousand(millions) + " Million")
        
        if thousands > 0:
            words_parts.append(convert_less_than_thousand(thousands) + " Thousand")
        
        if remainder > 0:
            words_parts.append(convert_less_than_thousand(remainder))
        
        words = " and ".join(words_parts)
        
        # Capitalize first letter
        words = words.strip()
    
    # Add decimal part if exists
    if decimal_part > 0:
        result = f"{words} Dirhams and {convert_less_than_thousand(decimal_part)} Fils Only"
    else:
        result = f"{words} Dirhams Only"
    
    return result

def generate_invoice_no(cur, invoice_type: str) -> str:
    """Generate invoice number with different systems for PROFORMA vs other invoices"""
    # Get the last 2 digits of current year
    year_short = str(datetime.now().year)[-2:]
    
    print(f"DEBUG generate_invoice_no: invoice_type='{invoice_type}', year_short='{year_short}'")
    
    # SPECIAL CASE FOR PROFORMA INVOICES - Reset each year starting from 001
    if invoice_type.upper() == 'PROFORMA':
        print("DEBUG: Generating PROFORMA invoice number")
        
        # Get the last PROFORMA invoice number for this year
        cur.execute("""
            SELECT invoice_no 
            FROM invoices 
            WHERE invoice_type = 'PROFORMA'
            AND invoice_no LIKE %s
            ORDER BY invoice_id DESC 
            LIMIT 1
        """, (f'%/{year_short}',))
        
        last_proforma = cur.fetchone()
        print(f"DEBUG: Last PROFORMA invoice found: {last_proforma}")
        
        if last_proforma and last_proforma[0]:
            try:
                # Get the number part
                last_number_str = last_proforma[0].split('/')[0]
                last_number = int(last_number_str)
                
                # If it's a new format invoice (001, 002, etc.), increment normally
                if last_number <= 999:
                    next_number = last_number + 1
                    invoice_no = f"{next_number:03d}/{year_short}"
                    print(f"DEBUG: Incremented new format PROFORMA: {last_number} -> {next_number}")
                else:
                    # Old format invoice, start new format from 001
                    next_number = 1
                    invoice_no = f"{next_number:03d}/{year_short}"
                    print(f"DEBUG: Old format found, starting new format from 001")
            except (ValueError, IndexError):
                # If parsing fails, start from 001
                next_number = 1
                invoice_no = f"{next_number:03d}/{year_short}"
                print(f"DEBUG: Parse failed, starting from 001")
        else:
            # No PROFORMA invoice for this year yet, start from 001
            next_number = 1
            invoice_no = f"{next_number:03d}/{year_short}"
            print(f"DEBUG: No existing PROFORMA, starting from 001")
        
        print(f"DEBUG: Generated PROFORMA invoice_no: {invoice_no}")
        return invoice_no
    
    # ORIGINAL LOGIC FOR OTHER INVOICE TYPES (CASH, CREDIT, TAX)
    print(f"DEBUG: Generating non-PROFORMA invoice number for type: {invoice_type}")
    
    # Get the max invoice number overall for non-PROFORMA invoices
    cur.execute("""
        SELECT invoice_no 
        FROM invoices 
        WHERE invoice_type != 'PROFORMA'
        ORDER BY invoice_id DESC 
        LIMIT 1
    """)
    
    last_invoice = cur.fetchone()
    print(f"DEBUG: Last non-PROFORMA invoice found: {last_invoice}")
    
    if last_invoice and last_invoice[0]:
        # Parse existing invoice number
        try:
            last_number_str = last_invoice[0].split('/')[0]
            last_number = int(last_number_str)
            next_number = last_number + 1
            print(f"DEBUG: Parsed last non-PROFORMA number: {last_number}, next: {next_number}")
        except (ValueError, IndexError) as e:
            print(f"DEBUG: Failed to parse non-PROFORMA number '{last_invoice[0]}': {e}")
            # If parsing fails, start from 36001
            next_number = 36001
    else:
        # First invoice ever (non-PROFORMA)
        print("DEBUG: First non-PROFORMA invoice ever")
        next_number = 36001
    
    invoice_no = f"{next_number}/{year_short}"
    print(f"DEBUG: Generated non-PROFORMA invoice_no: {invoice_no}")
    return invoice_no
def ensure_delivery_note_reports_table(cur):
    """Ensure the delivery_note_reports junction table exists"""
    cur.execute("""
        CREATE TABLE IF NOT EXISTS delivery_note_reports (
            id SERIAL PRIMARY KEY,
            delivery_note_no VARCHAR(50) NOT NULL,
            report_no VARCHAR(50) NOT NULL,  -- Changed from VARCHAR(100) to VARCHAR(50)
            UNIQUE(delivery_note_no, report_no)
        )
    """)
    # Note: removed included_at column since your SQL table doesn't have it


# ----------------------------
# Assignment Logic - SAME AS WORKSHEET GENERATION
# ----------------------------

def get_assigned_test_for_sample(sample_id: int, cur):
    """Get the assigned test for a sample - SAME LOGIC AS WORKSHEET GENERATION"""
    # Get sample's test request
    cur.execute("SELECT request_id FROM samples WHERE sample_id = %s", (sample_id,))
    sample_data = cur.fetchone()
    if not sample_data:
        return None
    
    request_id = sample_data[0]
    
    # Get all tests for this request
    cur.execute("""
        SELECT qi.item_id, qi.item_code, qi.description, qi.test_standard, qi.unit_rate,
               tri.quantity, tri.tri_id,
               ROW_NUMBER() OVER (ORDER BY tri.tri_id) as test_index
        FROM test_request_items tri
        JOIN quotation_items qi ON tri.quotation_item_id = qi.item_id
        WHERE tri.test_request_id = %s
        ORDER BY tri.tri_id
    """, (request_id,))
    
    test_items = cur.fetchall()
    
    if not test_items:
        return None
    
    # Get all samples for this request to find position
    cur.execute("SELECT sample_id FROM samples WHERE request_id = %s ORDER BY sample_id", (request_id,))
    all_samples = [row[0] for row in cur.fetchall()]
    
    # ✅ SAME LOGIC AS WORKSHEET: Calculate assigned test
    sample_position = all_samples.index(sample_id)
    assigned_test_index = sample_position % len(test_items)
    
    return test_items[assigned_test_index]


def get_sample_ids_for_reports(report_ids: List[int], cur) -> List[int]:
    """
    Resolve a list of report_id values to the underlying sample_id values
    they cover. A report always has a primary sample_id, and may additionally
    cover other samples via the covers_samples array (stored as sample_no text).
    """
    if not report_ids:
        return []

    cur.execute("""
        SELECT report_id, sample_id, covers_samples
        FROM reports
        WHERE report_id = ANY(%s)
    """, (report_ids,))
    rows = cur.fetchall()

    sample_id_set = set()
    covers_samples_nos = set()

    for report_id, sample_id, covers_samples in rows:
        if sample_id:
            sample_id_set.add(sample_id)
        if covers_samples:
            covers_samples_nos.update(covers_samples)

    if covers_samples_nos:
        cur.execute("""
            SELECT sample_id FROM samples WHERE sample_no = ANY(%s)
        """, (list(covers_samples_nos),))
        sample_id_set.update(row[0] for row in cur.fetchall())

    return list(sample_id_set)


def get_report_info_for_samples(sample_ids: List[int], cur) -> dict:
    """
    Reverse lookup of get_sample_ids_for_reports: given a list of sample_id
    values, return a dict mapping sample_id -> {"report_no", "created_at"}
    for the report that covers each sample (either as its primary sample_id,
    or via the covers_samples array).

    If a sample happens to be covered by more than one report, the most
    recently created one wins (ORDER BY r.created_at DESC, so the first
    match found per sample is kept).
    """
    if not sample_ids:
        return {}

    result = {}

    # Primary match: reports.sample_id = the invoice item's sample_id
    cur.execute("""
        SELECT sample_id, report_no, created_at
        FROM reports
        WHERE sample_id = ANY(%s)
        ORDER BY created_at DESC
    """, (sample_ids,))
    for sample_id, report_no, created_at in cur.fetchall():
        result.setdefault(sample_id, {"report_no": report_no, "created_at": created_at})

    # Remaining samples (not matched as a primary sample_id) may still be
    # covered via a report's covers_samples array (stored as sample_no text).
    remaining = [sid for sid in sample_ids if sid not in result]
    if remaining:
        cur.execute("""
            SELECT sample_id, sample_no
            FROM samples
            WHERE sample_id = ANY(%s)
        """, (remaining,))
        sample_no_by_id = {row[0]: row[1] for row in cur.fetchall()}
        sample_nos = [sample_no_by_id[sid] for sid in remaining if sid in sample_no_by_id]

        if sample_nos:
            cur.execute("""
                SELECT covers_samples, report_no, created_at
                FROM reports
                WHERE covers_samples IS NOT NULL
                AND EXISTS (
                    SELECT 1 FROM unnest(covers_samples) cs WHERE cs = ANY(%s)
                )
                ORDER BY created_at DESC
            """, (sample_nos,))
            for covers_samples, report_no, created_at in cur.fetchall():
                if not covers_samples:
                    continue
                for sid in remaining:
                    sno = sample_no_by_id.get(sid)
                    if sno and sno in covers_samples and sid not in result:
                        result[sid] = {"report_no": report_no, "created_at": created_at}

    return result


def get_project_quotation_items(project_id: int, cur, sample_ids: Optional[List[int]] = None):
    """
    Get invoiceable items using a single bulk SQL query instead of N+1 loops.

    Replicates the exact same sample→test assignment logic as the original:
      assigned_test_index = sample_position % num_tests_in_request
    where sample_position is the 0-based index of the sample within its
    test_request, ordered by sample_id.

    Returns list of tuples:
      (item_id, description, test_standard, unit_rate, 1,
       test_request_id, request_no, sample_id, sample_no, sample_status)
    """
    if sample_ids is not None and len(sample_ids) == 0:
        return []

    # Build the query in two variants to avoid mixing %s / %(name)s styles
    if sample_ids is not None:
        query = """
            WITH
            project_samples AS (
                SELECT
                    s.sample_id,
                    s.sample_no,
                    s.status            AS sample_status,
                    s.request_id,
                    tr.test_request_id,
                    tr.request_no,
                    (ROW_NUMBER() OVER (
                        PARTITION BY s.request_id
                        ORDER BY s.sample_id
                    ) - 1)              AS sample_pos
                FROM projects p
                JOIN test_requests tr ON p.project_id = tr.project_id
                JOIN samples s        ON tr.test_request_id = s.request_id
                WHERE p.project_id = %s
                  AND s.sample_id = ANY(%s)
            ),
            request_tests AS (
                SELECT
                    tri.test_request_id,
                    qi.item_id,
                    qi.description,
                    qi.test_standard,
                    qi.unit_rate,
                    (ROW_NUMBER() OVER (
                        PARTITION BY tri.test_request_id
                        ORDER BY tri.tri_id
                    ) - 1)              AS test_idx,
                    COUNT(*) OVER (PARTITION BY tri.test_request_id) AS num_tests
                FROM test_request_items tri
                JOIN quotation_items qi ON tri.quotation_item_id = qi.item_id
            )
            SELECT
                rt.item_id,
                rt.description,
                rt.test_standard,
                rt.unit_rate,
                ps.test_request_id,
                ps.request_no,
                ps.sample_id,
                ps.sample_no,
                ps.sample_status
            FROM project_samples ps
            JOIN request_tests rt
              ON rt.test_request_id = ps.test_request_id
             AND rt.test_idx = (ps.sample_pos %% rt.num_tests)
            ORDER BY ps.sample_id
        """
        cur.execute(query, (project_id, sample_ids))
    else:
        query = """
            WITH
            project_samples AS (
                SELECT
                    s.sample_id,
                    s.sample_no,
                    s.status            AS sample_status,
                    s.request_id,
                    tr.test_request_id,
                    tr.request_no,
                    (ROW_NUMBER() OVER (
                        PARTITION BY s.request_id
                        ORDER BY s.sample_id
                    ) - 1)              AS sample_pos
                FROM projects p
                JOIN test_requests tr ON p.project_id = tr.project_id
                JOIN samples s        ON tr.test_request_id = s.request_id
                WHERE p.project_id = %s
            ),
            request_tests AS (
                SELECT
                    tri.test_request_id,
                    qi.item_id,
                    qi.description,
                    qi.test_standard,
                    qi.unit_rate,
                    (ROW_NUMBER() OVER (
                        PARTITION BY tri.test_request_id
                        ORDER BY tri.tri_id
                    ) - 1)              AS test_idx,
                    COUNT(*) OVER (PARTITION BY tri.test_request_id) AS num_tests
                FROM test_request_items tri
                JOIN quotation_items qi ON tri.quotation_item_id = qi.item_id
            )
            SELECT
                rt.item_id,
                rt.description,
                rt.test_standard,
                rt.unit_rate,
                ps.test_request_id,
                ps.request_no,
                ps.sample_id,
                ps.sample_no,
                ps.sample_status
            FROM project_samples ps
            JOIN request_tests rt
              ON rt.test_request_id = ps.test_request_id
             AND rt.test_idx = (ps.sample_pos %% rt.num_tests)
            ORDER BY ps.sample_id
        """
        cur.execute(query, (project_id,))

    rows = cur.fetchall()
    return [
        (
            row[0],   # item_id
            row[1],   # description
            row[2],   # test_standard
            row[3],   # unit_rate
            1,        # quantity always 1 per sample
            row[4],   # test_request_id
            row[5],   # request_no
            row[6],   # sample_id
            row[7],   # sample_no
            row[8],   # sample_status
        )
        for row in rows
    ]

def get_invoice_complete(invoice_id: int, cur):
    """Get complete invoice details with items - FIXED for PROFORMA/TAX filtered totals with payment_method support"""
    # Get invoice header with payment_method
    cur.execute("""
        SELECT i.invoice_id, i.invoice_no, i.project_id, i.invoice_type, i.payment_method, i.invoice_date,
               i.client_reference, i.lpo_reference, i.lpo_date, i.payment_terms,
               i.subtotal, i.vat, i.total, i.amount_in_words, i.services_description, 
               i.remarks, i.payment_status, i.paid_date,
               p.project_no, p.project_name, p.location,
               c.client_id, c.name, c.contact_person, c.email, c.address, c.phone,
               i.generation_mode
        FROM invoices i
        JOIN projects p ON i.project_id = p.project_id
        JOIN clients c ON p.client_id = c.client_id
        WHERE i.invoice_id = %s
    """, (invoice_id,))
    
    header = cur.fetchone()
    if not header:
        raise HTTPException(status_code=404, detail="Invoice not found")
    
    # Get invoice items
    cur.execute("""
        SELECT 
            ii.item_id, 
            ii.description, 
            ii.test_standard, 
            ii.unit_rate, 
            ii.quantity, 
            ii.amount, 
            ii.sample_id, 
            s.sample_no,
            s.status as sample_status,
            tr.request_no,
            tr.test_request_id
        FROM invoice_items ii
        LEFT JOIN samples s ON ii.sample_id = s.sample_id
        LEFT JOIN test_requests tr ON ii.test_request_id = tr.test_request_id
        WHERE ii.invoice_id = %s
        ORDER BY ii.item_id
    """, (invoice_id,))
    
    items_data = cur.fetchall()
    
    # Get invoice type
    invoice_type = header[3]
    payment_method = header[4]  # NEW: Get payment_method
    
    # =====================================================
    # NOTE: invoice_items is now correctly scoped at invoice-creation
    # time (only samples belonging to the reports the user actually
    # selected are inserted), so no further fuzzy re-filtering by
    # description/test-type text matching happens here. That used to
    # exist as a second, independent filter and could wrongly include
    # or exclude items when two reports had similar test-type names.
    # =====================================================
    
    # Build items list
    items_list = []
    filtered_subtotal = 0.0
    
    for item in items_data:
        item_amount = float(item[5]) if isinstance(item[5], Decimal) else item[5]
        filtered_subtotal += item_amount
        
        items_list.append({
            "item_id": item[0],
            "description": item[1],
            "test_standard": item[2],
            "unit_rate": float(item[3]) if isinstance(item[3], Decimal) else item[3],
            "quantity": item[4],
            "amount": item_amount,
            "sample_id": item[6],
            "sample_no": item[7],
            "sample_status": item[8],
            "request_no": item[9],
            "test_request_id": item[10]
        })
    
    # =====================================================
    # NEW: Calculate CORRECT totals for filtered items
    # =====================================================
    if invoice_type in ["PROFORMA", "TAX"] and items_list:
        # Recalculate based on filtered items
        filtered_vat = filtered_subtotal * 0.05
        filtered_total = filtered_subtotal + filtered_vat
        filtered_amount_words = number_to_words(filtered_total)
        
        # Check if different from original
        original_subtotal = float(header[10]) if isinstance(header[10], Decimal) else header[10]
        
        if abs(original_subtotal - filtered_subtotal) > 0.01:
            print(f"DEBUG: Using FILTERED totals for {invoice_type} invoice {invoice_id}")
            print(f"  Original: Subtotal={original_subtotal}")
            print(f"  Filtered: Subtotal={filtered_subtotal}, Total={filtered_total}")
            
            subtotal = filtered_subtotal
            vat = filtered_vat
            total = filtered_total
            amount_words = filtered_amount_words
        else:
            # Use original totals
            subtotal = original_subtotal
            vat = float(header[11]) if isinstance(header[11], Decimal) else header[11]
            total = float(header[12]) if isinstance(header[12], Decimal) else header[12]
            amount_words = header[13]
    else:
        # For other invoice types, use original totals
        subtotal = float(header[10]) if isinstance(header[10], Decimal) else header[10]
        vat = float(header[11]) if isinstance(header[11], Decimal) else header[11]
        total = float(header[12]) if isinstance(header[12], Decimal) else header[12]
        amount_words = header[13]
    
    # Return the complete invoice data
    return {
        "invoice_id": header[0],
        "invoice_no": header[1],
        "project_id": header[2],
        "invoice_type": invoice_type,
        "payment_method": payment_method,  # NEW
        "invoice_date": header[5],
        "client_reference": header[6],
        "lpo_reference": header[7],
        "lpo_date": header[8],
        "payment_terms": header[9],
        "subtotal": subtotal,
        "vat": vat,
        "total": total,
        "amount_in_words": amount_words,
        "services_description": header[14],
        "remarks": header[15],
        "payment_status": header[16],
        "paid_date": header[17],
        "items": items_list,
        "generation_mode": header[27],
        "project_details": {
            "project_no": header[18],
            "project_name": header[19],
            "location": header[20],
            "client_id": header[21],
            "client_name": header[22],
            "client_contact": header[23],
            "client_email": header[24],
            "client_address": header[25],
            "client_phone": header[26]
        },
        # Add these new fields to help frontend understand what's happening
        "is_filtered": invoice_type in ["PROFORMA", "TAX"] and len(items_list) > 0,
        "filtered_item_count": len(items_list),
        "original_total": float(header[12]) if isinstance(header[12], Decimal) else header[12]
    }
# ----------------------------
# CREATE INVOICE (internal implementation)
# ----------------------------
def _create_invoice_with_payment_method_impl(
    payload: InvoiceCreate,
    sample_ids: Optional[List[int]] = None,
    shared_cur=None,          # ← NEW: pass caller's cursor to avoid cross-connection deadlocks
):
    """
    Shared implementation for creating an invoice.

    `sample_ids`  — when provided, restricts which project samples/tests get
                    pulled onto the invoice.
    `shared_cur`  — when provided by the caller (e.g. generate-with-reports-and-tests-v2),
                    all queries run on the *same* connection/transaction.
                    The caller is responsible for commit/rollback.
                    When None (standalone callers), this function manages its own
                    connection and commits internally.
    """
    _own_conn = shared_cur is None          # True  → we manage conn lifecycle
    conn = None
    cur  = shared_cur

    if _own_conn:
        conn = get_connection()
        cur  = conn.cursor()

    try:
        # ---------------------------------------------------
        # 1. Get project details with client info
        # ---------------------------------------------------
        cur.execute("""
            SELECT p.project_id, p.project_no, p.project_name, p.location, p.lpo_no, p.lpo_date,
                   c.client_id, c.name, c.contact_person, c.email, c.address, c.phone,
                   q.quotation_no
            FROM projects p
            JOIN clients c ON p.client_id = c.client_id
            LEFT JOIN quotations q ON p.quotation_id = q.quotation_id
            WHERE p.project_id = %s
        """, (payload.project_id,))
        project_data = cur.fetchone()
        if not project_data:
            raise HTTPException(status_code=404, detail="Project not found")

        # ---------------------------------------------------
        # 2. Get invoiceable items (optionally restricted to selected reports)
        # ---------------------------------------------------
        invoiceable_items = get_project_quotation_items(payload.project_id, cur, sample_ids=sample_ids)
        if not invoiceable_items:
            raise HTTPException(status_code=400, detail="No test items available for invoicing")

        # ---------------------------------------------------
        # 2a. Group by description / test_standard / unit_rate
        # ---------------------------------------------------
        from collections import defaultdict

        grouped_items = defaultdict(lambda: {
            "description": "", "test_standard": "", "unit_rate": 0,
            "quantity": 0, "sample_ids": []
        })

        for item in invoiceable_items:
            item_id, description, test_standard, unit_rate, quantity, test_request_id, request_no, sample_id, sample_no, sample_status = item
            key = (description, test_standard, unit_rate)
            grouped_items[key]["description"]  = description
            grouped_items[key]["test_standard"] = test_standard
            grouped_items[key]["unit_rate"]     = unit_rate
            grouped_items[key]["quantity"]     += quantity
            grouped_items[key]["sample_ids"].append(sample_id)

        final_items = [
            (desc, std, rate, data["quantity"], data["sample_ids"])
            for (desc, std, rate), data in grouped_items.items()
        ]

        # ---------------------------------------------------
        # 3. Generate invoice number
        # ---------------------------------------------------
        invoice_no = generate_invoice_no(cur, payload.invoice_type)

        # ---------------------------------------------------
        # 4. Calculate totals
        # ---------------------------------------------------
        subtotal = 0.0
        for desc, std, unit_rate, quantity, item_sample_ids in final_items:
            unit_rate_float = float(unit_rate) if isinstance(unit_rate, Decimal) else unit_rate
            subtotal += unit_rate_float * quantity

        vat          = subtotal * 0.05
        total        = subtotal + vat
        amount_words = number_to_words(total)

        # ---------------------------------------------------
        # 5. LPO and payment terms
        # ---------------------------------------------------
        lpo_reference = payload.lpo_reference or project_data[4]
        lpo_date      = payload.lpo_date      or project_data[5]
        payment_terms = payload.payment_terms or "30 days" if payload.payment_method == "CREDIT" else "Immediate"

        # ---------------------------------------------------
        # 6. Insert invoice header
        # ---------------------------------------------------
        cur.execute("""
            INSERT INTO invoices (
                invoice_no, project_id, invoice_type, payment_method, invoice_date,
                client_reference, lpo_reference, lpo_date, payment_terms,
                subtotal, vat, total, amount_in_words, services_description, remarks,
                payment_status
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING invoice_id
        """, (
            invoice_no,
            payload.project_id,
            payload.invoice_type,
            payload.payment_method,
            payload.invoice_date or date.today(),
            payload.client_reference,
            lpo_reference,
            lpo_date,
            payment_terms,
            subtotal, vat, total, amount_words,
            payload.services_description or f"Testing services for {project_data[2]}",
            payload.remarks,
            "UNPAID",
        ))
        invoice_id = cur.fetchone()[0]

        # ---------------------------------------------------
        # 7. Insert grouped invoice items
        # ---------------------------------------------------
        for description, test_standard, unit_rate, quantity, item_sample_ids in final_items:
            unit_rate_float = float(unit_rate) if isinstance(unit_rate, Decimal) else unit_rate
            amount    = unit_rate_float * quantity
            sample_id = item_sample_ids[0] if item_sample_ids else None

            cur.execute("""
                INSERT INTO invoice_items (
                    invoice_id, description, test_standard,
                    unit_rate, quantity, amount, sample_id
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s)
            """, (invoice_id, description, test_standard,
                  unit_rate_float, quantity, amount, sample_id))

        # ---------------------------------------------------
        # 8. Commit only when we own the connection
        # ---------------------------------------------------
        if _own_conn:
            conn.commit()

        # ---------------------------------------------------
        # 9. Return complete invoice
        # ---------------------------------------------------
        return get_invoice_complete(invoice_id, cur)

    except Exception as e:
        if _own_conn and conn:
            conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if _own_conn:
            cur.close()
            if conn:
                conn.close()

# ----------------------------
# ----------------------------
# DELETE INVOICE
# ----------------------------
from fastapi import Query

@router.delete("/{invoice_id}")
def delete_invoice(invoice_id: int):
    """
    Delete an invoice and its items from the database.
    """
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        # First, get invoice details for response
        cur.execute("""
            SELECT invoice_no FROM invoices WHERE invoice_id = %s
        """, (invoice_id,))
        
        invoice_data = cur.fetchone()
        if not invoice_data:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        # Delete invoice items first (foreign key constraint)
        cur.execute("DELETE FROM invoice_items WHERE invoice_id = %s", (invoice_id,))
        
        # Delete the invoice
        cur.execute("DELETE FROM invoices WHERE invoice_id = %s RETURNING invoice_id", (invoice_id,))
        
        if not cur.fetchone():
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        conn.commit()
        
        return {
            "message": f"Invoice {invoice_data[0]} deleted successfully",
            "deleted_invoice_id": invoice_id,
            "invoice_no": invoice_data[0]
        }
        
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()









# ----------------------------
# GET LATEST PROJECTS FOR INVOICE DROPDOWN - ENHANCED VERSION
# ----------------------------
@router.get("/projects/latest/")
def get_latest_projects():
    """
    Get the latest 10 projects with complete info for invoice creation.
    Returns: Array of project objects with details.
    """
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                p.project_id,
                p.project_name,
                p.project_no,
                COALESCE(c.name, p.client_name, p.walk_in_client) as client_name,
                p.location,
                q.quotation_no,
                p.is_walk_in
            FROM projects p
            LEFT JOIN clients c ON p.client_id = c.client_id
            LEFT JOIN quotations q ON p.quotation_id = q.quotation_id
            WHERE p.is_walk_in IS NOT TRUE
               OR (p.is_walk_in = TRUE AND p.quotation_id IS NOT NULL AND p.project_no != 'PENDING')
            ORDER BY p.project_id DESC
            LIMIT 10
        """)

        projects = []
        for row in cur.fetchall():
            project_id, project_name, project_no, client_name, location, quotation_no, is_walk_in = row
            display_label = f"{project_no} - {project_name} ({client_name})"
            projects.append({
                "project_id": project_id,
                "project_name": display_label,
                "project_no": project_no,
                "project_name_raw": project_name,
                "client_name": client_name,
                "location": location,
                "quotation_no": quotation_no,
                "is_walk_in": bool(is_walk_in),
                "value": project_id,
                "label": display_label,
            })

        return projects

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()


@router.get("/projects/search/")
def search_projects(q: str = ""):
    """
    Search projects by Plot Number (project_no) or LP number.
    Supports partial, case-insensitive matching.
    Returns up to 20 matching projects.
    Used by the Generate Invoices search feature.
    """
    conn = get_connection()
    cur = conn.cursor()

    try:
        term = f"%{q.strip()}%"
        limit_clause = "LIMIT 20" if q.strip() else ""
        cur.execute(f"""
            SELECT
                p.project_id,
                p.project_name,
                p.project_no,
                COALESCE(c.name, p.client_name, p.walk_in_client) as client_name,
                p.location,
                q.quotation_no,
                p.is_walk_in
            FROM projects p
            LEFT JOIN clients c ON p.client_id = c.client_id
            LEFT JOIN quotations q ON p.quotation_id = q.quotation_id
            WHERE (
                p.is_walk_in IS NOT TRUE
                OR (p.is_walk_in = TRUE AND p.quotation_id IS NOT NULL AND p.project_no != 'PENDING')
            )
            AND (
                p.project_no ILIKE %s
                OR p.project_name ILIKE %s
                OR COALESCE(c.name, p.client_name, p.walk_in_client) ILIKE %s
            )
            ORDER BY p.project_id DESC
            {limit_clause}
        """, (term, term, term))

        projects = []
        for row in cur.fetchall():
            project_id, project_name, project_no, client_name, location, quotation_no, is_walk_in = row
            display_label = f"{project_no} - {project_name} ({client_name})"
            projects.append({
                "project_id": project_id,
                "project_name": display_label,
                "project_no": project_no,
                "project_name_raw": project_name,
                "client_name": client_name,
                "location": location,
                "quotation_no": quotation_no,
                "is_walk_in": bool(is_walk_in),
                "value": project_id,
                "label": display_label,
            })

        return projects

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()



# Add this to your invoices.py file, after the existing endpoints

# ----------------------------
# DELIVERY NOTE GENERATION
# ----------------------------

# Replace the existing DeliveryNoteRequest class with this:
class DeliveryNoteRequest(BaseModel):
    project_id: int
    selected_report_ids: Optional[List[int]] = None
    include_all_reports: bool = True





def generate_delivery_note_number(cur):
    """Generate delivery note number: 13212/25, 13213/25, etc."""
    # Get the last 2 digits of current year
    year_short = str(datetime.now().year)[-2:]
    
    try:
        # Check if delivery_notes table exists
        cur.execute("""
            SELECT EXISTS (
                SELECT FROM information_schema.tables 
                WHERE table_name = 'delivery_notes'
            )
        """)
        table_exists = cur.fetchone()[0]
        
        if table_exists:
            # Get max delivery note number
            cur.execute("""
                SELECT delivery_note_no FROM delivery_notes 
                ORDER BY delivery_note_id DESC LIMIT 1
            """)
            last_dn = cur.fetchone()
            
            if last_dn and last_dn[0]:
                try:
                    # Extract number from format like "13212/25"
                    last_number = int(last_dn[0].split('/')[0])
                    next_number = last_number + 1
                except:
                    next_number = 13212
            else:
                next_number = 13212
        else:
            # Table doesn't exist, start from 13212
            next_number = 13212
            
    except Exception as e:
        print(f"Error in generate_delivery_note_number: {e}")
        next_number = 13212
    
    return f"{next_number}/{year_short}"




@router.get("/projects/{project_id}/reports-for-delivery")
def get_reports_for_delivery_note(project_id: int):
    """
    Get all approved reports for a project to select for delivery note
    Now includes delivery note status check
    """
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        # Ensure the junction table exists before querying
        ensure_delivery_note_reports_table(cur)
        
        # Get project details
        cur.execute(""" 
            SELECT p.project_no, p.project_name, c.name as client_name
            FROM projects p
            JOIN clients c ON p.client_id = c.client_id
            WHERE p.project_id = %s
        """, (project_id,))
        
        project_data = cur.fetchone()
        if not project_data:
            raise HTTPException(status_code=404, detail="Project not found")
        
        project_no, project_name, client_name = project_data
        
        print(f"DEBUG: Fetching reports for project {project_id} - {project_name}")
        
        # Get all approved reports for this project
        # SIMPLIFIED QUERY - Let's first get all reports, then check delivery note status
        cur.execute("""
            SELECT DISTINCT ON (r.report_no)
                r.report_id, 
                r.report_no, 
                r.created_at,
                r.covers_test_type as test_name,
                r.covers_samples,
                -- Count samples in the covers_samples array
                array_length(r.covers_samples, 1) as sample_count
            FROM reports r
            JOIN samples s ON r.sample_id = s.sample_id
            JOIN test_requests tr ON s.request_id = tr.test_request_id
            WHERE tr.project_id = %s 
            AND r.status = 'APPROVED'
            ORDER BY r.report_no, r.created_at DESC
        """, (project_id,))
        
        all_reports = cur.fetchall()
        
        print(f"DEBUG: Found {len(all_reports)} approved reports")

        # ── Bulk-fetch delivery-note status and sample_nos ───────────────────
        all_report_nos = [row[1] for row in all_reports]

        # Delivered status: single bulk query instead of N+1
        delivered_set = set()
        if all_report_nos:
            cur.execute(
                "SELECT report_no FROM delivery_note_reports WHERE report_no = ANY(%s)",
                (all_report_nos,)
            )
            delivered_set = {r[0] for r in cur.fetchall()}

        # covers_samples bulk validation
        covers_samples_map = {}  # report_no -> list of sample_no strings
        primary_report_nos_dn = []
        for row in all_reports:
            rno  = row[1]
            csmp = row[4]
            if csmp:
                covers_samples_map[rno] = list(csmp)
            else:
                primary_report_nos_dn.append(rno)

        all_covers_nos_dn = [s for nos in covers_samples_map.values() for s in nos]
        valid_covers_set_dn = set()
        if all_covers_nos_dn:
            cur.execute(
                "SELECT sample_no FROM samples WHERE sample_no = ANY(%s)",
                (all_covers_nos_dn,)
            )
            valid_covers_set_dn = {r[0] for r in cur.fetchall()}

        primary_sno_by_report_dn = {}
        if primary_report_nos_dn:
            cur.execute(
                """
                SELECT r.report_no, s.sample_no
                FROM reports r
                JOIN samples s ON r.sample_id = s.sample_id
                WHERE r.report_no = ANY(%s)
                """,
                (primary_report_nos_dn,)
            )
            for rno, sno in cur.fetchall():
                primary_sno_by_report_dn[rno] = sno
        # ────────────────────────────────────────────────────────────────────

        reports = []
        delivered_count = 0
        undelivered_count = 0
        
        for row in all_reports:
            report_id = row[0]
            report_no = row[1]
            created_date = row[2]
            test_name = row[3]
            covers_samples = row[4]
            sample_count = row[5] or 0
            
            already_in_delivery_note = report_no in delivered_set

            # Use pre-fetched sample_nos
            if covers_samples:
                sample_nos = [s for s in covers_samples_map.get(report_no, [])
                              if s in valid_covers_set_dn]
            else:
                sno = primary_sno_by_report_dn.get(report_no)
                if sno:
                    sample_nos   = [sno]
                    sample_count = 1
                else:
                    sample_nos = []
            
            if already_in_delivery_note:
                delivered_count += 1
                print(f"DEBUG: Report {report_no} is already in delivery note")
            else:
                undelivered_count += 1
                print(f"DEBUG: Report {report_no} is NOT in delivery note yet")
            
            reports.append({
                "report_id": report_id,
                "report_no": report_no,
                "created_date": created_date.strftime("%Y-%m-%d") if created_date else None,
                "test_name": test_name or "Test Report",
                "sample_count": sample_count,
                "covers_samples": sample_nos,
                "already_invoiced": already_in_delivery_note,  # Using same field name for compatibility
                "status": "Delivered" if already_in_delivery_note else "Not Delivered"
            })
        
        print(f"DEBUG: Total: {len(reports)}, Delivered: {delivered_count}, Undelivered: {undelivered_count}")
        
        return {
            "project_id": project_id,
            "project_no": project_no,
            "project_name": project_name,
            "client_name": client_name,
            "total_reports": len(reports),
            "delivered_count": delivered_count,
            "undelivered_count": undelivered_count,
            "reports": reports
        }
        
    except Exception as e:
        print(f"ERROR in get_reports_for_delivery_note: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    finally:
        cur.close()
        conn.close()








@router.post("/delivery-notes/generate-excel-template")
def generate_delivery_note_excel_template(payload: DeliveryNoteRequest):
    """
    Generate delivery note Excel file using the template for selected reports
    """
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        # Get project details
        cur.execute("""
            SELECT p.project_no, p.project_name, c.name as client_name,
                   c.address, c.phone, c.email
            FROM projects p
            JOIN clients c ON p.client_id = c.client_id
            WHERE p.project_id = %s
        """, (payload.project_id,))
        
        project_data = cur.fetchone()
        if not project_data:
            raise HTTPException(status_code=404, detail="Project not found")
        
        project_no, project_name, client_name, client_address, client_phone, client_email = project_data
        
        # Get selected reports or all if include_all_reports is True
        if payload.include_all_reports:
            # Get all approved reports BUT EXCLUDE ALREADY DELIVERED ONES
            cur.execute("""
                SELECT DISTINCT ON (r.report_no)
                    r.report_id, r.report_no, r.created_at,
                    r.covers_test_type as test_name,
                    r.covers_samples,
                    array_length(r.covers_samples, 1) as sample_count
                FROM reports r
                JOIN samples s ON r.sample_id = s.sample_id
                JOIN test_requests tr ON s.request_id = tr.test_request_id
                WHERE tr.project_id = %s 
                AND r.status = 'APPROVED'
                AND r.report_no NOT IN (
                    SELECT report_no FROM delivery_note_reports
                )
                ORDER BY r.report_no, r.created_at DESC
            """, (payload.project_id,))
        elif payload.selected_report_ids:
            # Get specific selected reports
            cur.execute("""
                SELECT DISTINCT ON (r.report_no)
                    r.report_id, r.report_no, r.created_at,
                    r.covers_test_type as test_name,
                    r.covers_samples,
                    array_length(r.covers_samples, 1) as sample_count
                FROM reports r
                WHERE r.report_id = ANY(%s)
                AND r.status = 'APPROVED'
                ORDER BY r.report_no, r.created_at DESC
            """, (payload.selected_report_ids,))
        else:
            raise HTTPException(status_code=400, detail="No reports selected")
        
        reports_data = cur.fetchall()
        
        if not reports_data:
            if payload.include_all_reports:
                raise HTTPException(status_code=400, detail="No undelivered reports found for this project")
            else:
                raise HTTPException(status_code=400, detail="No approved reports found for this project")
        
        # Generate delivery note number
        delivery_note_no = generate_delivery_note_number(cur)
        
        # =====================================================
        # 1. Load the Excel template
        # =====================================================
        template_path = download_template_from_supabase("delivery_note")
        
        if not os.path.exists(template_path):
            raise HTTPException(status_code=404, detail="Delivery note template not found")
        
        wb = openpyxl.load_workbook(template_path, data_only=False)
        ws = wb.active
        
        # =====================================================
        # 2. Fill Template Fields
        # =====================================================
        # Fill Ref. No.: delivery note number
        ws["B6"] = delivery_note_no  # Ref. No.: (row 10, column B)
        
        # Fill Lab Project No.: project number
        ws["B7"] = project_no  # Lab Project No.: (row 12, column B)
        
        # Fill Customer Name:
        ws["B8"] = client_name  # Customer Name: (row 13, column B)
        
        # P.O.Box: (already has Dubai - U.A.E.)
        # This is at ws["B14"] which already has "Dubai - U.A.E."
        
        # =====================================================
        # 3. Fill Report Table
        # =====================================================
        # Starting row for reports in the template (row 19 based on your template)
        START_ROW = 12
        
        for i, report in enumerate(reports_data, 1):
            row = START_ROW + i - 1
            report_id, report_no, created_date, test_name, covers_samples, sample_count = report
            
            # Use the actual sample count from the covers_samples array
            actual_sample_count = sample_count or 1
            if covers_samples and isinstance(covers_samples, list):
                actual_sample_count = len(covers_samples)
            
            # Fill columns according to template:
            # A = Report No.
            # B = Description
            # G = No. Tests (based on your template with columns A-K)
            
            ws[f"A{row}"] = report_no
            ws[f"B{row}"] = test_name or "Test Report"
            ws[f"G{row}"] = actual_sample_count
        
        # Clear any remaining rows in the template
        MAX_TEMPLATE_ROWS = 34  # Adjust based on your template
        for row in range(START_ROW + len(reports_data), MAX_TEMPLATE_ROWS + 1):
            ws[f"A{row}"].value = None
            ws[f"B{row}"].value = None
            ws[f"G{row}"].value = None
        
        # =====================================================
        # 4. Save Final File on Server
        # =====================================================
        output_dir = "generated_delivery_notes"
        os.makedirs(output_dir, exist_ok=True)
        
        # Clean filename — same approach as tax/proforma invoices
        import re, urllib.parse
        def clean_dn_filename(text):
            if not text:
                return ""
            text = str(text).strip()
            text = re.sub(r'[\\/*?:"<>|]', '-', text)
            text = re.sub(r'\s+', '-', text)
            text = text.strip('-')
            return text

        dn_clean = delivery_note_no.strip().replace('/', '-')
        clean_project_no = clean_dn_filename(project_no)
        filename = f"DN-{dn_clean}-{clean_project_no}.xlsx"
        filepath = os.path.join(output_dir, filename)
        
        wb.save(filepath)
        
        # =====================================================
        # 5. NEW: Track which reports were included in this delivery note
        # =====================================================
        try:
            # First ensure the delivery_note_reports junction table exists
            ensure_delivery_note_reports_table(cur)
            
            # Insert records for each report included
            for report in reports_data:
                report_no = report[1]  # report_no is at index 1
                cur.execute("""
                    INSERT INTO delivery_note_reports (delivery_note_no, report_no)
                    VALUES (%s, %s)
                    ON CONFLICT (delivery_note_no, report_no) DO NOTHING
                """, (delivery_note_no, report_no))
            
            # Update delivery_notes table - FIXED to match existing schema
            cur.execute("""
                CREATE TABLE IF NOT EXISTS delivery_notes (
                    delivery_note_id SERIAL PRIMARY KEY,
                    delivery_note_no VARCHAR(50) NOT NULL UNIQUE,
                    project_id INTEGER NOT NULL REFERENCES projects(project_id),
                    generated_by INTEGER REFERENCES users(user_id),
                    generated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    total_reports INTEGER DEFAULT 0,
                    file_path TEXT
                )
            """)
            
            # Use existing user ID or default to NULL
            # In a real app, you would get this from the current user session
            user_id = None  # You might want to pass this from the frontend or use a default
            
            cur.execute("""
                INSERT INTO delivery_notes 
                (delivery_note_no, project_id, generated_by, total_reports, file_path)
                VALUES (%s, %s, %s, %s, %s)
            """, (delivery_note_no, payload.project_id, user_id, len(reports_data), filepath))
            
            conn.commit()
        except Exception as db_error:
            print(f"Note: Could not save delivery note record: {db_error}")
            # Don't fail the request if we can't save the record
            if conn:
                conn.rollback()
        
        # =====================================================
        # 6. Return File for Download
        # =====================================================
        encoded_filename = urllib.parse.quote(filename)
        return FileResponse(
            filepath,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}; filename=\"{filename}\""
            }
        )
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in generate_delivery_note_excel_template: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating delivery note: {str(e)}")
    finally:
        cur.close()
        conn.close()









@router.get("/projects/{project_id}/reports-for-invoice/{invoice_type}")
def get_reports_for_invoice(project_id: int, invoice_type: str):
    """
    Get all approved reports for a project to select for PROFORMA or TAX invoice
    Similar to delivery notes but checks invoice_report_links table
    """
    if invoice_type not in ['PROFORMA', 'TAX']:
        raise HTTPException(status_code=400, detail="Invoice type must be PROFORMA or TAX")
    
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        # Ensure invoice_report_links exists before querying it
        ensure_walkin_proforma_records_table(cur)
        conn.commit()

        # Get project details
        cur.execute(""" 
            SELECT p.project_no, p.project_name, c.name as client_name
            FROM projects p
            JOIN clients c ON p.client_id = c.client_id
            WHERE p.project_id = %s
        """, (project_id,))
        
        project_data = cur.fetchone()
        if not project_data:
            raise HTTPException(status_code=404, detail="Project not found")
        
        project_no, project_name, client_name = project_data
        
        print(f"DEBUG: Fetching reports for {invoice_type} invoice - project {project_id}")
        
        # Get all approved reports for this project
        # Check if they're already in invoice_report_links for this invoice type
        cur.execute("""
            SELECT DISTINCT ON (r.report_no)
                r.report_id, 
                r.report_no, 
                r.created_at,
                r.covers_test_type as test_name,
                r.covers_samples,
                -- Count samples in the covers_samples array
                array_length(r.covers_samples, 1) as sample_count,
                -- Check if already in PROFORMA/TAX invoices
                EXISTS (
                    SELECT 1 FROM invoice_report_links irl
                    WHERE irl.report_no = r.report_no 
                    AND irl.invoice_type = %s
                ) as already_invoiced
            FROM reports r
            JOIN samples s ON r.sample_id = s.sample_id
            JOIN test_requests tr ON s.request_id = tr.test_request_id
            WHERE tr.project_id = %s 
            AND r.status = 'APPROVED'
            ORDER BY r.report_no, r.created_at DESC
        """, (invoice_type, project_id))
        
        all_reports = cur.fetchall()
        
        print(f"DEBUG: Found {len(all_reports)} approved reports for {invoice_type}")

        # ── Bulk-fetch sample_nos to avoid N+1 queries ──────────────────────
        # Collect all sample_nos referenced by covers_samples arrays, plus
        # the primary sample_id for reports that have no covers_samples array.
        covers_samples_map = {}   # report_no -> list[sample_no_str]
        primary_sample_ids = []   # sample_id ints for reports without covers_samples

        for row in all_reports:
            report_no   = row[1]
            covers_samp = row[4]
            if covers_samp:
                covers_samples_map[report_no] = list(covers_samp)
            else:
                # We'll look up by primary sample_id via a separate bulk query below
                primary_sample_ids.append(report_no)  # placeholder — resolved next

        # For reports with covers_samples: validate they exist (single bulk query)
        all_covers_nos = [sno for nos in covers_samples_map.values() for sno in nos]
        valid_sample_nos_set = set()
        if all_covers_nos:
            cur.execute(
                "SELECT sample_no FROM samples WHERE sample_no = ANY(%s)",
                (all_covers_nos,)
            )
            valid_sample_nos_set = {r[0] for r in cur.fetchall()}

        # For reports without covers_samples: bulk-fetch primary sample_no
        primary_report_nos = [row[1] for row in all_reports if not row[4]]
        primary_sample_no_by_report = {}
        if primary_report_nos:
            cur.execute(
                """
                SELECT r.report_no, s.sample_no
                FROM reports r
                JOIN samples s ON r.sample_id = s.sample_id
                WHERE r.report_no = ANY(%s)
                """,
                (primary_report_nos,)
            )
            for rno, sno in cur.fetchall():
                primary_sample_no_by_report[rno] = sno
        # ────────────────────────────────────────────────────────────────────

        reports = []
        invoiced_count = 0
        uninvoiced_count = 0
        
        for row in all_reports:
            report_id = row[0]
            report_no = row[1]
            created_date = row[2]
            test_name = row[3]
            covers_samples = row[4]
            sample_count = row[5] or 0
            already_invoiced = row[6]
            
            # Use pre-fetched data — no per-row queries
            if covers_samples:
                sample_nos = [sno for sno in covers_samples_map.get(report_no, [])
                              if sno in valid_sample_nos_set]
            else:
                sno = primary_sample_no_by_report.get(report_no)
                if sno:
                    sample_nos  = [sno]
                    sample_count = 1
                else:
                    sample_nos = []
            
            if already_invoiced:
                invoiced_count += 1
                print(f"DEBUG: Report {report_no} is already in {invoice_type} invoice")
            else:
                uninvoiced_count += 1
                print(f"DEBUG: Report {report_no} is NOT in {invoice_type} invoice yet")
            
            reports.append({
                "report_id": report_id,
                "report_no": report_no,
                "created_date": created_date.strftime("%Y-%m-%d") if created_date else None,
                "test_name": test_name or "Test Report",
                "sample_count": sample_count,
                "covers_samples": sample_nos,
                "already_invoiced": already_invoiced,
                "invoice_type": invoice_type,
                "status": "Invoiced" if already_invoiced else "Not Invoiced"
            })
        
        print(f"DEBUG: Total: {len(reports)}, {invoice_type} Invoiced: {invoiced_count}, Not Invoiced: {uninvoiced_count}")
        
        return {
            "project_id": project_id,
            "project_no": project_no,
            "project_name": project_name,
            "client_name": client_name,
            "invoice_type": invoice_type,
            "total_reports": len(reports),
            "invoiced_count": invoiced_count,
            "uninvoiced_count": uninvoiced_count,
            "reports": reports
        }
        
    except Exception as e:
        print(f"ERROR in get_reports_for_invoice: {str(e)}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Database error: {str(e)}")
    finally:
        cur.close()
        conn.close()


# =====================================================
# MODIFIED: Create Invoice - Record report links
# =====================================================
@router.patch("/{invoice_id}/payment-status")
def update_invoice_payment_status(invoice_id: int, status_update: dict):
    """
    Update invoice payment status (PAID/UNPAID)
    """
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        # Validate status
        new_status = status_update.get("payment_status")
        if new_status not in ["PAID", "UNPAID"]:
            raise HTTPException(status_code=400, detail="Status must be PAID or UNPAID")
        
        # Update payment status and paid_date
        paid_date = None
        if new_status == "PAID":
            paid_date = status_update.get("paid_date") or date.today()
        
        cur.execute("""
            UPDATE invoices 
            SET payment_status = %s, paid_date = %s
            WHERE invoice_id = %s
            RETURNING invoice_id, invoice_no, payment_status, paid_date
        """, (new_status, paid_date, invoice_id))
        
        result = cur.fetchone()
        if not result:
            raise HTTPException(status_code=404, detail="Invoice not found")
        
        conn.commit()
        
        return {
            "message": f"Invoice {result[1]} payment status updated",
            "invoice_id": result[0],
            "invoice_no": result[1],
            "payment_status": result[2],
            "paid_date": result[3]
        }
        
    except Exception as e:
        conn.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()


def download_template_from_supabase(template_type: str = "invoice"):
    """
    Download template from Supabase storage.
    template_type: "invoice" or "delivery_note"
    """
    template_urls = {
        "invoice": "https://hqwgkmbjmcxpxbwccclo.supabase.co/storage/v1/object/public/templates/invoices/invoice.xlsx",
        "delivery_note": "https://hqwgkmbjmcxpxbwccclo.supabase.co/storage/v1/object/public/templates/invoices/delivery_note.xlsx",
        "invoices_combined": "https://hqwgkmbjmcxpxbwccclo.supabase.co/storage/v1/object/public/templates/invoices/invoices.xlsx"
    }
    
    if template_type not in template_urls:
        raise ValueError(f"Template type {template_type} not supported")
    
    url = template_urls[template_type]
    
    try:
        # Download the file
        response = requests.get(url, timeout=30)
        response.raise_for_status()
        
        # Create a temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
            temp_file.write(response.content)
            temp_path = temp_file.name
        
        print(f"DEBUG: Downloaded {template_type} template from {url}")
        return temp_path
        
    except requests.exceptions.RequestException as e:
        print(f"ERROR: Failed to download template from {url}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to download template: {e}")

# =====================================================
# COMBINED PROFORMA + TAX INVOICE GENERATION
# =====================================================

# -------------------------------------------------------
# HELPER: Write to a cell that may be part of a merged range
# -------------------------------------------------------
def set_cell(ws, cell_addr: str, value):
    """
    Write value to a cell even if it is part of a merged region.
    openpyxl raises ReadOnlyCell errors when you try to write to
    a non-top-left cell of a merged range, so we always find the
    top-left anchor and write there.
    """
    from openpyxl.utils import column_index_from_string
    import re

    # Parse the address
    match = re.match(r"([A-Za-z]+)(\d+)", cell_addr)
    if not match:
        ws[cell_addr] = value
        return
    col_letter, row_str = match.group(1).upper(), int(match.group(2))
    col_idx = column_index_from_string(col_letter)

    # Check if this cell is inside a merged range
    for merged_range in ws.merged_cells.ranges:
        if (merged_range.min_row <= row_str <= merged_range.max_row and
                merged_range.min_col <= col_idx <= merged_range.max_col):
            # Write to the top-left anchor only
            anchor = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            anchor.value = value
            return

    # Not merged - write normally
    ws[cell_addr] = value


# -------------------------------------------------------
# HELPER: Clear a row's data cells (merge-safe)
# -------------------------------------------------------
def clear_row_cols(ws, row: int, col_letters: list):
    for col in col_letters:
        set_cell(ws, f"{col}{row}", None)


# -------------------------------------------------------
# CORE: Fill one sheet (Proforma or Tax) from invoice data
# -------------------------------------------------------
def _fill_proforma_sheet(ws, invoice: dict, project: dict, items: list):
    """
    Fill Sheet 1 (Proforma Invoice) of the combined template.

    Cell map (from user spec):
      A5  - Contractor (client_name)
      C11 - Consultant
      C12 - Client Name
      C13 - Plot No.
      C15 - Project Name
      I4  - Invoice Number (Proforma)
      I5  - Date
      I6  - LP No.
      Rows 18-34, cols A/B/D/I/J/K - line items
        A  - Report No. (of the report covering this test/sample)
        B  - Report creation date (dd-mm-yy)
        D  - Description
        I  - Quantity
        J  - Rate
        K  - Amount
      K35 - Subtotal
      K36 - VAT (5%)
      K37 - Grand Total
      C38 - Amount in words
    """
    from openpyxl.styles import Font, Alignment
    from datetime import date as date_type

    def fmt_date(d):
        if not d:
            return " - "
        if isinstance(d, str):
            return d
        if hasattr(d, "strftime"):
            return d.strftime("%d-%b-%Y")
        return str(d)

    def fmt_date_short(d):
        """dd-mm-yy, as requested for the report creation date in column B."""
        if not d:
            return " - "
        if isinstance(d, str):
            return d
        if hasattr(d, "strftime"):
            return d.strftime("%d-%m-%y")
        return str(d)

    # Header
    set_cell(ws, "A5",  project.get("client_name") or " - ")      # Contractor
    set_cell(ws, "C11", project.get("consultant") or " - ")        # Consultant
    set_cell(ws, "C12", project.get("client_name") or " - ")       # Client Name
    set_cell(ws, "C13", project.get("plot_no") or " - ")           # Plot No.
    set_cell(ws, "C15", project.get("project_name") or " - ")      # Project Name
    set_cell(ws, "I4",  invoice.get("invoice_no") or " - ")        # Invoice No.
    set_cell(ws, "I5",  fmt_date(invoice.get("invoice_date")))      # Date
    set_cell(ws, "I6",  project.get("project_no") or invoice.get("lpo_reference") or " - ")  # LP No.

    FIRST_ROW = 18
    LAST_TEMPLATE_ROW = 34
    ITEM_COLS = ["A", "B", "D", "I", "J", "K"]

    # Clear all item rows first
    for r in range(FIRST_ROW, LAST_TEMPLATE_ROW + 1):
        clear_row_cols(ws, r, ITEM_COLS)

    # If more items than template rows, insert extra rows (copy formatting from row 34)
    num_items = len(items)
    if num_items > (LAST_TEMPLATE_ROW - FIRST_ROW + 1):
        extra = num_items - (LAST_TEMPLATE_ROW - FIRST_ROW + 1)
        ws.insert_rows(LAST_TEMPLATE_ROW + 1, amount=extra)
        from copy import copy
        for i in range(extra):
            new_row = LAST_TEMPLATE_ROW + 1 + i
            for col in range(1, 14):
                src = ws.cell(row=LAST_TEMPLATE_ROW, column=col)
                tgt = ws.cell(row=new_row, column=col)
                tgt.font = copy(src.font)
                tgt.border = copy(src.border)
                tgt.fill = copy(src.fill)
                tgt.number_format = src.number_format
                tgt.alignment = copy(src.alignment)

    # Fill item rows
    # NOTE: Report No. / Report Date are only printed on the FIRST row of a
    # given report's items to avoid repeating the same value down every line
    # (matches the merged-cell look of the printed proforma). Rows belonging
    # to the same report are assumed contiguous, since items are ordered by
    # item_id which preserves report grouping.
    prev_report_no = None
    for idx, item in enumerate(items):
        row = FIRST_ROW + idx
        report_no  = item.get("report_no") or " - "
        report_dt  = fmt_date_short(item.get("report_created_at"))
        desc       = item.get("description") or " - "
        qty        = item.get("quantity") or 0
        rate       = float(item.get("unit_rate") or 0)
        amount     = float(item.get("amount") or qty * rate)

        is_new_report = (report_no != prev_report_no)
        set_cell(ws, f"A{row}", report_no if is_new_report else None)
        set_cell(ws, f"B{row}", report_dt if is_new_report else None)
        prev_report_no = report_no

        set_cell(ws, f"D{row}", desc)
        set_cell(ws, f"I{row}", qty)
        set_cell(ws, f"J{row}", rate)
        set_cell(ws, f"K{row}", round(amount, 2))

    # Totals
    subtotal = float(invoice.get("subtotal", 0))
    vat      = float(invoice.get("vat", 0))
    total    = float(invoice.get("total", 0))

    set_cell(ws, "K35", round(subtotal, 2))
    set_cell(ws, "K36", round(vat, 2))
    set_cell(ws, "K37", round(total, 2))
    set_cell(ws, "C38", invoice.get("amount_in_words") or " - ")


def _fill_tax_sheet(ws, invoice: dict, project: dict, items: list):
    """
    Fill Sheet 2 (Tax Invoice - renamed 'Tax-1') of the combined template.

    Cell map (from user spec):
      A5  - Contractor (client_name, from clients.name)
      C11 - Consultant
      C12 - Client Name
      C13 - Plot No.
      C15 - Project Name
      K4  - Invoice Number (Tax)
      K5  - Date
      K6  - LP No. (this is the project_no from the projects table, NOT the LPO number)
      Rows 18-34, cols A/B/D/I/J/K/L/M - line items
        A  - Report No. (of the report covering this test/sample)
        B  - Report creation date (dd-mm-yy)
        D  - Description
        I  - Quantity
        J  - Rate
        K  - Amount Excl. VAT
        L  - VAT Amount
        M  - Amount Incl. VAT
      M35 - Total Excl. VAT  (=SUM(K18:Kn))
      M36 - Total VAT        (=SUM(L18:Ln))
      M37 - Grand Total Incl VAT (=SUM(M18:Mn))
      D38 - Amount in words
    """
    from openpyxl.styles import Font, Alignment
    from datetime import date as date_type

    def fmt_date(d):
        if not d:
            return " - "
        if isinstance(d, str):
            return d
        if hasattr(d, "strftime"):
            return d.strftime("%d-%b-%Y")
        return str(d)

    def fmt_date_short(d):
        """dd-mm-yy, as requested for the report creation date in column B."""
        if not d:
            return " - "
        if isinstance(d, str):
            return d
        if hasattr(d, "strftime"):
            return d.strftime("%d-%m-%y")
        return str(d)

    # Header
    set_cell(ws, "A5",  project.get("contractor") or " - ")
    set_cell(ws, "C11", project.get("consultant") or " - ")
    set_cell(ws, "C12", project.get("client_name") or " - ")
    set_cell(ws, "C13", project.get("plot_no") or " - ")
    set_cell(ws, "C15", project.get("project_name") or " - ")
    set_cell(ws, "K4",  invoice.get("tax_invoice_no") or invoice.get("invoice_no") or " - ")
    set_cell(ws, "K5",  fmt_date(invoice.get("invoice_date")))
    set_cell(ws, "K6",  project.get("project_no") or " - ")

    FIRST_ROW = 18
    LAST_TEMPLATE_ROW = 34
    ITEM_COLS = ["A", "B", "D", "I", "J", "K", "L", "M"]

    # Clear all item rows first
    for r in range(FIRST_ROW, LAST_TEMPLATE_ROW + 1):
        clear_row_cols(ws, r, ITEM_COLS)

    # Insert extra rows if needed
    num_items = len(items)
    if num_items > (LAST_TEMPLATE_ROW - FIRST_ROW + 1):
        extra = num_items - (LAST_TEMPLATE_ROW - FIRST_ROW + 1)
        ws.insert_rows(LAST_TEMPLATE_ROW + 1, amount=extra)
        from copy import copy
        for i in range(extra):
            new_row = LAST_TEMPLATE_ROW + 1 + i
            for col in range(1, 14):
                src = ws.cell(row=LAST_TEMPLATE_ROW, column=col)
                tgt = ws.cell(row=new_row, column=col)
                tgt.font = copy(src.font)
                tgt.border = copy(src.border)
                tgt.fill = copy(src.fill)
                tgt.number_format = src.number_format
                tgt.alignment = copy(src.alignment)

    # Fill item rows
    # NOTE: Report No. / Report Date are only printed on the FIRST row of a
    # given report's items to avoid repeating the same value down every line
    # (matches the merged-cell look of the printed tax invoice). Rows
    # belonging to the same report are assumed contiguous, since items are
    # ordered by item_id which preserves report grouping.
    prev_report_no = None
    for idx, item in enumerate(items):
        row = FIRST_ROW + idx
        report_no  = item.get("report_no") or " - "
        report_dt  = fmt_date_short(item.get("report_created_at"))
        desc       = item.get("description") or " - "
        qty        = item.get("quantity") or 0
        rate       = float(item.get("unit_rate") or 0)
        excl_vat   = float(item.get("amount") or qty * rate)
        vat_amt    = round(excl_vat * 0.05, 2)
        incl_vat   = round(excl_vat + vat_amt, 2)

        is_new_report = (report_no != prev_report_no)
        set_cell(ws, f"A{row}", report_no if is_new_report else None)
        set_cell(ws, f"B{row}", report_dt if is_new_report else None)
        prev_report_no = report_no

        set_cell(ws, f"D{row}", desc)
        set_cell(ws, f"I{row}", qty)
        set_cell(ws, f"J{row}", rate)
        set_cell(ws, f"K{row}", round(excl_vat, 2))
        set_cell(ws, f"L{row}", vat_amt)
        set_cell(ws, f"M{row}", incl_vat)

    # Totals
    subtotal = float(invoice.get("subtotal", 0))
    vat      = float(invoice.get("vat", 0))
    total    = float(invoice.get("total", 0))

    set_cell(ws, "M35", round(subtotal, 2))   # Total Excl. VAT
    set_cell(ws, "M36", round(vat, 2))         # Total VAT
    set_cell(ws, "M37", round(total, 2))        # Grand Total Incl. VAT
    set_cell(ws, "D38", invoice.get("amount_in_words") or " - ")


# -------------------------------------------------------
# MAIN: Generate combined Proforma + Tax workbook
# -------------------------------------------------------
@router.get("/{invoice_id}/excel-combined")
def generate_excel_invoice_combined(invoice_id: int):
    """
    Generate a combined Proforma + Tax Invoice workbook (2 sheets).
    Sheet 1 = Proforma Invoice, Sheet 2 = Tax Invoice (Tax-1).
    Uses the invoices.xlsx 2-sheet template from Supabase.
    """
    import re, urllib.parse, os, traceback
    from fastapi.responses import FileResponse

    template_path = download_template_from_supabase("invoices_combined")
    if not os.path.exists(template_path):
        raise HTTPException(status_code=404, detail="Combined invoice template not found.")

    conn = get_connection()
    cur = conn.cursor()

    try:
        # 1. Load full invoice
        invoice = get_invoice_complete(invoice_id, cur)
        invoice_type = invoice.get("invoice_type", "")
        project_details = invoice.get("project_details", {})

        # 2. Enrich project_details with consultant / plot_no from projects table.
        #    NOTE: client_name is intentionally NOT overwritten here — it already
        #    comes from clients.name via get_invoice_complete, which is the
        #    correct company/client name. projects.client_name is a separate,
        #    unrelated free-text column and was previously clobbering the
        #    correct value (showing client_id-like text on the invoice).
        cur.execute("""
            SELECT consultant, plot_no
            FROM projects
            WHERE project_id = %s
        """, (invoice.get("project_id"),))
        proj_extra = cur.fetchone()
        if proj_extra:
            project_details["consultant"]  = proj_extra[0] or " - "
            project_details["plot_no"]     = proj_extra[1] or " - "

        # 3. Items for the workbook.
        #    invoice["items"] (from get_invoice_complete) already reflects
        #    exactly the samples/tests belonging to the reports the user
        #    selected when the invoice was created — no further filtering
        #    or report-matching needed here.
        items = invoice.get("items", [])

        # 3b. Attach the governing report's report_no / created_at to each
        #     item, for display in columns A/B of the Tax-1 sheet.
        item_sample_ids = [it.get("sample_id") for it in items if it.get("sample_id")]
        report_info_by_sample = get_report_info_for_samples(item_sample_ids, cur)
        for it in items:
            info = report_info_by_sample.get(it.get("sample_id"))
            it["report_no"] = info["report_no"] if info else None
            it["report_created_at"] = info["created_at"] if info else None

        # 4. Generate tax invoice number (TAX numbering sequence)
        tax_invoice_no = generate_invoice_no(cur, "TAX")
        # Don't commit yet - we're generating only, not persisting a new invoice row.
        # Just use it for the Excel label.
        invoice["tax_invoice_no"] = tax_invoice_no

        # 5. Load the 2-sheet workbook
        wb = openpyxl.load_workbook(template_path, data_only=False)

        if len(wb.worksheets) < 2:
            raise HTTPException(
                status_code=500,
                detail="Combined template must have 2 sheets (Proforma + Tax). "
                       "Please upload the correct invoices.xlsx to Supabase."
            )

        ws_proforma = wb.worksheets[0]   # Sheet 1
        ws_tax      = wb.worksheets[1]   # Sheet 2

        # 6. Fill both sheets
        _fill_proforma_sheet(ws_proforma, invoice, project_details, items)
        _fill_tax_sheet(ws_tax, invoice, project_details, items)

        # 7. Save & return
        output_dir = "generated_invoices"
        os.makedirs(output_dir, exist_ok=True)

        invoice_no        = invoice.get("invoice_no", "invoice").replace("/", "-")
        project_name      = project_details.get("project_name", "")
        lpo_ref           = invoice.get("lpo_reference", "")

        def _clean(text):
            if not text:
                return ""
            text = re.sub(r'[\\/*?:"<>|]', '-', text)
            return re.sub(r'\s+', '-', text).strip('- ')

        if lpo_ref and lpo_ref not in (" - ", ""):
            filename = f"{invoice_no}-{_clean(project_name)}-{_clean(lpo_ref)}-Combined.xlsx"
        else:
            filename = f"{invoice_no}-{_clean(project_name)}-Combined.xlsx"

        output_path = os.path.join(output_dir, f"{invoice_no}-combined.xlsx")
        wb.save(output_path)

        encoded = urllib.parse.quote(filename)
        return FileResponse(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}; filename=\"{filename}\""
            }
        )

    except Exception as e:
        print(f"ERROR in generate_excel_invoice_combined: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()

# =====================================================
# WALK-IN LP: Combined Proforma + Tax Invoice
#
# Endpoint: POST /invoices/walkin/generate-invoice
#
# Flow:
#   1. Validate project is is_walk_in = TRUE and has an LP number
#   2. Resolve which walk_in_items to include
#   3. Calculate totals (subtotal / VAT 5% / grand total)
#   4. Generate PROFORMA invoice number + TAX invoice number
#   5. Insert invoice header + items into invoices / invoice_items tables
#   6. Fill the 2-sheet combined workbook (invoices.xlsx template)
#      PAGE 1 (Proforma):
#        A5  = walk_in_client (Contractor)
#        B7  = walk_in_phone
#        C11 = consultant
#        C12 = client_name   (Client / Owner)
#        C13 = plot_no
#        C15 = project_name
#        I4  = proforma invoice number
#        I5  = date
#        I6  = project_no (LP number)
#        D18:D34 = description, I = qty, J = rate, K = amount
#        K35 = subtotal, K36 = VAT, K37 = grand total, C38 = words
#      PAGE 2 (Tax-1):
#        same header layout but K4/K5/K6 instead of I4/I5/I6
#        D/I/J/K/L/M line items with VAT split
#        M35/M36/M37 = totals, D38 = words
#   7. Return FileResponse (xlsx download)
# =====================================================


class WalkInInvoiceRequest(BaseModel):
    project_id: int
    payment_method: Optional[Literal['CASH', 'CREDIT']] = 'CASH'
    selected_item_ids: Optional[List[int]] = None
    include_all_items: bool = True


@router.post("/walkin/generate-invoice")
def generate_walkin_invoice(payload: WalkInInvoiceRequest):
    """
    Generate a combined Proforma + Tax Invoice workbook for a Walk-In LP.
    Uses walk_in_items directly — no test_requests / samples / reports involved.
    """
    import re as _re
    import urllib.parse as _urlparse

    conn = get_connection()
    cur = conn.cursor()

    try:
        # ──────────────────────────────────────────────────────────
        # 1. Load walk-in project details
        # ──────────────────────────────────────────────────────────
        cur.execute("""
            SELECT project_id, project_no, walk_in_client, walk_in_phone,
                   consultant, plot_no, client_name, project_name
            FROM projects
            WHERE project_id = %s AND is_walk_in = TRUE
        """, (payload.project_id,))
        proj = cur.fetchone()
        if not proj:
            raise HTTPException(404, "Walk-in LP not found")

        (project_id, project_no, walk_in_client, walk_in_phone,
         consultant, plot_no, client_name, project_name) = proj

        if not project_no or project_no == "PENDING":
            raise HTTPException(400, "LP number has not been generated for this walk-in yet")

        contractor    = walk_in_client or project_name or "—"
        phone         = walk_in_phone or ""
        consultant    = consultant or ""
        plot_no       = plot_no or ""
        client_field  = client_name or ""   # Client/Owner → C12
        proj_name     = project_name or contractor

        # ──────────────────────────────────────────────────────────
        # 2. Resolve items
        #
        # BUSINESS RULE: Only tests that have a Payment Advice
        # (pa_generated = TRUE) can be invoiced. Tests without a PA
        # are not yet invoiceable and must be excluded regardless of
        # what the frontend sends.
        # ──────────────────────────────────────────────────────────
        if payload.include_all_items:
            # "Select All" from the UI = all ADVISED items for this LP
            cur.execute("""
                SELECT item_id, description, test_standard, unit_rate, quantity, amount
                FROM walk_in_items
                WHERE project_id = %s AND pa_generated = TRUE
                ORDER BY item_id
            """, (project_id,))
        else:
            if not payload.selected_item_ids:
                raise HTTPException(400, "No tests selected")
            # Explicit selection — enforce pa_generated = TRUE as a safety guard
            cur.execute("""
                SELECT item_id, description, test_standard, unit_rate, quantity, amount
                FROM walk_in_items
                WHERE project_id = %s
                  AND item_id = ANY(%s)
                  AND pa_generated = TRUE
                ORDER BY item_id
            """, (project_id, payload.selected_item_ids))

        rows = cur.fetchall()
        if not rows:
            raise HTTPException(
                400,
                "No invoiceable tests found. Only tests with a Payment Advice (Already Advised) "
                "can be included on an invoice. Please generate a Payment Advice first."
            )

        items = [
            {
                "item_id":      r[0],
                "description":  r[1] or "—",
                "test_standard": r[2] or "",
                "unit_rate":    float(r[3]) if r[3] is not None else 0.0,
                "quantity":     r[4] or 1,
                "amount":       float(r[5]) if r[5] is not None else 0.0,
            }
            for r in rows
        ]

        # ──────────────────────────────────────────────────────────
        # 3. Totals
        # ──────────────────────────────────────────────────────────
        subtotal    = sum(i["amount"] for i in items)
        vat         = round(subtotal * 0.05, 2)
        grand_total = round(subtotal + vat, 2)
        words       = number_to_words(grand_total)

        # ──────────────────────────────────────────────────────────
        # 4. Invoice numbers
        # ──────────────────────────────────────────────────────────
        proforma_no = generate_invoice_no(cur, "PROFORMA")
        tax_no      = generate_invoice_no(cur, "TAX")

        payment_terms = "30 days" if payload.payment_method == "CREDIT" else "Immediate"

        # ──────────────────────────────────────────────────────────
        # 5. Persist invoice header (walk-in projects have no client_id /
        #    quotation_id so we use the PROFORMA invoice row only; the Tax
        #    number is recorded in the remarks field for traceability).
        # ──────────────────────────────────────────────────────────
        cur.execute("""
            INSERT INTO invoices (
                invoice_no, project_id, invoice_type, payment_method, invoice_date,
                lpo_reference, payment_terms,
                subtotal, vat, total, amount_in_words,
                services_description, remarks, payment_status
            )
            VALUES (%s, %s, 'PROFORMA', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'UNPAID')
            RETURNING invoice_id
        """, (
            proforma_no,
            project_id,
            payload.payment_method,
            date.today(),
            project_no,          # LP number stored as lpo_reference for walk-ins
            payment_terms,
            subtotal,
            vat,
            grand_total,
            words,
            f"Testing services — {proj_name}",
            f"Tax Invoice No: {tax_no}",
        ))
        invoice_id = cur.fetchone()[0]

        # Insert invoice items (one row per walk_in_item)
        for item in items:
            cur.execute("""
                INSERT INTO invoice_items (
                    invoice_id, description, test_standard,
                    unit_rate, quantity, amount
                )
                VALUES (%s, %s, %s, %s, %s, %s)
            """, (
                invoice_id,
                item["description"],
                item["test_standard"],
                item["unit_rate"],
                item["quantity"],
                item["amount"],
            ))

        conn.commit()

        # ──────────────────────────────────────────────────────────
        # 6. Fill the 2-sheet Excel workbook
        # ──────────────────────────────────────────────────────────
        template_path = download_template_from_supabase("invoices_combined")
        if not os.path.exists(template_path):
            raise HTTPException(404, "Combined invoice template not found. "
                                     "Upload invoices.xlsx to Supabase templates/invoices/.")

        wb = openpyxl.load_workbook(template_path, data_only=False)

        if len(wb.worksheets) < 2:
            raise HTTPException(
                500,
                "Combined template must have 2 sheets (Proforma + Tax). "
                "Please upload the correct invoices.xlsx to Supabase."
            )

        ws_proforma = wb.worksheets[0]
        ws_tax      = wb.worksheets[1]

        today_str = date.today().strftime("%d-%b-%Y")

        # ── Helper already defined at module level as set_cell() ──────────

        # ── PAGE 1: Proforma Invoice ──────────────────────────────────────
        _wi_fill_proforma(ws_proforma, {
            "contractor":   contractor,
            "phone":        phone,
            "consultant":   consultant,
            "client_name":  client_field,
            "plot_no":      plot_no,
            "project_name": proj_name,
            "invoice_no":   proforma_no,
            "date":         today_str,
            "lp_number":    project_no,
            "subtotal":     subtotal,
            "vat":          vat,
            "grand_total":  grand_total,
            "words":        words,
        }, items)

        # ── PAGE 2: Tax Invoice (Tax-1) ───────────────────────────────────
        _wi_fill_tax(ws_tax, {
            "contractor":   contractor,
            "phone":        phone,
            "consultant":   consultant,
            "client_name":  client_field,
            "plot_no":      plot_no,
            "project_name": proj_name,
            "invoice_no":   tax_no,
            "date":         today_str,
            "lp_number":    project_no,
            "subtotal":     subtotal,
            "vat":          vat,
            "grand_total":  grand_total,
            "words":        words,
        }, items)

        # ──────────────────────────────────────────────────────────
        # 7. Save & return
        # ──────────────────────────────────────────────────────────
        output_dir = "generated_invoices"
        os.makedirs(output_dir, exist_ok=True)

        def _clean(text):
            if not text:
                return ""
            text = _re.sub(r'[\\/*?":<>|]', '-', text)
            return _re.sub(r'\s+', '-', text).strip('- ')

        proforma_no_hyphen = proforma_no.replace('/', '-')
        filename    = f"{_clean(contractor)}.xlsx"
        output_path = os.path.join(output_dir, f"{proforma_no_hyphen}-walkin.xlsx")
        wb.save(output_path)

        encoded = _urlparse.quote(filename)
        from fastapi.responses import FileResponse as _FR
        return _FR(
            output_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded}; filename=\"{filename}\""
            }
        )

    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        print(f"ERROR in generate_walkin_invoice: {e}")
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error generating walk-in invoice: {str(e)}")
    finally:
        cur.close()
        conn.close()


# ──────────────────────────────────────────────────────────────────────────────
# INTERNAL HELPERS: fill Proforma and Tax sheets for walk-in invoices
# (separate from _fill_proforma_sheet / _fill_tax_sheet which are used for
#  standard report-based invoices — different header cell layout)
# ──────────────────────────────────────────────────────────────────────────────

def _wi_fill_proforma(ws, meta: dict, items: list):
    """
    Fill Sheet 1 (Proforma Invoice) for a Walk-In LP.

    Header cell map:
      A5  = Contractor (walk_in_client)
      B7  = Phone
      C11 = Consultant
      C12 = Client/Owner name  (client_name column)
      C13 = Plot No.
      C15 = Project Name
      I4  = Proforma Invoice Number
      I5  = Date
      I6  = LP No. (project_no)

    Line items (rows 18–34, extendable):
      D = Description
      I = Quantity
      J = Rate
      K = Amount (= Qty × Rate)

    Summary:
      K35 = Subtotal
      K36 = VAT (5%)
      K37 = Grand Total
      C38 = Amount in Words
    """
    FIRST_ROW         = 18
    LAST_TEMPLATE_ROW = 34
    ITEM_COLS         = ["D", "I", "J", "K"]

    # Header
    set_cell(ws, "A5",  meta.get("contractor") or "—")
    set_cell(ws, "B7",  meta.get("phone") or "")
    set_cell(ws, "C11", meta.get("consultant") or "")
    set_cell(ws, "C12", meta.get("client_name") or "")
    set_cell(ws, "C13", meta.get("plot_no") or "")
    set_cell(ws, "C15", meta.get("project_name") or "")
    set_cell(ws, "I4",  meta.get("invoice_no") or "")
    set_cell(ws, "I5",  meta.get("date") or "")
    set_cell(ws, "I6",  meta.get("lp_number") or "")

    # Clear existing item rows
    for r in range(FIRST_ROW, LAST_TEMPLATE_ROW + 1):
        clear_row_cols(ws, r, ITEM_COLS)

    # Insert extra rows if more items than template allows
    num_items = len(items)
    template_slots = LAST_TEMPLATE_ROW - FIRST_ROW + 1
    if num_items > template_slots:
        extra = num_items - template_slots
        ws.insert_rows(LAST_TEMPLATE_ROW + 1, amount=extra)
        from copy import copy as _copy
        for i in range(extra):
            new_row = LAST_TEMPLATE_ROW + 1 + i
            for col in range(1, 14):
                src = ws.cell(row=LAST_TEMPLATE_ROW, column=col)
                tgt = ws.cell(row=new_row, column=col)
                tgt.font        = _copy(src.font)
                tgt.border      = _copy(src.border)
                tgt.fill        = _copy(src.fill)
                tgt.number_format = src.number_format
                tgt.alignment   = _copy(src.alignment)

    # Fill item rows
    for idx, item in enumerate(items):
        row = FIRST_ROW + idx
        set_cell(ws, f"D{row}", item.get("description") or "—")
        set_cell(ws, f"I{row}", item.get("quantity") or 0)
        set_cell(ws, f"J{row}", item.get("unit_rate") or 0)
        set_cell(ws, f"K{row}", f'=IF(I{row}="","",(I{row}*J{row}))')

    # Blank remaining IF formulas inside K35 SUM range
    for row in range(FIRST_ROW + num_items, LAST_TEMPLATE_ROW + 1):
        set_cell(ws, f"K{row}", f'=IF(I{row}="","",(I{row}*J{row}))')

    # Totals
    set_cell(ws, "K35", f"=SUM(K{FIRST_ROW}:K{LAST_TEMPLATE_ROW})")
    set_cell(ws, "K36", "=K35*5%")
    set_cell(ws, "K37", "=K35+K36")
    set_cell(ws, "C38", meta.get("words") or "")


def _wi_fill_tax(ws, meta: dict, items: list):
    """
    Fill Sheet 2 (Tax Invoice – 'Tax-1') for a Walk-In LP.

    Header cell map:
      A5  = Contractor (walk_in_client)
      B7  = Phone
      C11 = Consultant
      C12 = Client/Owner name
      C13 = Plot No.
      C15 = Project Name
      K4  = Tax Invoice Number
      K5  = Date
      K6  = LP No. (project_no)

    Line items (rows 18–34, extendable):
      D = Description
      I = Quantity
      J = Rate
      K = Amount Excl. VAT  (= Qty × Rate)
      L = VAT Amount        (= K × 5%)
      M = Amount Incl. VAT  (= K + L)

    Summary:
      M35 = Total Excl. VAT  (=SUM(K18:K34))
      M36 = Total VAT        (=SUM(L18:L34))
      M37 = Grand Total      (=SUM(M18:M34))
      D38 = Amount in Words
    """
    FIRST_ROW         = 18
    LAST_TEMPLATE_ROW = 34
    ITEM_COLS         = ["D", "I", "J", "K", "L", "M"]

    # Header
    set_cell(ws, "A5",  meta.get("contractor") or "—")
    set_cell(ws, "B7",  meta.get("phone") or "")
    set_cell(ws, "C11", meta.get("consultant") or "")
    set_cell(ws, "C12", meta.get("client_name") or "")
    set_cell(ws, "C13", meta.get("plot_no") or "")
    set_cell(ws, "C15", meta.get("project_name") or "")
    set_cell(ws, "K4",  meta.get("invoice_no") or "")
    set_cell(ws, "K5",  meta.get("date") or "")
    set_cell(ws, "K6",  meta.get("lp_number") or "")

    # Clear existing item rows
    for r in range(FIRST_ROW, LAST_TEMPLATE_ROW + 1):
        clear_row_cols(ws, r, ITEM_COLS)

    # Insert extra rows if needed
    num_items = len(items)
    template_slots = LAST_TEMPLATE_ROW - FIRST_ROW + 1
    if num_items > template_slots:
        extra = num_items - template_slots
        ws.insert_rows(LAST_TEMPLATE_ROW + 1, amount=extra)
        from copy import copy as _copy
        for i in range(extra):
            new_row = LAST_TEMPLATE_ROW + 1 + i
            for col in range(1, 14):
                src = ws.cell(row=LAST_TEMPLATE_ROW, column=col)
                tgt = ws.cell(row=new_row, column=col)
                tgt.font        = _copy(src.font)
                tgt.border      = _copy(src.border)
                tgt.fill        = _copy(src.fill)
                tgt.number_format = src.number_format
                tgt.alignment   = _copy(src.alignment)

    # Fill item rows with formulas for VAT split
    for idx, item in enumerate(items):
        row = FIRST_ROW + idx
        set_cell(ws, f"D{row}", item.get("description") or "—")
        set_cell(ws, f"I{row}", item.get("quantity") or 0)
        set_cell(ws, f"J{row}", item.get("unit_rate") or 0)
        set_cell(ws, f"K{row}", f'=IF(I{row}="","",(I{row}*J{row}))')
        set_cell(ws, f"L{row}", f'=IF(K{row}="","",K{row}*5%)')
        set_cell(ws, f"M{row}", f'=IF(K{row}="","",K{row}+L{row})')

    # Blank remaining formula rows
    for row in range(FIRST_ROW + num_items, LAST_TEMPLATE_ROW + 1):
        set_cell(ws, f"K{row}", f'=IF(I{row}="","",(I{row}*J{row}))')
        set_cell(ws, f"L{row}", f'=IF(K{row}="","",K{row}*5%)')
        set_cell(ws, f"M{row}", f'=IF(K{row}="","",K{row}+L{row})')

    # Totals
    set_cell(ws, "M35", f"=SUM(K{FIRST_ROW}:K{LAST_TEMPLATE_ROW})")
    set_cell(ws, "M36", f"=SUM(L{FIRST_ROW}:L{LAST_TEMPLATE_ROW})")
    set_cell(ws, "M37", f"=SUM(M{FIRST_ROW}:M{LAST_TEMPLATE_ROW})")
    set_cell(ws, "D38", meta.get("words") or "")



# ============================================================================
# NEW ENDPOINTS — paste these into invoices.py
#
# Changes summary:
#   1. New Pydantic models for the 3-mode requests
#   2. DB migration helper (run once at startup or manually)
#   3. POST /walkin/generate-proforma-only
#   4. GET  /walkin/pending-proformas/{project_id}
#   5. POST /walkin/generate-tax-only
#   6. Updated POST /generate-with-reports-and-tests  (replaces old version)
#      — handles PROFORMA_ONLY and TAX_ONLY for non-walk-in LPs as well
#
# Required NEW table (run once):
#   CREATE TABLE IF NOT EXISTS walkin_proforma_records (
#       id                 SERIAL PRIMARY KEY,
#       project_id         INTEGER NOT NULL,
#       invoice_id         INTEGER NOT NULL REFERENCES invoices(invoice_id),
#       proforma_no        TEXT    NOT NULL,
#       item_ids           INTEGER[] NOT NULL,          -- walk_in_items.item_id list
#       tax_invoice_id     INTEGER REFERENCES invoices(invoice_id),
#       tax_invoice_no     TEXT,
#       created_at         TIMESTAMP DEFAULT NOW()
#   );
#
# Required NEW column on invoices (run once):
#   ALTER TABLE invoices ADD COLUMN IF NOT EXISTS generation_mode TEXT
#       CHECK (generation_mode IN ('BOTH', 'PROFORMA_ONLY', 'TAX_ONLY'));
# ============================================================================

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Optional, List, Literal
from datetime import date, datetime
from decimal import Decimal
import traceback, os, re as _re
import urllib.parse as _urlparse

# ── these are already imported in invoices.py; kept here for clarity ─────────
# from db import get_connection
# from utils import resource_path
# import openpyxl, requests, tempfile
# from fastapi.responses import FileResponse


# ============================================================================
# NEW Pydantic models
# ============================================================================

class WalkInProformaOnlyRequest(BaseModel):
    """Option B — Proforma Only (walk-in LP)"""
    project_id:         int
    payment_method:     Optional[Literal['CASH', 'CREDIT']] = 'CASH'
    selected_item_ids:  Optional[List[int]] = None   # None / empty → all advised
    include_all_items:  bool = True


class WalkInTaxOnlyRequest(BaseModel):
    """Option C — Tax Only (walk-in LP)"""
    project_id:         int
    proforma_record_id: int          # walkin_proforma_records.id chosen by user
    payment_method:     Optional[Literal['CASH', 'CREDIT']] = 'CASH'


class NonWalkInInvoiceRequest(BaseModel):
    """Unified request for non-walk-in LPs (Options A / B / C)"""
    project_id:          int
    generation_mode:     Literal['BOTH', 'PROFORMA_ONLY', 'TAX_ONLY']
    payment_method:      Optional[Literal['CASH', 'CREDIT']] = 'CASH'
    include_all_reports: bool = True
    selected_report_ids: Optional[List[int]] = None
    # Only required for TAX_ONLY
    proforma_invoice_id: Optional[int] = None        # invoices.invoice_id of the existing proforma
    services_description: Optional[str] = None


# ============================================================================
# DB migration helper — call once
# ============================================================================

def ensure_walkin_proforma_records_table(cur):
    cur.execute("""
        CREATE TABLE IF NOT EXISTS walkin_proforma_records (
            id               SERIAL PRIMARY KEY,
            project_id       INTEGER NOT NULL,
            invoice_id       INTEGER NOT NULL,
            proforma_no      TEXT    NOT NULL,
            item_ids         INTEGER[] NOT NULL,
            tax_invoice_id   INTEGER,
            tax_invoice_no   TEXT,
            created_at       TIMESTAMP DEFAULT NOW()
        )
    """)
    # Add generation_mode column to invoices if not present
    cur.execute("""
        ALTER TABLE invoices
        ADD COLUMN IF NOT EXISTS generation_mode TEXT
            CHECK (generation_mode IN ('BOTH', 'PROFORMA_ONLY', 'TAX_ONLY'))
    """)
    # Ensure invoice_report_links table exists — used by _link_reports and
    # list_nonwi_pending_proformas.  Without this guard the SELECT in
    # list_nonwi_pending_proformas crashes with "relation does not exist"
    # (surfaced as "tuple index out of range" at the HTTP layer).
    cur.execute("""
        CREATE TABLE IF NOT EXISTS invoice_report_links (
            id           SERIAL PRIMARY KEY,
            invoice_id   INTEGER NOT NULL,
            report_no    TEXT    NOT NULL,
            invoice_type TEXT    NOT NULL,
            created_at   TIMESTAMP DEFAULT NOW(),
            UNIQUE (invoice_id, report_no, invoice_type)
        )
    """)


# ============================================================================
# HELPER: build walk-in project meta dict from a project row
# ============================================================================

def _wi_project_meta(proj_row, proforma_no, tax_no, subtotal, vat, grand_total, words, today_str):
    """Return the two meta dicts expected by _wi_fill_proforma / _wi_fill_tax."""
    (project_id, project_no, walk_in_client, walk_in_phone,
     consultant, plot_no, client_name, project_name) = proj_row

    contractor   = walk_in_client or project_name or "—"
    client_field = client_name or ""
    proj_name    = project_name or contractor

    base = dict(
        contractor   = contractor,
        phone        = walk_in_phone or "",
        consultant   = consultant or "",
        client_name  = client_field,
        plot_no      = plot_no or "",
        project_name = proj_name,
        date         = today_str,
        lp_number    = project_no,
        subtotal     = subtotal,
        vat          = vat,
        grand_total  = grand_total,
        words        = words,
    )
    proforma_meta = {**base, "invoice_no": proforma_no}
    tax_meta      = {**base, "invoice_no": tax_no}
    return proforma_meta, tax_meta, proj_name


# ============================================================================
# HELPER: blank-out Sheet 2 (Tax-1) keeping template structure intact
# ============================================================================

def _wi_blank_tax_sheet(ws):
    """
    Leave Sheet 2 (Tax Invoice) completely unfilled — just clear any
    data cells that might carry over from a previous fill.
    Header cells (invoice number, date, line items, totals, words) are cleared.
    Template borders / formatting remain intact.
    """
    # Header number / date / LP cells
    for addr in ("K4", "K5", "K6", "A5", "B7", "C11", "C12", "C13", "C15", "D38"):
        set_cell(ws, addr, None)

    FIRST_ROW = 18
    LAST_ROW  = 34
    for r in range(FIRST_ROW, LAST_ROW + 1):
        clear_row_cols(ws, r, ["D", "I", "J", "K", "L", "M"])

    for addr in ("M35", "M36", "M37"):
        set_cell(ws, addr, None)


def _nonwi_blank_tax_sheet(ws):
    """Same idea for non-walk-in Sheet 2 (different header cells)."""
    for addr in ("K4", "K5", "K6", "A5", "C11", "C12", "C13", "C15", "D38"):
        set_cell(ws, addr, None)

    FIRST_ROW = 18
    LAST_ROW  = 34
    for r in range(FIRST_ROW, LAST_ROW + 1):
        clear_row_cols(ws, r, ["A", "B", "D", "I", "J", "K", "L", "M"])

    for addr in ("M35", "M36", "M37"):
        set_cell(ws, addr, None)


# ============================================================================
# OPTION B — Walk-In: Generate Proforma Only
# POST /invoices/walkin/generate-proforma-only
# ============================================================================

@router.post("/walkin/generate-proforma-only")
def generate_walkin_proforma_only(payload: WalkInProformaOnlyRequest):
    """
    Option B for walk-in LPs.
    Output: 2-sheet workbook — Sheet 1 filled, Sheet 2 LEFT BLANK.
    Persists a PROFORMA invoice row + walkin_proforma_records entry.
    """
    conn = get_connection()
    cur  = conn.cursor()

    try:
        ensure_walkin_proforma_records_table(cur)

        # 1. Load project
        cur.execute("""
            SELECT project_id, project_no, walk_in_client, walk_in_phone,
                   consultant, plot_no, client_name, project_name
            FROM projects
            WHERE project_id = %s AND is_walk_in = TRUE
        """, (payload.project_id,))
        proj = cur.fetchone()
        if not proj:
            raise HTTPException(404, "Walk-in LP not found")
        if not proj[1] or proj[1] == "PENDING":
            raise HTTPException(400, "LP number has not been generated yet")

        # 2. Resolve items (pa_generated = TRUE guard)
        if payload.include_all_items or not payload.selected_item_ids:
            cur.execute("""
                SELECT item_id, description, test_standard, unit_rate, quantity, amount
                FROM walk_in_items
                WHERE project_id = %s AND pa_generated = TRUE
                ORDER BY item_id
            """, (payload.project_id,))
        else:
            cur.execute("""
                SELECT item_id, description, test_standard, unit_rate, quantity, amount
                FROM walk_in_items
                WHERE project_id = %s
                  AND item_id = ANY(%s)
                  AND pa_generated = TRUE
                ORDER BY item_id
            """, (payload.project_id, payload.selected_item_ids))

        rows = cur.fetchall()
        if not rows:
            raise HTTPException(
                400,
                "No invoiceable tests found. Only tests with a Payment Advice can be invoiced."
            )

        items = [
            {
                "item_id":       r[0],
                "description":   r[1] or "—",
                "test_standard": r[2] or "",
                "unit_rate":     float(r[3]) if r[3] is not None else 0.0,
                "quantity":      r[4] or 1,
                "amount":        float(r[5]) if r[5] is not None else 0.0,
            }
            for r in rows
        ]
        item_ids = [i["item_id"] for i in items]

        # 3. Totals
        subtotal    = sum(i["amount"] for i in items)
        vat         = round(subtotal * 0.05, 2)
        grand_total = round(subtotal + vat, 2)
        words       = number_to_words(grand_total)

        # 4. Proforma number only (no tax number yet)
        proforma_no   = generate_invoice_no(cur, "PROFORMA")
        payment_terms = "30 days" if payload.payment_method == "CREDIT" else "Immediate"
        today         = date.today()
        today_str     = today.strftime("%d-%b-%Y")
        project_no    = proj[1]

        # 5. Persist invoice row (PROFORMA only)
        cur.execute("""
            INSERT INTO invoices (
                invoice_no, project_id, invoice_type, payment_method, invoice_date,
                lpo_reference, payment_terms,
                subtotal, vat, total, amount_in_words,
                services_description, remarks, payment_status, generation_mode
            )
            VALUES (%s, %s, 'PROFORMA', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'UNPAID', 'PROFORMA_ONLY')
            RETURNING invoice_id
        """, (
            proforma_no, payload.project_id, payload.payment_method, today,
            project_no, payment_terms,
            subtotal, vat, grand_total, words,
            f"Testing services — {proj[7] or proj[2] or ''}",
            "Proforma Invoice — Tax Invoice pending",
        ))
        invoice_id = cur.fetchone()[0]

        # Invoice items
        for item in items:
            cur.execute("""
                INSERT INTO invoice_items (
                    invoice_id, description, test_standard,
                    unit_rate, quantity, amount
                ) VALUES (%s, %s, %s, %s, %s, %s)
            """, (invoice_id, item["description"], item["test_standard"],
                  item["unit_rate"], item["quantity"], item["amount"]))

        # walkin_proforma_records
        cur.execute("""
            INSERT INTO walkin_proforma_records
                (project_id, invoice_id, proforma_no, item_ids)
            VALUES (%s, %s, %s, %s)
            RETURNING id
        """, (payload.project_id, invoice_id, proforma_no, item_ids))
        record_id = cur.fetchone()[0]

        conn.commit()

        # 6. Build Excel workbook (Sheet 1 filled, Sheet 2 blank)
        template_path = download_template_from_supabase("invoices_combined")
        wb = openpyxl.load_workbook(template_path, data_only=False)
        if len(wb.worksheets) < 2:
            raise HTTPException(500, "Combined template must have 2 sheets")

        ws_proforma = wb.worksheets[0]
        ws_tax      = wb.worksheets[1]

        proforma_meta, _, proj_name = _wi_project_meta(
            proj, proforma_no, "", subtotal, vat, grand_total, words, today_str
        )
        _wi_fill_proforma(ws_proforma, proforma_meta, items)
        _wi_blank_tax_sheet(ws_tax)          # Sheet 2 intentionally unfilled

        # 7. Save & return
        output_dir = "generated_invoices"
        os.makedirs(output_dir, exist_ok=True)

        def _clean(t):
            if not t: return ""
            t = _re.sub(r'[\\/*?":<>|]', '-', t)
            return _re.sub(r'\s+', '-', t).strip('- ')

        pno_h    = proforma_no.replace('/', '-')
        filename = f"{pno_h}_{_clean(proj[2] or proj[7] or '')}.xlsx"
        out_path = os.path.join(output_dir, f"{pno_h}-walkin-proforma.xlsx")
        wb.save(out_path)

        encoded = _urlparse.quote(filename)
        from fastapi.responses import FileResponse as _FR
        return _FR(
            out_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":
                     f"attachment; filename*=UTF-8''{encoded}; filename=\"{filename}\""}
        )

    except HTTPException:
        conn.rollback(); raise
    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        raise HTTPException(500, f"Error generating proforma: {str(e)}")
    finally:
        cur.close(); conn.close()


# ============================================================================
# LIST PENDING PROFORMAS — Walk-In LP
# GET /invoices/walkin/pending-proformas/{project_id}
# ============================================================================

@router.get("/walkin/pending-proformas/{project_id}")
def list_walkin_pending_proformas(project_id: int):
    """
    Return all Proforma-Only records for this walk-in LP that have NOT yet
    had a Tax Invoice generated.  Used to populate the "Choose Proforma"
    selector in Tax Only mode.
    """
    conn = get_connection()
    cur  = conn.cursor()
    try:
        ensure_walkin_proforma_records_table(cur)

        cur.execute("""
            SELECT
                wpr.id,
                wpr.proforma_no,
                wpr.invoice_id,
                wpr.item_ids,
                wpr.created_at,
                i.subtotal,
                i.vat,
                i.total,
                i.payment_method
            FROM walkin_proforma_records wpr
            JOIN invoices i ON wpr.invoice_id = i.invoice_id
            WHERE wpr.project_id = %s
              AND wpr.tax_invoice_id IS NULL       -- not yet converted
            ORDER BY wpr.created_at DESC
        """, (project_id,))

        rows = cur.fetchall()
        result = []

        for row in rows:
            (rec_id, proforma_no, invoice_id, item_ids,
             created_at, subtotal, vat, total, payment_method) = row

            # Fetch the item descriptions so the UI can show the test list
            if item_ids:
                cur.execute("""
                    SELECT item_id, description, test_standard, unit_rate, quantity, amount
                    FROM walk_in_items
                    WHERE item_id = ANY(%s)
                    ORDER BY item_id
                """, (item_ids,))
                items = [
                    {
                        "item_id":       r[0],
                        "description":   r[1] or "—",
                        "test_standard": r[2] or "",
                        "unit_rate":     float(r[3]) if r[3] is not None else 0.0,
                        "quantity":      r[4] or 1,
                        "amount":        float(r[5]) if r[5] is not None else 0.0,
                    }
                    for r in cur.fetchall()
                ]
            else:
                items = []

            result.append({
                "record_id":      rec_id,
                "proforma_no":    proforma_no,
                "invoice_id":     invoice_id,
                "created_at":     created_at.strftime("%d-%b-%Y") if created_at else None,
                "subtotal":       float(subtotal) if subtotal else 0.0,
                "vat":            float(vat) if vat else 0.0,
                "total":          float(total) if total else 0.0,
                "payment_method": payment_method or "CASH",
                "items":          items,
            })

        return {"project_id": project_id, "pending_proformas": result}

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))
    finally:
        cur.close(); conn.close()


# ============================================================================
# OPTION C — Walk-In: Generate Tax Only
# POST /invoices/walkin/generate-tax-only
# ============================================================================

@router.post("/walkin/generate-tax-only")
def generate_walkin_tax_only(payload: WalkInTaxOnlyRequest):
    """
    Option C for walk-in LPs.
    Loads an existing Proforma record, generates a new Tax Invoice number,
    fills BOTH sheets (Sheet 1 = existing proforma data, Sheet 2 = new tax).
    Marks the proforma record as converted.
    """
    conn = get_connection()
    cur  = conn.cursor()

    try:
        ensure_walkin_proforma_records_table(cur)

        # 1. Load the proforma record
        cur.execute("""
            SELECT wpr.id, wpr.project_id, wpr.invoice_id, wpr.proforma_no,
                   wpr.item_ids, wpr.tax_invoice_id
            FROM walkin_proforma_records wpr
            WHERE wpr.id = %s
        """, (payload.proforma_record_id,))
        rec = cur.fetchone()
        if not rec:
            raise HTTPException(404, "Proforma record not found")

        rec_id, project_id, proforma_invoice_id, proforma_no, item_ids, existing_tax_id = rec

        if existing_tax_id:
            raise HTTPException(
                400,
                f"This Proforma ({proforma_no}) has already been converted to a Tax Invoice."
            )
        if project_id != payload.project_id:
            raise HTTPException(400, "Proforma record does not belong to this project")

        # 2. Load project meta
        cur.execute("""
            SELECT project_id, project_no, walk_in_client, walk_in_phone,
                   consultant, plot_no, client_name, project_name
            FROM projects
            WHERE project_id = %s AND is_walk_in = TRUE
        """, (project_id,))
        proj = cur.fetchone()
        if not proj:
            raise HTTPException(404, "Walk-in LP not found")

        # 3. Load original proforma invoice for financials
        cur.execute("""
            SELECT subtotal, vat, total, amount_in_words, payment_method
            FROM invoices WHERE invoice_id = %s
        """, (proforma_invoice_id,))
        inv_row = cur.fetchone()
        if not inv_row:
            raise HTTPException(404, "Original proforma invoice record not found")

        subtotal    = float(inv_row[0]) if inv_row[0] is not None else 0.0
        vat         = float(inv_row[1]) if inv_row[1] is not None else 0.0
        grand_total = float(inv_row[2]) if inv_row[2] is not None else 0.0
        words       = inv_row[3] or number_to_words(grand_total)
        orig_method = inv_row[4] or payload.payment_method

        # 4. Load original items from walk_in_items
        cur.execute("""
            SELECT item_id, description, test_standard, unit_rate, quantity, amount
            FROM walk_in_items
            WHERE item_id = ANY(%s)
            ORDER BY item_id
        """, (item_ids,))
        items = [
            {
                "item_id":       r[0],
                "description":   r[1] or "—",
                "test_standard": r[2] or "",
                "unit_rate":     float(r[3]) if r[3] is not None else 0.0,
                "quantity":      r[4] or 1,
                "amount":        float(r[5]) if r[5] is not None else 0.0,
            }
            for r in cur.fetchall()
        ]

        # 5. Generate Tax Invoice number
        tax_no        = generate_invoice_no(cur, "TAX")
        payment_terms = "30 days" if payload.payment_method == "CREDIT" else "Immediate"
        today         = date.today()
        today_str     = today.strftime("%d-%b-%Y")
        project_no    = proj[1]
        proj_name     = proj[7] or proj[2] or "—"

        # 6. Persist Tax Invoice row
        cur.execute("""
            INSERT INTO invoices (
                invoice_no, project_id, invoice_type, payment_method, invoice_date,
                lpo_reference, payment_terms,
                subtotal, vat, total, amount_in_words,
                services_description, remarks, payment_status, generation_mode
            )
            VALUES (%s, %s, 'TAX', %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'UNPAID', 'TAX_ONLY')
            RETURNING invoice_id
        """, (
            tax_no, project_id, payload.payment_method, today,
            project_no, payment_terms,
            subtotal, vat, grand_total, words,
            f"Testing services — {proj_name}",
            f"Proforma Invoice No: {proforma_no}",
        ))
        tax_invoice_id = cur.fetchone()[0]

        for item in items:
            cur.execute("""
                INSERT INTO invoice_items (
                    invoice_id, description, test_standard,
                    unit_rate, quantity, amount
                ) VALUES (%s, %s, %s, %s, %s, %s)
            """, (tax_invoice_id, item["description"], item["test_standard"],
                  item["unit_rate"], item["quantity"], item["amount"]))

        # 7. Mark proforma record as converted
        cur.execute("""
            UPDATE walkin_proforma_records
            SET tax_invoice_id = %s, tax_invoice_no = %s
            WHERE id = %s
        """, (tax_invoice_id, tax_no, rec_id))

        # 8. Update original proforma's remarks to cross-reference the tax no
        cur.execute("""
            UPDATE invoices SET remarks = %s
            WHERE invoice_id = %s
        """, (f"Tax Invoice No: {tax_no}", proforma_invoice_id))

        conn.commit()

        # 9. Build Excel workbook — both sheets filled
        template_path = download_template_from_supabase("invoices_combined")
        wb = openpyxl.load_workbook(template_path, data_only=False)
        if len(wb.worksheets) < 2:
            raise HTTPException(500, "Combined template must have 2 sheets")

        ws_proforma = wb.worksheets[0]
        ws_tax      = wb.worksheets[1]

        # Fetch the ORIGINAL proforma date for Sheet 1
        cur.execute("SELECT invoice_date FROM invoices WHERE invoice_id = %s", (proforma_invoice_id,))
        orig_date_row = cur.fetchone()
        orig_date_str = (orig_date_row[0].strftime("%d-%b-%Y")
                         if orig_date_row and orig_date_row[0] else today_str)

        proforma_meta, tax_meta, _ = _wi_project_meta(
            proj, proforma_no, tax_no, subtotal, vat, grand_total, words, orig_date_str
        )
        # Tax sheet uses today's date
        tax_meta["date"] = today_str

        _wi_fill_proforma(ws_proforma, proforma_meta, items)
        _wi_fill_tax(ws_tax, tax_meta, items)

        # 10. Save & return
        output_dir = "generated_invoices"
        os.makedirs(output_dir, exist_ok=True)

        def _clean(t):
            if not t: return ""
            t = _re.sub(r'[\\/*?":<>|]', '-', t)
            return _re.sub(r'\s+', '-', t).strip('- ')

        pno_h    = proforma_no.replace('/', '-')
        tno_h    = tax_no.replace('/', '-')
        filename = f"{tno_h}_{_clean(proj[2] or proj[7] or '')}.xlsx"
        out_path = os.path.join(output_dir, f"{pno_h}-{tno_h}-walkin.xlsx")
        wb.save(out_path)

        encoded = _urlparse.quote(filename)
        from fastapi.responses import FileResponse as _FR
        return _FR(
            out_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":
                     f"attachment; filename*=UTF-8''{encoded}; filename=\"{filename}\""}
        )

    except HTTPException:
        conn.rollback(); raise
    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        raise HTTPException(500, f"Error generating tax invoice: {str(e)}")
    finally:
        cur.close(); conn.close()


# ============================================================================
# UPDATED /generate-with-reports-and-tests
# Handles BOTH / PROFORMA_ONLY / TAX_ONLY for NON-walk-in LPs
# Replaces the existing @router.post("/generate-with-reports-and-tests")
# ============================================================================

@router.post("/generate-with-reports-and-tests-v2")
def generate_invoice_with_reports_and_tests_v2(payload: NonWalkInInvoiceRequest):
    """
    Unified endpoint for non-walk-in LP invoice generation.

    generation_mode:
      BOTH          → creates PROFORMA invoice, generates both sheets filled
      PROFORMA_ONLY → creates PROFORMA invoice, Sheet 1 filled, Sheet 2 blank
      TAX_ONLY      → requires proforma_invoice_id; creates TAX invoice,
                       Sheet 1 = original proforma data, Sheet 2 = new tax
    """
    conn = get_connection()
    cur  = conn.cursor()

    try:
        ensure_walkin_proforma_records_table(cur)

        project_id      = payload.project_id
        generation_mode = payload.generation_mode
        payment_method  = payload.payment_method or "CASH"

        # Reject walk-in projects
        cur.execute("SELECT is_walk_in FROM projects WHERE project_id = %s", (project_id,))
        proj_row = cur.fetchone()
        if not proj_row:
            raise HTTPException(404, "Project not found")
        if proj_row[0]:
            raise HTTPException(400,
                "Walk-in projects use the /walkin/ endpoints (proforma-only / tax-only).")

        today = date.today()

        # ── Shared: resolve report → sample scope ────────────────────────────
        def _resolve_reports(invoice_type_str):
            """
            Returns (report_nos, sample_ids_filter).
            For TAX_ONLY we exclude reports already converted from proforma.
            """
            inc_all = payload.include_all_reports
            sel_ids = payload.selected_report_ids

            if inc_all:
                # All approved reports not yet linked to this invoice_type
                cur.execute("""
                    SELECT DISTINCT r.report_id, r.report_no
                    FROM reports r
                    JOIN samples s ON r.sample_id = s.sample_id
                    JOIN test_requests tr ON s.request_id = tr.test_request_id
                    WHERE tr.project_id = %s
                      AND r.status = 'APPROVED'
                      AND r.report_no NOT IN (
                          SELECT report_no FROM invoice_report_links
                          WHERE invoice_type = %s
                      )
                """, (project_id, invoice_type_str))
            elif sel_ids:
                cur.execute("""
                    SELECT DISTINCT r.report_id, r.report_no
                    FROM reports r
                    WHERE r.report_id = ANY(%s)
                      AND r.status = 'APPROVED'
                """, (sel_ids,))
            else:
                return [], []

            rows        = cur.fetchall()
            rpt_ids     = [r[0] for r in rows]
            rpt_nos     = [r[1] for r in rows]
            samp_filter = get_sample_ids_for_reports(rpt_ids, cur)
            return rpt_nos, samp_filter

        # ────────────────────────────────────────────────────────────────────
        # MODE: BOTH
        # ────────────────────────────────────────────────────────────────────
        if generation_mode == "BOTH":
            report_nos, sample_ids_filter = _resolve_reports("PROFORMA")
            if not sample_ids_filter:
                raise HTTPException(400, "No uninvoiced reports selected.")

            inv_payload = InvoiceCreate(
                project_id          = project_id,
                invoice_type        = "PROFORMA",
                payment_method      = payment_method,
                invoice_date        = today,
                payment_terms       = "30 days" if payment_method == "CREDIT" else "Immediate",
                services_description= payload.services_description or "Professional services rendered",
            )
            invoice_result = _create_invoice_with_payment_method_impl(inv_payload, sample_ids=sample_ids_filter, shared_cur=cur)
            invoice_id = invoice_result["invoice_id"]

            # Record report links + generation mode
            _link_reports(cur, invoice_id, report_nos, "PROFORMA")
            cur.execute("UPDATE invoices SET generation_mode='BOTH' WHERE invoice_id=%s", (invoice_id,))

            # Commit to release the row lock before generating the TAX number
            conn.commit()

            # Now safe to read invoices table for the next TAX number
            tax_no = generate_invoice_no(cur, "TAX")
            invoice_result["tax_invoice_no"] = tax_no

            # Persist the tax number on the proforma row's remarks. No separate
            # TAX invoice row is created for BOTH mode, but the regenerate-download
            # logic (and the department revenue view) both look for
            # "Tax Invoice No: ..." in remarks, so it must be saved here.
            cur.execute("UPDATE invoices SET remarks = %s WHERE invoice_id = %s",
                        (f"Tax Invoice No: {tax_no}", invoice_id))
            conn.commit()

            # Build Excel — both sheets
            return _build_nonwi_combined_excel(invoice_id, invoice_result, cur,
                                               fill_tax=True, mode="BOTH")

        # ────────────────────────────────────────────────────────────────────
        # MODE: PROFORMA_ONLY
        # ────────────────────────────────────────────────────────────────────
        elif generation_mode == "PROFORMA_ONLY":
            report_nos, sample_ids_filter = _resolve_reports("PROFORMA")
            if not sample_ids_filter:
                raise HTTPException(400, "No uninvoiced reports selected.")

            inv_payload = InvoiceCreate(
                project_id          = project_id,
                invoice_type        = "PROFORMA",
                payment_method      = payment_method,
                invoice_date        = today,
                payment_terms       = "30 days" if payment_method == "CREDIT" else "Immediate",
                services_description= payload.services_description or "Professional services rendered",
            )
            invoice_result = _create_invoice_with_payment_method_impl(inv_payload, sample_ids=sample_ids_filter, shared_cur=cur)
            invoice_id = invoice_result["invoice_id"]

            _link_reports(cur, invoice_id, report_nos, "PROFORMA")
            cur.execute("UPDATE invoices SET generation_mode='PROFORMA_ONLY' WHERE invoice_id=%s", (invoice_id,))

            # Commit before building Excel
            conn.commit()

            return _build_nonwi_combined_excel(invoice_id, invoice_result, cur,
                                               fill_tax=False, mode="PROFORMA_ONLY")

        # ────────────────────────────────────────────────────────────────────
        # MODE: TAX_ONLY
        # ────────────────────────────────────────────────────────────────────
        elif generation_mode == "TAX_ONLY":
            if not payload.proforma_invoice_id:
                raise HTTPException(400, "proforma_invoice_id is required for TAX_ONLY mode.")

            # Load the existing proforma invoice
            proforma_invoice_id = payload.proforma_invoice_id
            cur.execute("""
                SELECT invoice_no, subtotal, vat, total, amount_in_words,
                       payment_method, invoice_date, generation_mode
                FROM invoices
                WHERE invoice_id = %s AND invoice_type = 'PROFORMA'
            """, (proforma_invoice_id,))
            pfm = cur.fetchone()
            if not pfm:
                raise HTTPException(404, "Proforma invoice not found.")
            if pfm[7] not in (None, "BOTH", "PROFORMA_ONLY"):
                # Already a TAX_ONLY row itself — shouldn't happen
                raise HTTPException(400, "Selected invoice is not a Proforma.")

            # Validate it hasn't been converted already
            cur.execute("""
                SELECT COUNT(*) FROM invoices
                WHERE project_id = %s
                  AND invoice_type = 'TAX'
                  AND remarks LIKE %s
            """, (project_id, f"%{pfm[0]}%"))
            if cur.fetchone()[0] > 0:
                raise HTTPException(400,
                    f"Proforma {pfm[0]} has already been converted to a Tax Invoice.")

            proforma_no    = pfm[0]
            subtotal       = float(pfm[1]) if pfm[1] else 0.0
            vat            = float(pfm[2]) if pfm[2] else 0.0
            grand_total    = float(pfm[3]) if pfm[3] else 0.0
            words          = pfm[4] or number_to_words(grand_total)
            orig_method    = pfm[5] or payment_method
            proforma_date  = pfm[6]

            # Get the report links from the original proforma
            cur.execute("""
                SELECT report_no FROM invoice_report_links
                WHERE invoice_id = %s AND invoice_type = 'PROFORMA'
            """, (proforma_invoice_id,))
            proforma_report_nos = [r[0] for r in cur.fetchall()]

            # Resolve samples for the tax invoice (restrict to proforma's reports)
            cur.execute("""
                SELECT r.report_id FROM reports r
                WHERE r.report_no = ANY(%s)
            """, (proforma_report_nos,))
            pfm_report_ids  = [r[0] for r in cur.fetchall()]
            sample_ids_filter = get_sample_ids_for_reports(pfm_report_ids, cur)

            if not sample_ids_filter:
                raise HTTPException(400,
                    "Could not resolve samples from the selected proforma's reports.")

            # Create Tax Invoice — generate number first, then insert on shared cursor
            tax_no = generate_invoice_no(cur, "TAX")
            inv_payload = InvoiceCreate(
                project_id          = project_id,
                invoice_type        = "TAX",
                payment_method      = payment_method,
                invoice_date        = today,
                payment_terms       = "30 days" if payment_method == "CREDIT" else "Immediate",
                services_description= payload.services_description or "Professional services rendered",
                remarks             = f"Proforma Invoice No: {proforma_no}",
            )
            tax_invoice_result = _create_invoice_with_payment_method_impl(inv_payload, sample_ids=sample_ids_filter, shared_cur=cur)
            tax_invoice_id = tax_invoice_result["invoice_id"]

            # Link same reports under TAX type + mark mode
            _link_reports(cur, tax_invoice_id, proforma_report_nos, "TAX")
            cur.execute("UPDATE invoices SET generation_mode='TAX_ONLY' WHERE invoice_id=%s", (tax_invoice_id,))

            # Update proforma remarks to cross-reference tax no
            cur.execute("""
                UPDATE invoices SET remarks = %s WHERE invoice_id = %s
            """, (f"Tax Invoice No: {tax_no}", proforma_invoice_id))
            conn.commit()

            # Build Excel — both sheets
            # Sheet 1: original proforma data, Sheet 2: new tax data
            return _build_nonwi_tax_only_excel(
                proforma_invoice_id  = proforma_invoice_id,
                proforma_no          = proforma_no,
                proforma_date        = proforma_date,
                tax_invoice_id       = tax_invoice_id,
                tax_no               = tax_no,
                cur                  = cur,
            )

    except HTTPException:
        conn.rollback(); raise
    except Exception as e:
        conn.rollback()
        traceback.print_exc()
        raise HTTPException(500, f"Error generating invoice: {str(e)}")
    finally:
        cur.close(); conn.close()


# ── Internal helper: insert invoice_report_links ─────────────────────────────

def _link_reports(cur, invoice_id: int, report_nos: list, invoice_type: str):
    for rno in report_nos:
        try:
            cur.execute("""
                INSERT INTO invoice_report_links (invoice_id, report_no, invoice_type)
                VALUES (%s, %s, %s)
                ON CONFLICT DO NOTHING
            """, (invoice_id, rno, invoice_type))
        except Exception as e:
            print(f"WARNING: Could not link report {rno}: {e}")


# ── Internal helper: build combined excel for non-walk-in ────────────────────

def _build_nonwi_combined_excel(invoice_id: int, invoice_result: dict,
                                 cur, fill_tax: bool, mode: str):
    """
    Builds the 2-sheet workbook for non-walk-in BOTH / PROFORMA_ONLY modes.
    Calls the existing _fill_proforma_sheet / _fill_tax_sheet helpers.
    """
    import re, urllib.parse, os
    from fastapi.responses import FileResponse

    template_path = download_template_from_supabase("invoices_combined")
    wb = openpyxl.load_workbook(template_path, data_only=False)
    if len(wb.worksheets) < 2:
        raise HTTPException(500, "Combined template must have 2 sheets")

    ws_proforma = wb.worksheets[0]
    ws_tax      = wb.worksheets[1]

    invoice        = invoice_result
    project_details = invoice.get("project_details", {})

    # Enrich with consultant / plot_no
    cur.execute("SELECT consultant, plot_no FROM projects WHERE project_id=%s",
                (invoice.get("project_id"),))
    ex = cur.fetchone()
    if ex:
        project_details["consultant"] = ex[0] or " - "
        project_details["plot_no"]    = ex[1] or " - "

    # Attach report info to items
    items = invoice.get("items", [])
    sample_ids = [it.get("sample_id") for it in items if it.get("sample_id")]
    report_info = get_report_info_for_samples(sample_ids, cur)
    for it in items:
        info = report_info.get(it.get("sample_id"))
        it["report_no"]         = info["report_no"]  if info else None
        it["report_created_at"] = info["created_at"] if info else None

    _fill_proforma_sheet(ws_proforma, invoice, project_details, items)

    if fill_tax:
        _fill_tax_sheet(ws_tax, invoice, project_details, items)
    else:
        _nonwi_blank_tax_sheet(ws_tax)

    output_dir = "generated_invoices"
    os.makedirs(output_dir, exist_ok=True)

    inv_no      = invoice.get("invoice_no", "invoice").replace("/", "-")
    client_name = project_details.get("client_name", "")
    proj_name   = project_details.get("project_name", "")
    contractor  = client_name or proj_name

    def _cl(t):
        if not t: return ""
        t = re.sub(r'[\\/*?:"<>|]', '-', t)
        return re.sub(r'\s+', ' ', t).strip()

    suffix = "Combined" if fill_tax else "Proforma"

    # Naming convention:
    # mode == "BOTH"          -> {contractor_name}.xlsx
    # mode == "PROFORMA_ONLY" -> {invoice_no} {contractor_name}.xlsx
    if mode == "BOTH":
        filename = f"{_cl(contractor)}.xlsx"
    else:  # PROFORMA_ONLY
        filename = f"{inv_no}_{_cl(contractor)}.xlsx"

    out_path  = os.path.join(output_dir, f"{inv_no}-{suffix.lower()}.xlsx")
    wb.save(out_path)

    enc = urllib.parse.quote(filename)
    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename*=UTF-8''{enc}; filename=\"{filename}\""}
    )


def _build_nonwi_tax_only_excel(proforma_invoice_id: int, proforma_no: str,
                                  proforma_date, tax_invoice_id: int,
                                  tax_no: str, cur):
    """
    Builds the 2-sheet workbook for TAX_ONLY mode.
    Sheet 1 = proforma data (original date + proforma no).
    Sheet 2 = new tax invoice data.
    """
    import re, urllib.parse, os
    from fastapi.responses import FileResponse

    template_path = download_template_from_supabase("invoices_combined")
    wb = openpyxl.load_workbook(template_path, data_only=False)
    if len(wb.worksheets) < 2:
        raise HTTPException(500, "Combined template must have 2 sheets")

    ws_proforma = wb.worksheets[0]
    ws_tax      = wb.worksheets[1]

    # Load proforma invoice from DB
    proforma_inv = get_invoice_complete(proforma_invoice_id, cur)
    proforma_project = proforma_inv.get("project_details", {})

    # Enrich
    cur.execute("SELECT consultant, plot_no FROM projects WHERE project_id=%s",
                (proforma_inv.get("project_id"),))
    ex = cur.fetchone()
    if ex:
        proforma_project["consultant"] = ex[0] or " - "
        proforma_project["plot_no"]    = ex[1] or " - "

    pfm_items = proforma_inv.get("items", [])
    pfm_sample_ids = [it.get("sample_id") for it in pfm_items if it.get("sample_id")]
    pfm_report_info = get_report_info_for_samples(pfm_sample_ids, cur)
    for it in pfm_items:
        info = pfm_report_info.get(it.get("sample_id"))
        it["report_no"]         = info["report_no"]  if info else None
        it["report_created_at"] = info["created_at"] if info else None

    # Load tax invoice from DB
    tax_inv = get_invoice_complete(tax_invoice_id, cur)
    tax_project = tax_inv.get("project_details", {})
    cur.execute("SELECT consultant, plot_no FROM projects WHERE project_id=%s",
                (tax_inv.get("project_id"),))
    ex2 = cur.fetchone()
    if ex2:
        tax_project["consultant"] = ex2[0] or " - "
        tax_project["plot_no"]    = ex2[1] or " - "

    tax_items = tax_inv.get("items", [])
    tax_sample_ids = [it.get("sample_id") for it in tax_items if it.get("sample_id")]
    tax_report_info = get_report_info_for_samples(tax_sample_ids, cur)
    for it in tax_items:
        info = tax_report_info.get(it.get("sample_id"))
        it["report_no"]         = info["report_no"]  if info else None
        it["report_created_at"] = info["created_at"] if info else None

    # Override tax_invoice_no on the tax object so Sheet 2 prints the correct number
    tax_inv["tax_invoice_no"] = tax_no

    _fill_proforma_sheet(ws_proforma, proforma_inv, proforma_project, pfm_items)
    _fill_tax_sheet(ws_tax, tax_inv, tax_project, tax_items)

    output_dir = "generated_invoices"
    os.makedirs(output_dir, exist_ok=True)

    pfm_h       = proforma_no.replace("/", "-")
    tax_h       = tax_no.replace("/", "-")
    client_name = proforma_project.get("client_name", "")
    proj_name   = proforma_project.get("project_name", "")
    contractor  = client_name or proj_name

    def _cl(t):
        if not t: return ""
        t = re.sub(r'[\\/*?:"<>|]', '-', t)
        return re.sub(r'\s+', ' ', t).strip()

    # TAX_ONLY naming: {tax_invoice_no}_{contractor_name}.xlsx
    filename = f"{tax_h}_{_cl(contractor)}.xlsx"
    out_path = os.path.join(output_dir, f"{pfm_h}-{tax_h}-taxonly.xlsx")
    wb.save(out_path)

    enc = urllib.parse.quote(filename)
    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition":
                 f"attachment; filename*=UTF-8''{enc}; filename=\"{filename}\""}
    )
# ============================================================================
# NEW: List pending proformas for NON-walk-in LPs (Tax Only selector)
# GET /invoices/projects/{project_id}/pending-proformas
# ============================================================================

@router.get("/projects/{project_id}/pending-proformas")
def list_nonwi_pending_proformas(project_id: int):
    """
    Returns PROFORMA invoices for this project that have NOT yet had a
    TAX invoice generated from them.
    Used to populate the "Choose Proforma" selector in Tax Only mode.
    """
    conn = get_connection()
    cur  = conn.cursor()
    try:
        # Ensure generation_mode column exists before querying it
        ensure_walkin_proforma_records_table(cur)
        conn.commit()

        # Find PROFORMA invoices where no TAX invoice remarks reference them.
        # Include generation_mode IS NULL so proformas created before the column
        # was added are not silently excluded.
        cur.execute("""
            SELECT
                i.invoice_id,
                i.invoice_no,
                i.invoice_date,
                i.subtotal,
                i.vat,
                i.total,
                i.payment_method,
                i.generation_mode
            FROM invoices i
            WHERE i.project_id = %s
              AND i.invoice_type = 'PROFORMA'
              AND (i.generation_mode IN ('PROFORMA_ONLY', 'BOTH') OR i.generation_mode IS NULL)
              AND NOT EXISTS (
                  SELECT 1 FROM invoices t
                  WHERE t.project_id = %s
                    AND t.invoice_type = 'TAX'
                    AND t.remarks LIKE '%%' || i.invoice_no || '%%'
              )
            ORDER BY i.invoice_date DESC, i.invoice_id DESC
        """, (project_id, project_id))

        rows = cur.fetchall()
        result = []
        for row in rows:
            inv_id, inv_no, inv_date, subtotal, vat, total, pay_method, mode = row

            # Get linked reports for display
            cur.execute("""
                SELECT report_no FROM invoice_report_links
                WHERE invoice_id = %s AND invoice_type = 'PROFORMA'
                ORDER BY report_no
            """, (inv_id,))
            report_nos = [r[0] for r in cur.fetchall()]

            result.append({
                "invoice_id":     inv_id,
                "proforma_no":    inv_no,
                "invoice_date":   inv_date.strftime("%d-%b-%Y") if inv_date else None,
                "subtotal":       float(subtotal) if subtotal else 0.0,
                "vat":            float(vat) if vat else 0.0,
                "total":          float(total) if total else 0.0,
                "payment_method": pay_method or "CASH",
                "generation_mode": mode,
                "reports":        report_nos,
            })

        return {"project_id": project_id, "pending_proformas": result}

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))
    finally:
        cur.close(); conn.close()


# ============================================================================
# GET /invoices/list
# Returns all invoices with generation_mode, computed contractor name,
# payment status, paid_date, invoice_date — for the Invoice Database view.
# Only invoices with a recognised generation_mode are meaningful here;
# the frontend filters out rows where it can't derive a valid filename.
# ============================================================================

@router.get("/list")
def list_all_invoices():
    """
    Returns a flat list of every invoice row enriched with:
      - contractor_name  (clients.name for non-walk-in, walk_in_client/project_name for walk-in)
      - generation_mode  (BOTH | PROFORMA_ONLY | TAX_ONLY | NULL)
      - payment_status, paid_date, invoice_date
      - linked_invoice_no  (for TAX_ONLY rows: the proforma number embedded in remarks)
      - total        (invoice grand total, for daily/monthly financial summaries)
      - is_walk_in   (True = Walk-In Customer invoice, False = Credit/Non-Walk-In invoice)

    The frontend uses generation_mode + contractor_name to reconstruct the
    downloaded filename and filters out rows that don't match the known
    naming convention (i.e. old / unrelated invoices).
    """
    conn = get_connection()
    cur  = conn.cursor()
    try:
        cur.execute("""
            SELECT
                i.invoice_id,
                i.invoice_no,
                i.invoice_type,
                i.generation_mode,
                i.invoice_date,
                i.payment_status,
                i.paid_date,
                i.remarks,
                i.total,
                p.is_walk_in,
                -- Contractor name: walk-in uses walk_in_client / project_name,
                -- non-walk-in uses clients.name
                CASE
                    WHEN p.is_walk_in = TRUE
                        THEN COALESCE(p.walk_in_client, p.project_name)
                    ELSE
                        COALESCE(c.name, p.client_name)
                END AS contractor_name
            FROM invoices i
            JOIN projects p ON i.project_id = p.project_id
            LEFT JOIN clients c ON p.client_id = c.client_id
            WHERE i.generation_mode IS NOT NULL          -- only formally generated invoices
            ORDER BY i.invoice_id DESC
        """)

        rows = cur.fetchall()
        result = []
        for row in rows:
            (invoice_id, invoice_no, invoice_type, generation_mode,
             invoice_date, payment_status, paid_date, remarks,
             total, is_walk_in,
             contractor_name) = row

            # For TAX_ONLY invoices the proforma number is stored in remarks;
            # surface it so the frontend can show "Linked: 001/25"
            linked_invoice_no = None
            if generation_mode == "TAX_ONLY" and remarks:
                import re as _re
                m = _re.search(r'(\d{3}/\d{2})', remarks)
                if m:
                    linked_invoice_no = m.group(1)

            result.append({
                "invoice_id":       invoice_id,
                "invoice_no":       invoice_no,
                "invoice_type":     invoice_type,
                "generation_mode":  generation_mode,
                "invoice_date":     invoice_date.isoformat() if invoice_date else None,
                "payment_status":   payment_status or "UNPAID",
                "paid_date":        paid_date.isoformat() if paid_date else None,
                "contractor_name":  contractor_name or "",
                "linked_invoice_no": linked_invoice_no,
                "total":            float(total) if total is not None else 0.0,
                "is_walk_in":       bool(is_walk_in),
            })

        return result

    except Exception as e:
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cur.close()
        conn.close()


# ============================================================================
# GET /invoices/{invoice_id}/regenerate-excel
#
# Regenerates the exact Excel file for any invoice from DB data alone.
# No stored files needed. Handles all 6 cases:
#
#   Walk-in    × BOTH          → both sheets filled (_wi_fill_*)
#   Walk-in    × PROFORMA_ONLY → Sheet 1 filled, Sheet 2 blank
#   Walk-in    × TAX_ONLY      → Sheet 1 = orig proforma, Sheet 2 = tax
#   Non-walk-in × BOTH         → both sheets filled (_fill_*_sheet)
#   Non-walk-in × PROFORMA_ONLY→ Sheet 1 filled, Sheet 2 blank
#   Non-walk-in × TAX_ONLY     → Sheet 1 = orig proforma, Sheet 2 = tax
#
# Paste this endpoint into invoices.py, BEFORE the @router.delete("/{invoice_id}") line.
# ============================================================================

@router.get("/{invoice_id}/regenerate-excel")
def regenerate_invoice_excel(invoice_id: int):
    """
    Regenerate the exact Excel file for a stored invoice, reading all data
    fresh from the database. No file storage needed — everything is rebuilt
    from invoice_items, projects, clients, walk_in_items, etc.
    """
    import re as _re, urllib.parse as _urlparse, os, traceback
    from fastapi.responses import FileResponse

    conn = get_connection()
    cur  = conn.cursor()

    try:
        # ── 1. Load core invoice row + project/client info ───────────────────
        cur.execute("""
            SELECT
                i.invoice_id, i.invoice_no, i.invoice_type, i.generation_mode,
                i.invoice_date, i.payment_method, i.payment_terms,
                i.subtotal, i.vat, i.total, i.amount_in_words,
                i.services_description, i.remarks,
                i.project_id,
                p.project_no, p.project_name, p.is_walk_in,
                p.walk_in_client, p.walk_in_phone, p.client_name AS proj_client_name,
                p.consultant, p.plot_no,
                c.name AS client_name, c.contact_person, c.email, c.address, c.phone
            FROM invoices i
            JOIN projects p ON i.project_id = p.project_id
            LEFT JOIN clients c ON p.client_id = c.client_id
            WHERE i.invoice_id = %s
        """, (invoice_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Invoice not found")

        (inv_id, invoice_no, invoice_type, generation_mode,
         invoice_date, payment_method, payment_terms,
         subtotal, vat, total, amount_in_words,
         services_description, remarks,
         project_id,
         project_no, project_name, is_walk_in,
         walk_in_client, walk_in_phone, proj_client_name,
         consultant, plot_no,
         client_name, contact_person, client_email, client_address, client_phone) = row

        subtotal = float(subtotal) if subtotal else 0.0
        vat      = float(vat)      if vat      else 0.0
        total    = float(total)    if total    else 0.0

        def _cl(t):
            if not t: return ""
            t = _re.sub(r'[\\/*?:"<>|]', '-', t)
            return _re.sub(r'\s+', ' ', t).strip()

        inv_no_h = invoice_no.replace("/", "-") if invoice_no else "invoice"

        # ── 2. Load invoice items ────────────────────────────────────────────
        cur.execute("""
            SELECT ii.description, ii.test_standard, ii.unit_rate, ii.quantity, ii.amount, ii.sample_id
            FROM invoice_items ii
            WHERE ii.invoice_id = %s
            ORDER BY ii.item_id
        """, (invoice_id,))
        item_rows = cur.fetchall()
        items = [
            {
                "description":   r[0] or "—",
                "test_standard": r[1] or "",
                "unit_rate":     float(r[2]) if r[2] is not None else 0.0,
                "quantity":      r[3] or 1,
                "amount":        float(r[4]) if r[4] is not None else 0.0,
                "sample_id":     r[5],
            }
            for r in item_rows
        ]

        # For non-walk-in: attach report_no / report_created_at to each item
        if not is_walk_in:
            sample_ids = [it["sample_id"] for it in items if it.get("sample_id")]
            report_info = get_report_info_for_samples(sample_ids, cur)
            for it in items:
                info = report_info.get(it.get("sample_id"))
                it["report_no"]         = info["report_no"]  if info else None
                it["report_created_at"] = info["created_at"] if info else None

        # ── 3. Load template ─────────────────────────────────────────────────
        template_path = download_template_from_supabase("invoices_combined")
        wb = openpyxl.load_workbook(template_path, data_only=False)
        if len(wb.worksheets) < 2:
            raise HTTPException(500, "Combined template must have at least 2 sheets")
        ws_proforma = wb.worksheets[0]
        ws_tax      = wb.worksheets[1]

        output_dir = "generated_invoices"
        os.makedirs(output_dir, exist_ok=True)

        # ════════════════════════════════════════════════════════════════════
        # WALK-IN INVOICES
        # ════════════════════════════════════════════════════════════════════
        if is_walk_in:
            contractor  = walk_in_client or project_name or "—"
            phone       = walk_in_phone or ""
            client_fld  = proj_client_name or ""
            proj_nm     = project_name or contractor

            def fmt_d(d):
                if not d: return ""
                if hasattr(d, "strftime"): return d.strftime("%d-%b-%Y")
                return str(d)

            if generation_mode == "BOTH":
                # Both sheets filled. Sheet 1 = proforma, Sheet 2 = tax.
                # For BOTH the tax number is stored in a sibling TAX invoice's
                # invoice_no whose remarks reference this proforma_no.
                # Look it up:
                cur.execute("""
                    SELECT invoice_no, invoice_date FROM invoices
                    WHERE project_id = %s AND invoice_type = 'TAX'
                      AND remarks LIKE %s
                    ORDER BY invoice_id DESC LIMIT 1
                """, (project_id, f"%{invoice_no}%"))
                tax_row = cur.fetchone()
                tax_no   = tax_row[0] if tax_row else invoice_no
                tax_date = fmt_d(tax_row[1]) if tax_row else fmt_d(invoice_date)

                proforma_meta = dict(
                    contractor=contractor, phone=phone, consultant=consultant or "",
                    client_name=client_fld, plot_no=plot_no or "",
                    project_name=proj_nm, date=fmt_d(invoice_date),
                    lp_number=project_no, invoice_no=invoice_no,
                    subtotal=subtotal, vat=vat, grand_total=total,
                    words=amount_in_words,
                )
                tax_meta = {**proforma_meta, "invoice_no": tax_no, "date": tax_date}

                _wi_fill_proforma(ws_proforma, proforma_meta, items)
                _wi_fill_tax(ws_tax, tax_meta, items)

                filename = f"{_cl(contractor)}.xlsx"
                out_path = os.path.join(output_dir, f"{inv_no_h}-regen-walkin-both.xlsx")

            elif generation_mode == "PROFORMA_ONLY":
                proforma_meta = dict(
                    contractor=contractor, phone=phone, consultant=consultant or "",
                    client_name=client_fld, plot_no=plot_no or "",
                    project_name=proj_nm, date=fmt_d(invoice_date),
                    lp_number=project_no, invoice_no=invoice_no,
                    subtotal=subtotal, vat=vat, grand_total=total,
                    words=amount_in_words,
                )
                _wi_fill_proforma(ws_proforma, proforma_meta, items)
                _wi_blank_tax_sheet(ws_tax)

                filename = f"{inv_no_h}_{_cl(contractor)}.xlsx"
                out_path = os.path.join(output_dir, f"{inv_no_h}-regen-walkin-proforma.xlsx")

            elif generation_mode == "TAX_ONLY":
                # This is a TAX invoice. Find the linked proforma from remarks.
                import re as _re2
                proforma_no = None
                if remarks:
                    m = _re2.search(r'Proforma Invoice No[:\s]+(\S+)', remarks)
                    if m:
                        proforma_no = m.group(1).strip()

                if proforma_no:
                    cur.execute("""
                        SELECT i.invoice_id, i.invoice_no, i.invoice_date,
                               i.subtotal, i.vat, i.total, i.amount_in_words
                        FROM invoices i
                        WHERE i.project_id = %s AND i.invoice_no = %s
                          AND i.invoice_type = 'PROFORMA'
                        LIMIT 1
                    """, (project_id, proforma_no))
                    pfm = cur.fetchone()
                else:
                    pfm = None

                if pfm:
                    # Sheet 1: original proforma data
                    pfm_date = pfm[2]
                    pfm_subtotal = float(pfm[3]) if pfm[3] else subtotal
                    pfm_vat      = float(pfm[4]) if pfm[4] else vat
                    pfm_total    = float(pfm[5]) if pfm[5] else total
                    pfm_words    = pfm[6] or amount_in_words

                    # Load proforma items
                    cur.execute("""
                        SELECT description, test_standard, unit_rate, quantity, amount
                        FROM invoice_items WHERE invoice_id = %s ORDER BY item_id
                    """, (pfm[0],))
                    pfm_items = [
                        {"description": r[0] or "—", "test_standard": r[1] or "",
                         "unit_rate": float(r[2] or 0), "quantity": r[3] or 1,
                         "amount": float(r[4] or 0)}
                        for r in cur.fetchall()
                    ]

                    proforma_meta = dict(
                        contractor=contractor, phone=phone, consultant=consultant or "",
                        client_name=client_fld, plot_no=plot_no or "",
                        project_name=proj_nm,
                        date=pfm_date.strftime("%d-%b-%Y") if hasattr(pfm_date,"strftime") else str(pfm_date or ""),
                        lp_number=project_no, invoice_no=pfm[1],
                        subtotal=pfm_subtotal, vat=pfm_vat, grand_total=pfm_total,
                        words=pfm_words,
                    )
                    tax_meta = dict(
                        contractor=contractor, phone=phone, consultant=consultant or "",
                        client_name=client_fld, plot_no=plot_no or "",
                        project_name=proj_nm,
                        date=invoice_date.strftime("%d-%b-%Y") if hasattr(invoice_date,"strftime") else str(invoice_date or ""),
                        lp_number=project_no, invoice_no=invoice_no,
                        subtotal=subtotal, vat=vat, grand_total=total,
                        words=amount_in_words,
                    )
                    _wi_fill_proforma(ws_proforma, proforma_meta, pfm_items)
                    _wi_fill_tax(ws_tax, tax_meta, items)
                else:
                    # Proforma not found — fill both sheets with what we have
                    meta = dict(
                        contractor=contractor, phone=phone, consultant=consultant or "",
                        client_name=client_fld, plot_no=plot_no or "",
                        project_name=proj_nm,
                        date=invoice_date.strftime("%d-%b-%Y") if hasattr(invoice_date,"strftime") else "",
                        lp_number=project_no, invoice_no=invoice_no,
                        subtotal=subtotal, vat=vat, grand_total=total,
                        words=amount_in_words,
                    )
                    _wi_fill_proforma(ws_proforma, meta, items)
                    _wi_fill_tax(ws_tax, meta, items)

                filename = f"{inv_no_h}_{_cl(contractor)}.xlsx"
                out_path = os.path.join(output_dir, f"{inv_no_h}-regen-walkin-tax.xlsx")

            else:
                raise HTTPException(400, f"Unrecognised generation_mode: {generation_mode}")

        # ════════════════════════════════════════════════════════════════════
        # NON-WALK-IN INVOICES
        # ════════════════════════════════════════════════════════════════════
        else:
            contractor = client_name or proj_client_name or project_name or "—"

            project_details = {
                "project_no":      project_no,
                "project_name":    project_name,
                "client_name":     client_name or proj_client_name or "",
                "contractor":      contractor,
                "client_contact":  contact_person or "",
                "client_email":    client_email or "",
                "client_address":  client_address or "",
                "client_phone":    client_phone or "",
                "consultant":      consultant or " - ",
                "plot_no":         plot_no or " - ",
            }

            invoice_dict = {
                "invoice_id":       inv_id,
                "invoice_no":       invoice_no,
                "invoice_type":     invoice_type,
                "generation_mode":  generation_mode,
                "invoice_date":     invoice_date,
                "payment_method":   payment_method,
                "subtotal":         subtotal,
                "vat":              vat,
                "total":            total,
                "amount_in_words":  amount_in_words,
                "lpo_reference":    project_no,
            }

            if generation_mode == "BOTH":
                # For BOTH, this invoice IS the proforma. The tax number was
                # generated at creation time but NOT stored (it was just printed
                # on the Excel). We can look it up from a sibling TAX invoice
                # whose remarks reference this proforma_no, or regenerate one.
                # Since we don't create a new TAX invoice row on BOTH mode,
                # we just use the next available tax number for display only.
                # Actually: check remarks on this invoice for "Tax Invoice No:"
                tax_no = None
                if remarks:
                    import re as _re3
                    m = _re3.search(r'Tax Invoice No[:\s]+(\S+)', remarks)
                    if m:
                        tax_no = m.group(1).strip()
                if not tax_no:
                    # Try to find a sibling TAX row
                    cur.execute("""
                        SELECT invoice_no FROM invoices
                        WHERE project_id = %s AND invoice_type = 'TAX'
                          AND remarks LIKE %s
                        ORDER BY invoice_id DESC LIMIT 1
                    """, (project_id, f"%{invoice_no}%"))
                    tr = cur.fetchone()
                    tax_no = tr[0] if tr else invoice_no  # fallback

                invoice_dict["tax_invoice_no"] = tax_no
                project_details["contractor"]  = contractor

                _fill_proforma_sheet(ws_proforma, invoice_dict, project_details, items)
                _fill_tax_sheet(ws_tax, invoice_dict, project_details, items)

                filename = f"{_cl(contractor)}.xlsx"
                out_path = os.path.join(output_dir, f"{inv_no_h}-regen-both.xlsx")

            elif generation_mode == "PROFORMA_ONLY":
                _fill_proforma_sheet(ws_proforma, invoice_dict, project_details, items)
                _nonwi_blank_tax_sheet(ws_tax)

                filename = f"{inv_no_h}_{_cl(contractor)}.xlsx"
                out_path = os.path.join(output_dir, f"{inv_no_h}-regen-proforma.xlsx")

            elif generation_mode == "TAX_ONLY":
                # This is a TAX invoice. Find the linked proforma.
                import re as _re4
                proforma_no = None
                if remarks:
                    m = _re4.search(r'Proforma Invoice No[:\s]+(\S+)', remarks)
                    if m:
                        proforma_no = m.group(1).strip()

                if proforma_no:
                    cur.execute("""
                        SELECT i.invoice_id, i.invoice_no, i.invoice_date,
                               i.subtotal, i.vat, i.total, i.amount_in_words
                        FROM invoices i
                        WHERE i.project_id = %s AND i.invoice_no = %s
                          AND i.invoice_type = 'PROFORMA'
                        LIMIT 1
                    """, (project_id, proforma_no))
                    pfm = cur.fetchone()
                else:
                    pfm = None

                if pfm:
                    pfm_inv_id   = pfm[0]
                    pfm_subtotal = float(pfm[3]) if pfm[3] else subtotal
                    pfm_vat      = float(pfm[4]) if pfm[4] else vat
                    pfm_total    = float(pfm[5]) if pfm[5] else total
                    pfm_words    = pfm[6] or amount_in_words

                    # Load proforma items + attach report info
                    cur.execute("""
                        SELECT ii.description, ii.test_standard, ii.unit_rate,
                               ii.quantity, ii.amount, ii.sample_id
                        FROM invoice_items ii
                        WHERE ii.invoice_id = %s ORDER BY ii.item_id
                    """, (pfm_inv_id,))
                    pfm_items = [
                        {"description": r[0] or "—", "test_standard": r[1] or "",
                         "unit_rate": float(r[2] or 0), "quantity": r[3] or 1,
                         "amount": float(r[4] or 0), "sample_id": r[5]}
                        for r in cur.fetchall()
                    ]
                    pfm_sample_ids = [it["sample_id"] for it in pfm_items if it.get("sample_id")]
                    pfm_report_info = get_report_info_for_samples(pfm_sample_ids, cur)
                    for it in pfm_items:
                        info = pfm_report_info.get(it.get("sample_id"))
                        it["report_no"]         = info["report_no"]  if info else None
                        it["report_created_at"] = info["created_at"] if info else None

                    proforma_dict = {
                        **invoice_dict,
                        "invoice_no":   pfm[1],
                        "invoice_date": pfm[2],
                        "subtotal":     pfm_subtotal,
                        "vat":          pfm_vat,
                        "total":        pfm_total,
                        "amount_in_words": pfm_words,
                    }
                    # Sheet 1 = original proforma
                    _fill_proforma_sheet(ws_proforma, proforma_dict, project_details, pfm_items)

                    # Sheet 2 = this tax invoice
                    invoice_dict["tax_invoice_no"] = invoice_no
                    project_details["contractor"]  = contractor
                    _fill_tax_sheet(ws_tax, invoice_dict, project_details, items)
                else:
                    # Proforma not found; best effort
                    invoice_dict["tax_invoice_no"] = invoice_no
                    project_details["contractor"]  = contractor
                    _fill_proforma_sheet(ws_proforma, invoice_dict, project_details, items)
                    _fill_tax_sheet(ws_tax, invoice_dict, project_details, items)

                filename = f"{inv_no_h}_{_cl(contractor)}.xlsx"
                out_path = os.path.join(output_dir, f"{inv_no_h}-regen-tax.xlsx")

            else:
                raise HTTPException(400, f"Unrecognised generation_mode: {generation_mode}")

        # ── 4. Save & stream ─────────────────────────────────────────────────
        wb.save(out_path)
        enc = _urlparse.quote(filename)
        return FileResponse(
            out_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition":
                    f"attachment; filename*=UTF-8''{enc}; filename=\"{filename}\""
            }
        )

    except HTTPException:
        raise
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, f"Error regenerating invoice: {str(e)}")
    finally:
        cur.close()
        conn.close()






# ─────────────────────────────────────────────────────────────────────────────
# ADD THIS ENDPOINT to invoices.py, right after the /projects/latest/ endpoint
# (around line 961 in the original file).
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/projects/{project_id}/reports-invoiced-status")
def get_project_reports_invoiced_status(project_id: int):
    """
    Returns all reports for a project (Approved + Pending), each with an
    is_invoiced flag.  Walk-in projects are detected early and a dedicated
    response is returned so the frontend can display the correct notice.

    Used by ViewReports > "Reports by LP" section.
    """
    conn = get_connection()
    cur  = conn.cursor()

    try:
        # ── 1. Fetch project basics ───────────────────────────────────────────
        cur.execute("""
            SELECT p.project_id, p.project_no, p.project_name,
                   p.is_walk_in,
                   COALESCE(c.name, p.client_name, p.walk_in_client) AS client_name,
                   p.location
            FROM   projects p
            LEFT JOIN clients c ON p.client_id = c.client_id
            WHERE  p.project_id = %s
        """, (project_id,))

        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Project not found")

        (proj_id, project_no, project_name,
         is_walk_in, client_name, location) = row

        # ── 2. Walk-in early-return ───────────────────────────────────────────
        if is_walk_in:
            return {
                "project_id":   proj_id,
                "project_no":   project_no,
                "project_name": project_name,
                "client_name":  client_name,
                "location":     location,
                "is_walk_in":   True,
                "reports":      [],
                "total_reports": 0,
            }

        # ── 3. All reports (Approved + Pending) ──────────────────────────────
        cur.execute("""
            SELECT DISTINCT ON (r.report_no)
                r.report_id,
                r.report_no,
                r.status,
                r.created_at,
                r.covers_test_type   AS test_name,
                r.covers_samples,
                COALESCE(array_length(r.covers_samples, 1), 1) AS sample_count,
                EXISTS (
                    SELECT 1
                    FROM   invoice_report_links irl
                    WHERE  irl.report_no   = r.report_no
                    AND    irl.invoice_type = 'PROFORMA'
                ) AS is_invoiced
            FROM   reports r
            JOIN   samples      s  ON r.sample_id   = s.sample_id
            JOIN   test_requests tr ON s.request_id  = tr.test_request_id
            WHERE  tr.project_id = %s
            AND    r.status IN ('APPROVED', 'UNDER_REVIEW', 'DRAFT')
            ORDER  BY r.report_no, r.created_at DESC
        """, (project_id,))

        rows = cur.fetchall()

        reports = []
        for row in rows:
            (report_id, report_no, status, created_at,
             test_name, covers_samples, sample_count, is_invoiced) = row

            # Resolve sample list
            if covers_samples:
                sample_nos = list(covers_samples)
            else:
                cur.execute(
                    "SELECT sample_no FROM samples WHERE sample_id = "
                    "(SELECT sample_id FROM reports WHERE report_id = %s LIMIT 1)",
                    (report_id,)
                )
                srow = cur.fetchone()
                sample_nos   = [srow[0]] if srow else []
                sample_count = len(sample_nos)

            reports.append({
                "report_id":     report_id,
                "report_no":     report_no,
                "status":        status,
                "created_date":  created_at.strftime("%Y-%m-%d") if created_at else None,
                "test_name":     test_name or "Test Report",
                "sample_count":  sample_count,
                "covers_samples": sample_nos,
                "is_invoiced":   bool(is_invoiced),
            })

        return {
            "project_id":    proj_id,
            "project_no":    project_no,
            "project_name":  project_name,
            "client_name":   client_name,
            "location":      location,
            "is_walk_in":    False,
            "total_reports": len(reports),
            "reports":       reports,
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Failed to fetch project reports: {str(e)}")
    finally:
        cur.close()
        conn.close()


# ============================================================================
# "Test Wise Department" — department revenue breakdown for View Invoices
#
# Revenue is attributed to a department via:
#     invoice_items.sample_id  ->  samples.department
# This only covers non-walk-in invoices, since walk-in items are not tied to
# a lab sample (walk_in_items has no department column). Walk-in revenue is
# therefore intentionally excluded from this view.
#
# PRF / Tax number pairing is read off invoices.remarks, using the same
# cross-reference convention already used elsewhere in this file:
#   - a PROFORMA row's remarks contains "Tax Invoice No: <no>" once a tax
#     invoice has actually been created against it (BOTH mode, or a later
#     PROFORMA_ONLY -> TAX_ONLY conversion)
#   - a TAX row's remarks contains "Proforma Invoice No: <no>" when it was
#     converted from an existing proforma
# If no such reference exists, the corresponding number is left blank —
# i.e. "if tax is not made, do not show it".
# ============================================================================

import re as _dept_re


def _prf_tax_numbers(invoice_type: str, remarks: str, invoice_no: str):
    """Given one invoice row, return (prf_number, tax_number) using the
    remarks cross-reference convention described above."""
    prf_number = None
    tax_number = None
    if invoice_type == "PROFORMA":
        prf_number = invoice_no
        if remarks:
            m = _dept_re.search(r"Tax Invoice No[:\s]+(\S+)", remarks)
            if m:
                tax_number = m.group(1).strip()
    elif invoice_type == "TAX":
        tax_number = invoice_no
        if remarks:
            m = _dept_re.search(r"Proforma Invoice No[:\s]+(\S+)", remarks)
            if m:
                prf_number = m.group(1).strip()
    return prf_number, tax_number


@router.get("/departments/summary")
def get_department_revenue_summary():
    """
    Returns every department that has invoiced tests, with the total cost
    and test count across all (non-walk-in) invoices.

    [{ "department": "GeoTech", "test_count": 3, "total_cost": 300.0 }, ...]
    """
    conn = get_connection()
    cur  = conn.cursor()
    try:
        cur.execute("""
            SELECT s.department,
                   COUNT(*)                    AS test_count,
                   COALESCE(SUM(ii.amount), 0) AS total_cost
            FROM invoice_items ii
            JOIN samples  s ON ii.sample_id = s.sample_id
            JOIN invoices i ON ii.invoice_id = i.invoice_id
            WHERE s.department IS NOT NULL
              AND s.department <> ''
              AND i.generation_mode IS NOT NULL
            GROUP BY s.department
            ORDER BY total_cost DESC
        """)
        rows = cur.fetchall()
        return [
            {
                "department": dept,
                "test_count": int(count),
                "total_cost": float(total) if total is not None else 0.0,
            }
            for dept, count, total in rows
        ]
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))
    finally:
        cur.close(); conn.close()


@router.get("/departments/breakdown")
def get_department_revenue_breakdown(department: str):
    """
    Month-wise drill-down for a single department: for each month, the list
    of tests (test name, invoice date, PRF no., Tax no. if made, cost) plus
    a month total, newest month first.
    """
    conn = get_connection()
    cur  = conn.cursor()
    try:
        cur.execute("""
            SELECT ii.description,
                   ii.amount,
                   i.invoice_no,
                   i.invoice_type,
                   i.invoice_date,
                   i.remarks
            FROM invoice_items ii
            JOIN samples  s ON ii.sample_id = s.sample_id
            JOIN invoices i ON ii.invoice_id = i.invoice_id
            WHERE s.department = %s
              AND i.generation_mode IS NOT NULL
            ORDER BY i.invoice_date DESC NULLS LAST, ii.item_id DESC
        """, (department,))
        rows = cur.fetchall()

        months = []          # ordered list of month buckets
        by_key = {}           # month_key -> bucket dict
        total_cost = 0.0

        for description, amount, invoice_no, invoice_type, invoice_date, remarks in rows:
            cost = float(amount) if amount is not None else 0.0
            total_cost += cost

            prf_number, tax_number = _prf_tax_numbers(invoice_type, remarks, invoice_no)

            if invoice_date:
                month_key   = invoice_date.strftime("%Y-%m")
                month_label = invoice_date.strftime("%B %Y")
                date_str    = invoice_date.strftime("%d-%b-%Y")
            else:
                month_key = month_label = "Undated"
                date_str  = None

            if month_key not in by_key:
                bucket = {"month_key": month_key, "month_label": month_label,
                          "total_cost": 0.0, "tests": []}
                by_key[month_key] = bucket
                months.append(bucket)

            bucket = by_key[month_key]
            bucket["total_cost"] += cost
            bucket["tests"].append({
                "test_name":    description or "—",
                "invoice_date": date_str,
                "prf_number":   prf_number,
                "tax_number":   tax_number,
                "cost":         round(cost, 2),
            })

        for bucket in months:
            bucket["total_cost"] = round(bucket["total_cost"], 2)

        return {
            "department": department,
            "total_cost": round(total_cost, 2),
            "test_count": len(rows),
            "months":     months,
        }
    except Exception as e:
        traceback.print_exc()
        raise HTTPException(500, str(e))
    finally:
        cur.close(); conn.close()