# samples_workflow.py - FIXED VERSION WITH CONSISTENT TEST ASSIGNMENT with excel template
# Each sample gets ONE test at creation and keeps it forever

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime
from db import get_connection
from utils import resource_path

from fastapi import UploadFile, File
import shutil
import os
import sys
import uuid
import tempfile 
import secrets
from decimal import Decimal
from fastapi.responses import FileResponse
from supabase import create_client, Client

import openpyxl
from openpyxl import load_workbook

import requests
import tempfile

router = APIRouter(prefix="/samples-workflow", tags=["Samples Workflow"])

if hasattr(sys, "_MEIPASS"):
    WORKSHEET_TEMPLATES_DIR = os.path.join(tempfile.gettempdir(), "lab_app_worksheets")
else:
    WORKSHEET_TEMPLATES_DIR = resource_path("templates/worksheets")
os.makedirs(WORKSHEET_TEMPLATES_DIR, exist_ok=True)


SUPABASE_URL = "https://hqwgkmbjmcxpxbwccclo.supabase.co"
SUPABASE_KEY = "sb_secret_-8uQCdQSiUgDFO_MUEsTWg_TPWtsyy3"

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)


def download_worksheet_template_from_supabase(item_code: str):
    try:
        template_urls = [
            f"https://hqwgkmbjmcxpxbwccclo.supabase.co/storage/v1/object/public/templates/worksheets/{item_code}.xlsx",
            f"https://hqwgkmbjmcxpxbwccclo.supabase.co/storage/v1/object/public/templates/worksheets/{item_code}_Worksheet.xlsx",
            f"https://hqwgkmbjmcxpxbwccclo.supabase.co/storage/v1/object/public/templates/worksheets/{item_code}.xls"
        ]
        
        template_path = None
        
        for url in template_urls:
            try:
                response = requests.get(url, timeout=30)
                if response.status_code == 200:
                    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                        temp_file.write(response.content)
                        template_path = temp_file.name
                    break
            except requests.exceptions.RequestException:
                continue
        
        if not template_path:
            generic_urls = [
                "https://hqwgkmbjmcxpxbwccclo.supabase.co/storage/v1/object/public/templates/worksheets/DEFAULT_Worksheet.xlsx",
                "https://hqwgkmbjmcxpxbwccclo.supabase.co/storage/v1/object/public/templates/worksheets/GENERIC_Worksheet.xlsx"
            ]
            for url in generic_urls:
                try:
                    response = requests.get(url, timeout=30)
                    if response.status_code == 200:
                        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                            temp_file.write(response.content)
                            template_path = temp_file.name
                        break
                except requests.exceptions.RequestException:
                    continue
        
        if not template_path:
            raise HTTPException(status_code=404, detail=f"No worksheet template found for {item_code} in Supabase storage")
        
        return template_path
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to download worksheet template: {str(e)}")


def generate_sample_no(cur, request_id: int, sequence_num: int):
    cur.execute("""
        SELECT request_no, created_at 
        FROM test_requests 
        WHERE test_request_id = %s
    """, (request_id,))
    row = cur.fetchone()
    
    if not row:
        return f"GS-{datetime.now().strftime('%d%m%y')}-REQ{request_id:04d}-{sequence_num:02d}"
    
    request_no, created_at = row
    date_part = ""
    
    if len(request_no) >= 9 and '-' in request_no:
        parts = request_no.split('-')
        if len(parts) >= 2:
            date_part = parts[1]
        else:
            date_part = created_at.strftime("%d%m%y") if created_at else datetime.now().strftime("%d%m%y")
    else:
        date_part = created_at.strftime("%d%m%y") if created_at else datetime.now().strftime("%d%m%y")
    
    request_seq = "01"
    if len(request_no) >= 12 and request_no.count('-') >= 2:
        try:
            request_seq = request_no.split('-')[2]
        except (IndexError, AttributeError):
            request_seq = "01"
    
    return f"GS-{date_part}-{request_seq}-{sequence_num}"


def generate_worksheet_no(cur, sample_id: int):
    year = datetime.utcnow().year
    cur.execute("""
        SELECT COUNT(*) 
        FROM worksheets 
        WHERE EXTRACT(YEAR FROM created_at) = %s
    """, (year,))
    seq = cur.fetchone()[0] + 1
    return f"WKS-{year}-{sample_id:04d}-{seq:03d}"


def generate_barcode():
    return secrets.token_hex(8).upper()


class GenerateSamplesIn(BaseModel):
    collected_by: Optional[str] = None

class AcceptSampleIn(BaseModel):
    storage_location: Optional[str] = None
    note: Optional[str] = None

class RejectSampleIn(BaseModel):
    reason: Optional[str] = None
    inform_client: Optional[bool] = False

class GenerateWorksheetIn(BaseModel):
    technician: Optional[str] = None


def assign_tests_to_samples(cur, test_request_id: int):
    cur.execute("""
        SELECT tri.tri_id, tri.quotation_item_id, tri.quantity,
               qi.item_code, qi.description, qi.test_standard, qi.unit_rate
        FROM test_request_items tri
        JOIN quotation_items qi ON tri.quotation_item_id = qi.item_id
        WHERE tri.test_request_id = %s
        ORDER BY tri.tri_id
    """, (test_request_id,))
    
    tests = cur.fetchall()
    test_distribution = []
    sample_counter = 1
    
    for tri_id, quotation_item_id, quantity, item_code, description, test_standard, unit_rate in tests:
        for _ in range(quantity):
            test_distribution.append({
                "sample_sequence": sample_counter,
                "tri_id": tri_id,
                "quotation_item_id": quotation_item_id,
                "item_code": item_code,
                "description": description,
                "test_standard": test_standard,
                "unit_rate": unit_rate
            })
            sample_counter += 1
    
    return test_distribution


# ---------------------------
# 1) Generate Samples
# ---------------------------
@router.post("/generate-samples-by-request-no/{request_no}")
def generate_samples_by_request_no(request_no: str, payload: GenerateSamplesIn):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT test_request_id, project_id 
            FROM test_requests 
            WHERE request_no = %s
        """, (request_no,))
        req = cur.fetchone()
        if not req:
            raise HTTPException(404, f"Test request with number '{request_no}' not found")

        test_request_id, project_id = req
        test_distribution = assign_tests_to_samples(cur, test_request_id)
        
        if not test_distribution:
            raise HTTPException(400, "This request has no items")

        created_samples = []
        test_assignments = []

        for test_info in test_distribution:
            sample_sequence = test_info["sample_sequence"]
            sample_no = generate_sample_no(cur, test_request_id, sample_sequence)

            cur.execute("""
                INSERT INTO samples (
                    sample_no, request_id, collected_by, received_date, status,
                    assigned_tri_id, assigned_quotation_item_id
                )
                VALUES (%s, %s, %s, NULL, 'PENDING', %s, %s)
                RETURNING sample_id
            """, (
                sample_no, test_request_id, payload.collected_by,
                test_info["tri_id"], test_info["quotation_item_id"]
            ))

            new_sample_id = cur.fetchone()[0]
            created_samples.append(new_sample_id)
            test_assignments.append({
                "sample_id": new_sample_id,
                "sample_no": sample_no,
                "assigned_test": test_info["item_code"],
                "test_name": test_info["description"],
                "tri_id": test_info["tri_id"],
                "quotation_item_id": test_info["quotation_item_id"],
                "sequence": sample_sequence
            })

        conn.commit()

        return {
            "message": f"Samples generated for request {request_no}",
            "count": len(created_samples),
            "test_request_id": test_request_id,
            "request_no": request_no,
            "sample_ids": created_samples,
            "test_distribution": test_assignments,
            "note": "Each sample has a permanently assigned test."
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, str(e))
    finally:
        cur.close()
        conn.close()


# ---------------------------
# 2) Accept sample
# ---------------------------
@router.post("/samples/{sample_id}/accept")
def accept_sample(sample_id: int, payload: AcceptSampleIn):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT s.sample_id, s.sample_no, s.request_id, 
                   s.assigned_tri_id, s.assigned_quotation_item_id,
                   qi.item_code, qi.description
            FROM samples s
            LEFT JOIN quotation_items qi ON s.assigned_quotation_item_id = qi.item_id
            WHERE s.sample_id = %s
        """, (sample_id,))
        row = cur.fetchone()

        if not row:
            raise HTTPException(404, "Sample not found")

        sample_id, existing_sample_no, request_id, assigned_tri_id, assigned_quotation_item_id, item_code, test_name = row
        sample_no = existing_sample_no or generate_sample_no(cur, request_id, 1)
        barcode = generate_barcode()

        cur.execute("""
            UPDATE samples
            SET sample_no = %s,
                barcode = %s,
                received_date = NOW(),
                status = 'ACCEPTED',
                storage_location = COALESCE(%s, storage_location)
            WHERE sample_id = %s
            RETURNING sample_no, barcode
        """, (sample_no, barcode, payload.storage_location, sample_id))

        updated = cur.fetchone()
        conn.commit()

        return {
            "message": "Sample accepted",
            "sample_id": sample_id,
            "sample_no": updated[0],
            "barcode": updated[1],
            "assigned_test": item_code,
            "test_name": test_name,
        }

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, str(e))
    finally:
        cur.close()
        conn.close()


# ---------------------------
# 3) Reject sample
# ---------------------------
@router.post("/samples/{sample_id}/reject")
def reject_sample(sample_id: int, payload: RejectSampleIn):
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("SELECT sample_id FROM samples WHERE sample_id = %s", (sample_id,))
        if cur.fetchone() is None:
            raise HTTPException(404, "Sample not found")

        cur.execute("""
            UPDATE samples
            SET status = 'REJECTED',
                reason_rejected = %s,
                received_date = NOW()
            WHERE sample_id = %s
        """, (payload.reason, sample_id))

        conn.commit()
        return {"message": "Sample rejected"}

    except Exception as e:
        conn.rollback()
        raise HTTPException(500, str(e))
    finally:
        cur.close()
        conn.close()


# ---------------------------
# 4) Generate Worksheet
# ---------------------------
@router.post("/samples/{sample_id}/generate-worksheet")
def generate_worksheet(sample_id: int, payload: GenerateWorksheetIn):
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT s.sample_id, s.sample_no, s.request_id,
                   s.assigned_tri_id, s.assigned_quotation_item_id,
                   qi.item_code, qi.description, qi.test_standard, qi.unit_rate
            FROM samples s
            LEFT JOIN quotation_items qi ON s.assigned_quotation_item_id = qi.item_id
            WHERE s.sample_id = %s
        """, (sample_id,))
        
        sample_row = cur.fetchone()
        if not sample_row:
            raise HTTPException(404, f"Sample {sample_id} not found")
        
        sample_id_db, sample_no, request_id, assigned_tri_id, assigned_quotation_item_id, item_code, description, test_standard, unit_rate = sample_row
        
        if not assigned_quotation_item_id:
            raise HTTPException(400, f"Sample {sample_id} has no assigned test. Please regenerate samples.")
        
        cur.execute("""
            SELECT worksheet_id, worksheet_no, status, created_at
            FROM worksheets 
            WHERE sample_id = %s AND quotation_item_id = %s
        """, (sample_id, assigned_quotation_item_id))
        
        existing = cur.fetchone()
        if existing:
            existing_id, existing_no, existing_status, existing_created = existing
            return {
                "message": f"Worksheet {existing_no} already exists for this sample/test combination.",
                "existing_worksheet": {
                    "worksheet_id": existing_id,
                    "worksheet_no": existing_no,
                    "status": existing_status,
                    "created_at": existing_created,
                    "download_url": f"/samples-workflow/worksheets/{existing_id}/download"
                },
                "download_available": True,
            }
        
        year = datetime.utcnow().year
        cur.execute("""
            SELECT COUNT(*) FROM worksheets WHERE EXTRACT(YEAR FROM created_at) = %s
        """, (year,))
        seq = cur.fetchone()[0] + 1
        worksheet_no = f"WKS-{year}-{sample_id:04d}-{seq:03d}"
        
        template_available = False
        template_path = None
        
        if item_code:
            try:
                template_path = download_worksheet_template_from_supabase(item_code)
                template_available = True
            except Exception:
                template_available = False
        
        cur.execute("""
            INSERT INTO worksheets (
                worksheet_no, sample_id, quotation_item_id, test_name,
                standard, unit_rate, quantity, technician, status, created_at
            )
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'GENERATED', NOW())
            RETURNING worksheet_id
        """, (
            worksheet_no, sample_id, assigned_quotation_item_id, description,
            test_standard, float(unit_rate) if isinstance(unit_rate, Decimal) else unit_rate,
            1, payload.technician or "Lab Technician"
        ))
        
        worksheet_id_new = cur.fetchone()[0]
        conn.commit()
        
        return {
            "message": f"Worksheet generated for {description}",
            "worksheet_id": worksheet_id_new,
            "worksheet_no": worksheet_no,
            "sample_id": sample_id,
            "sample_no": sample_no,
            "test_name": description,
            "item_code": item_code,
            "test_standard": test_standard,
            "status": "GENERATED",
            "template_available": template_available,
        }
        
    except HTTPException:
        conn.rollback()
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"Error generating worksheet: {str(e)}")
    finally:
        cur.close()
        conn.close()


# ---------------------------
# 5) Get pending samples — THE ONLY DEFINITION (includes picture fields)
# ---------------------------
@router.get("/pending-samples")
def get_pending_samples():
    """Get all PENDING samples — includes picture_path and picture_uploaded_at"""
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("""
            SELECT
                s.sample_id,
                s.sample_no,
                s.request_id,
                s.collected_by,
                s.received_date,
                s.status,
                s.reason_rejected,
                s.barcode,
                s.storage_location,
                tr.request_no,
                s.assigned_quotation_item_id,
                qi.item_code,
                qi.description,
                s.picture_path,
                s.picture_uploaded_at
            FROM samples s
            JOIN test_requests tr ON s.request_id = tr.test_request_id
            LEFT JOIN quotation_items qi ON s.assigned_quotation_item_id = qi.item_id
            WHERE s.status = 'PENDING'
            ORDER BY s.sample_id DESC
        """)

        samples = cur.fetchall()

        result = []
        for sample in samples:
            (sample_id, sample_no, request_id, collected_by, received_date, status,
             reason_rejected, barcode, storage_location, request_no,
             assigned_quotation_item_id, item_code, description,
             picture_path, picture_uploaded_at) = sample

            result.append({
                "sample_id": sample_id,
                "sample_no": sample_no,
                "request_id": request_id,
                "collected_by": collected_by,
                "received_date": received_date,
                "status": status,
                "reason_rejected": reason_rejected,
                "barcode": barcode,
                "storage_location": storage_location,
                "request_no": request_no,
                "assigned_test": item_code or "Not Assigned",
                "test_name": description or "No test name",
                "assigned_from_storage": assigned_quotation_item_id is not None,
                "picture_path": picture_path,
                "picture_uploaded_at": picture_uploaded_at.isoformat() if picture_uploaded_at else None,
            })

        return result

    except Exception as e:
        raise HTTPException(500, f"Database error: {str(e)}")
    finally:
        cur.close()
        conn.close()


# ---------------------------
# 6) Debug endpoint
# ---------------------------
@router.get("/debug/worksheet/{sample_id}")
def debug_worksheet_data(sample_id: int):
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT s.sample_id, s.sample_no, s.request_id,
                   s.assigned_tri_id, s.assigned_quotation_item_id,
                   tr.test_request_id, tr.request_no,
                   qi.item_code, qi.description, qi.test_standard, qi.unit_rate
            FROM samples s
            JOIN test_requests tr ON s.request_id = tr.test_request_id
            LEFT JOIN quotation_items qi ON s.assigned_quotation_item_id = qi.item_id
            WHERE s.sample_id = %s
        """, (sample_id,))
        
        sample_info = cur.fetchone()
        if not sample_info:
            raise HTTPException(404, "Sample not found")
        
        (sample_id_db, sample_no, request_id, assigned_tri_id, assigned_quotation_item_id,
         test_request_id, request_no, item_code, description, test_standard, unit_rate) = sample_info
        
        cur.execute("""
            SELECT qi.item_id, qi.item_code, qi.description, qi.test_standard, qi.unit_rate,
                   tri.quantity as requested_quantity, tri.tri_id
            FROM quotation_items qi
            JOIN test_request_items tri ON qi.item_id = tri.quotation_item_id
            WHERE tri.test_request_id = %s
            ORDER BY tri.tri_id
        """, (test_request_id,))
        all_tests = cur.fetchall()
        
        cur.execute("""
            SELECT w.worksheet_id, w.worksheet_no, w.test_name, w.standard, w.created_at, qi.item_code
            FROM worksheets w
            LEFT JOIN quotation_items qi ON w.quotation_item_id = qi.item_id
            WHERE w.sample_id = %s
        """, (sample_id,))
        existing_worksheets = cur.fetchall()
        
        return {
            "sample_info": {
                "sample_id": sample_id_db,
                "sample_no": sample_no,
                "stored_assignment": {
                    "assigned_tri_id": assigned_tri_id,
                    "assigned_quotation_item_id": assigned_quotation_item_id,
                    "item_code": item_code,
                    "description": description,
                }
            },
            "all_tests_in_request": [
                {"item_id": i[0], "item_code": i[1], "description": i[2], "tri_id": i[6]}
                for i in all_tests
            ],
            "existing_worksheets": [
                {"worksheet_id": ws[0], "worksheet_no": ws[1], "test_name": ws[2], "item_code": ws[5]}
                for ws in existing_worksheets
            ]
        }
        
    except Exception as e:
        raise HTTPException(500, str(e))
    finally:
        cur.close()
        conn.close()


# ---------------------------
# Upload / Download worksheet file
# ---------------------------
@router.post("/worksheets/{worksheet_id}/upload")
async def upload_worksheet_file(worksheet_id: int, worksheet_file: UploadFile = File(...)):
    conn = get_connection()
    cur = conn.cursor()
    TARGET_BUCKET = "lab_worksheets"
    
    try:
        cur.execute("""
            SELECT w.worksheet_no, qi.item_code, w.test_name
            FROM worksheets w
            LEFT JOIN quotation_items qi ON w.quotation_item_id = qi.item_id
            WHERE w.worksheet_id = %s
        """, (worksheet_id,))
        worksheet_info = cur.fetchone()
        if not worksheet_info:
            raise HTTPException(404, f"Worksheet {worksheet_id} not found")
        
        worksheet_no, item_code, test_name = worksheet_info
        file_content = await worksheet_file.read()
        file_ext = os.path.splitext(worksheet_file.filename)[1].lower()
        safe_code = (item_code or 'worksheet').replace('/', '_').replace('\\', '_').replace(' ', '_')
        filename = f"WS_{worksheet_no}_{safe_code}_{worksheet_id}{file_ext}"
        
        try:
            supabase.storage.from_(TARGET_BUCKET).remove([filename])
        except:
            pass
        
        supabase.storage.from_(TARGET_BUCKET).upload(
            path=filename,
            file=file_content,
            file_options={"content-type": worksheet_file.content_type or "application/octet-stream", "x-upsert": "true"}
        )
        
        public_url = f"https://hqwgkmbjmcxpxbwccclo.supabase.co/storage/v1/object/public/{TARGET_BUCKET}/{filename}"
        
        cur.execute("""
            UPDATE worksheets SET template_path = %s, updated_at = NOW(), status = 'UPLOADED'
            WHERE worksheet_id = %s
        """, (public_url, worksheet_id))
        conn.commit()
        
        return {"status": "success", "url": public_url, "worksheet_id": worksheet_id, "worksheet_no": worksheet_no}
        
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, f"Upload failed: {str(e)}")
    finally:
        cur.close()
        conn.close()


@router.get("/worksheets/{worksheet_id}/download")
def download_worksheet(worksheet_id: int):
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT w.worksheet_no, w.template_path, w.test_name, s.sample_no, w.status, qi.item_code
            FROM worksheets w
            JOIN samples s ON w.sample_id = s.sample_id
            LEFT JOIN quotation_items qi ON w.quotation_item_id = qi.item_id
            WHERE w.worksheet_id = %s
        """, (worksheet_id,))
        worksheet = cur.fetchone()
        if not worksheet:
            raise HTTPException(404, "Worksheet not found")
        
        worksheet_no, template_path, test_name, sample_no, status, item_code = worksheet
        
        if template_path and template_path.startswith('http'):
            filename = template_path.split('/')[-1]
            correct_url = f"https://hqwgkmbjmcxpxbwccclo.supabase.co/storage/v1/object/public/lab_worksheets/{filename}"
            return {
                "has_file": True,
                "download_url": correct_url,
                "worksheet_id": worksheet_id,
                "worksheet_no": worksheet_no,
                "sample_no": sample_no,
                "test_name": test_name,
                "item_code": item_code,
                "status": status,
            }
        
        return {
            "has_file": False,
            "message": "No worksheet file uploaded yet.",
            "worksheet_id": worksheet_id,
            "worksheet_no": worksheet_no,
            "sample_no": sample_no,
            "test_name": test_name,
            "status": status,
        }
        
    except Exception as e:
        raise HTTPException(500, f"Download error: {str(e)}")
    finally:
        cur.close()
        conn.close()


@router.get("/recent-samples")
def get_recent_samples(limit: int = 5):
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT s.sample_id, s.sample_no, s.status, s.barcode,
                   tr.request_no, s.request_id, qi.item_code, qi.description
            FROM samples s
            JOIN test_requests tr ON s.request_id = tr.test_request_id
            LEFT JOIN quotation_items qi ON s.assigned_quotation_item_id = qi.item_id
            WHERE s.status IN ('PENDING', 'ACCEPTED')
            ORDER BY s.sample_id DESC
            LIMIT %s
        """, (limit,))
        
        rows = cur.fetchall()
        return [
            {
                "sample_id": r[0], "sample_no": r[1], "status": r[2], "barcode": r[3],
                "request_no": r[4], "assigned_test": r[7] or "Test not assigned", "item_code": r[6] or "N/A"
            }
            for r in rows
        ]
        
    except Exception as e:
        raise HTTPException(500, f"Error fetching recent samples: {str(e)}")
    finally:
        cur.close()
        conn.close()


@router.get("/samples/{sample_id}/download-template")
def download_worksheet_template(sample_id: int):
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT s.sample_id, s.sample_no, s.assigned_quotation_item_id,
                   qi.item_code, qi.description
            FROM samples s
            LEFT JOIN quotation_items qi ON s.assigned_quotation_item_id = qi.item_id
            WHERE s.sample_id = %s
        """, (sample_id,))
        sample_row = cur.fetchone()
        if not sample_row:
            raise HTTPException(404, f"Sample {sample_id} not found")
        
        sample_id_db, sample_no, assigned_quotation_item_id, item_code, test_name = sample_row
        
        if not assigned_quotation_item_id or not item_code:
            raise HTTPException(400, f"Sample {sample_id} has no assigned test. Please regenerate samples.")

        template_path = None
        for variation in [item_code, item_code.upper(), item_code.lower()]:
            try:
                template_path = download_worksheet_template_from_supabase(variation)
                break
            except Exception:
                continue

        if not template_path:
            return {
                "has_template": False,
                "message": f"No standard template found for {test_name} ({item_code}).",
                "item_code": item_code,
                "test_name": test_name,
                "sample_no": sample_no,
            }
        
        return FileResponse(
            path=template_path,
            filename=os.path.basename(template_path),
            media_type='application/octet-stream'
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error downloading template: {str(e)}")
    finally:
        cur.close()
        conn.close()


@router.get("/all-samples")
def get_all_samples():
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT s.sample_id, s.sample_no, s.request_id, s.collected_by, s.received_date,
                   s.status, s.reason_rejected, s.barcode, s.storage_location,
                   tr.request_no, s.assigned_quotation_item_id, qi.item_code, qi.description
            FROM samples s
            JOIN test_requests tr ON s.request_id = tr.test_request_id
            LEFT JOIN quotation_items qi ON s.assigned_quotation_item_id = qi.item_id
            ORDER BY s.sample_id DESC
        """)
        samples = cur.fetchall()
        
        return [
            {
                "sample_id": s[0], "sample_no": s[1], "request_id": s[2],
                "collected_by": s[3], "received_date": s[4], "status": s[5],
                "reason_rejected": s[6], "barcode": s[7], "storage_location": s[8],
                "request_no": s[9], "assigned_test": s[11] or "Not Assigned",
                "test_name": s[12] or "No test name",
                "assigned_from_storage": s[10] is not None,
            }
            for s in samples
        ]
        
    except Exception as e:
        raise HTTPException(500, f"Database error: {str(e)}")
    finally:
        cur.close()
        conn.close()


# ---------------------------
# Upload sample picture
# ---------------------------
@router.post("/samples/{sample_id}/upload-picture", summary="Upload Sample Picture to Supabase")
async def upload_sample_picture(sample_id: int, file: UploadFile = File(...)):
    SAMPLE_PICS_BUCKET = "sample_pics"
    conn = get_connection()
    cur = conn.cursor()

    try:
        cur.execute("SELECT sample_id, sample_no FROM samples WHERE sample_id = %s", (sample_id,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Sample not found")

        sample_no = row[1]
        contents = await file.read()

        if len(contents) > 20 * 1024 * 1024:
            raise HTTPException(400, "File too large. Maximum size is 20MB.")

        original_ext = os.path.splitext(file.filename)[1].lower() if file.filename else ".bin"
        allowed_exts = {".jpg", ".jpeg", ".png", ".webp"}
        if original_ext not in allowed_exts:
            raise HTTPException(400, f"File type not allowed. Allowed: {', '.join(allowed_exts)}")

        unique_name = f"{sample_no}_{uuid.uuid4().hex[:8]}{original_ext}"
        storage_path = f"samples/{unique_name}"

        try:
            supabase.storage.from_(SAMPLE_PICS_BUCKET).remove([storage_path])
        except:
            pass

        supabase.storage.from_(SAMPLE_PICS_BUCKET).upload(
            path=storage_path,
            file=contents,
            file_options={"content-type": file.content_type or "image/jpeg", "x-upsert": "true"}
        )

        uploaded_at = datetime.utcnow()
        cur.execute(
            "UPDATE samples SET picture_path = %s, picture_uploaded_at = %s WHERE sample_id = %s",
            (storage_path, uploaded_at, sample_id)
        )
        conn.commit()

        public_url = f"{SUPABASE_URL}/storage/v1/object/public/{SAMPLE_PICS_BUCKET}/{storage_path}"

        return {
            "message": "Sample picture uploaded successfully",
            "sample_id": sample_id,
            "picture_path": storage_path,
            "picture_url": public_url,
            "picture_uploaded_at": uploaded_at.isoformat(),
        }

    except HTTPException:
        raise
    except Exception as e:
        conn.rollback()
        raise HTTPException(500, str(e))
    finally:
        cur.close()
        conn.close()


# ---------------------------
# Worksheet population helpers
# ---------------------------
def populate_worksheet_template(template_path: str, worksheet_id: int, output_path: str):
    conn = get_connection()
    cur = conn.cursor()
    
    try:
        cur.execute("""
            SELECT w.sample_id, s.sample_no, s.request_id
            FROM worksheets w
            JOIN samples s ON w.sample_id = s.sample_id
            WHERE w.worksheet_id = %s
        """, (worksheet_id,))
        worksheet_data = cur.fetchone()
        if not worksheet_data:
            raise HTTPException(404, f"Worksheet {worksheet_id} not found")
        
        sample_id, sample_no, request_id = worksheet_data
        
        cur.execute("""
            SELECT s.assigned_quotation_item_id, qi.item_code, qi.description
            FROM samples s
            LEFT JOIN quotation_items qi ON s.assigned_quotation_item_id = qi.item_id
            WHERE s.sample_id = %s
        """, (sample_id,))
        sample_info = cur.fetchone()
        if not sample_info:
            raise HTTPException(404, f"Sample {sample_id} not found")
        
        assigned_quotation_item_id, assigned_item_code, assigned_test_name = sample_info
        
        cur.execute("""
            SELECT s.sample_id, s.sample_no, qi.item_code, qi.description,
                   ROW_NUMBER() OVER (ORDER BY s.sample_id) as sequence_num
            FROM samples s
            LEFT JOIN quotation_items qi ON s.assigned_quotation_item_id = qi.item_id
            WHERE s.request_id = %s AND s.assigned_quotation_item_id = %s
            ORDER BY s.sample_id
        """, (request_id, assigned_quotation_item_id))
        test_samples = cur.fetchall()
        
        cur.execute("""
            SELECT s.sample_no, s.collected_by,
                   TO_CHAR(s.received_date, 'DD-MM-YYYY'),
                   tr.request_no, p.project_no,
                   w.worksheet_no, w.technician
            FROM samples s
            JOIN test_requests tr ON s.request_id = tr.test_request_id
            JOIN projects p ON tr.project_id = p.project_id
            LEFT JOIN worksheets w ON w.sample_id = s.sample_id
                AND w.quotation_item_id = s.assigned_quotation_item_id
            WHERE s.sample_id = %s
        """, (sample_id,))
        meta = cur.fetchone()
        
        workbook = load_workbook(template_path)
        sheet = workbook.active
        
        if meta:
            sheet['D7'] = meta[3]  # request_no
            sheet['D8'] = meta[4]  # project_no
            sheet['E39'] = meta[1]  # collected_by
            sheet['J9'] = meta[2]  # received_date
        
        start_col = 6  # Column F
        for idx, sample in enumerate(test_samples):
            col_letter = openpyxl.utils.get_column_letter(start_col + idx)
            sheet[f'{col_letter}14'] = sample[1]  # sample_no
            sheet[f'{col_letter}15'] = sample[4]  # sequence
        
        workbook.save(output_path)
        return {"output_path": output_path, "sample_count": len(test_samples)}
    
    except Exception as e:
        raise HTTPException(500, f"Error populating worksheet: {str(e)}")
    finally:
        cur.close()
        conn.close()


@router.get("/worksheets/{worksheet_id}/download-filled-worksheet")
def download_filled_worksheet(worksheet_id: int):
    try:
        conn = get_connection()
        cur = conn.cursor()
        
        cur.execute("""
            SELECT w.worksheet_id, w.worksheet_no, w.quotation_item_id,
                   qi.item_code, s.sample_no, s.sample_id
            FROM worksheets w
            JOIN samples s ON w.sample_id = s.sample_id
            LEFT JOIN quotation_items qi ON w.quotation_item_id = qi.item_id
            WHERE w.worksheet_id = %s
            LIMIT 1
        """, (worksheet_id,))
        worksheet_info = cur.fetchone()
        
        if not worksheet_info:
            raise HTTPException(404, f"No worksheet found with ID {worksheet_id}")
        
        worksheet_id_db, worksheet_no, quotation_item_id, item_code, sample_no, sample_id = worksheet_info
        
        template_path = None
        for variation in [item_code, item_code.upper(), item_code.lower()]:
            try:
                template_path = download_worksheet_template_from_supabase(variation)
                break
            except HTTPException:
                continue
        
        if not template_path:
            raise HTTPException(404, f"No template found for {item_code}")
        
        output_filename = f"{sample_no}_{worksheet_no}_FILLED.xlsx"
        temp_dir = "temp_filled_worksheets"
        os.makedirs(temp_dir, exist_ok=True)
        output_path = os.path.join(temp_dir, output_filename)
        
        populate_worksheet_template(template_path, worksheet_id_db, output_path)
        
        return FileResponse(
            path=output_path,
            filename=output_filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Error generating filled worksheet: {str(e)}")
    finally:
        cur.close()
        conn.close()